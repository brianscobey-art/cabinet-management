from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.models import Role
from tests.conftest import login, make_user
from tests.test_jobs import make_job, setup_account


def setup_accepted_quote(client, db):
    make_user(db, email="sales@example.com", role=Role.sales)
    headers = login(client, "sales@example.com")
    account_id, community_id = setup_account(client, headers)
    job = make_job(client, headers, account_id, community_id)
    client.post(
        f"/jobs/{job['id']}/rooms", headers=headers,
        json={"room": "Kitchen", "door_style": "Shaker", "finish": "Chestnut"},
    )
    quote = client.post(f"/jobs/{job['id']}/quotes", headers=headers, json={"name": "Option A"}).json()
    client.post(
        f"/quotes/{quote['id']}/lines", headers=headers,
        json={"sku": "B18", "qty": 2, "list_price": "461.75", "product_code": "B18-SC"},
    )
    client.post(
        f"/quotes/{quote['id']}/lines", headers=headers,
        json={"sku": "RANGE1.30", "qty": 1, "list_price": "0"},  # appliance placeholder
    )
    client.post(f"/quotes/{quote['id']}/accept", headers=headers)
    return headers, job, quote


def test_order_requires_accepted_quote(client, db):
    make_user(db, email="sales@example.com", role=Role.sales)
    headers = login(client, "sales@example.com")
    account_id, community_id = setup_account(client, headers)
    job = make_job(client, headers, account_id, community_id)
    quote = client.post(f"/jobs/{job['id']}/quotes", headers=headers, json={"name": "Draft"}).json()
    client.post(f"/quotes/{quote['id']}/lines", headers=headers, json={"sku": "B18", "list_price": "100"})
    resp = client.post(f"/jobs/{job['id']}/orders", headers=headers, json={"quote_id": quote["id"]})
    assert resp.status_code == 409


def test_generate_everluxe_order(client, db, tmp_path):
    get_settings().generated_dir = str(tmp_path)  # keep test artifacts out of backend/generated
    headers, job, quote = setup_accepted_quote(client, db)

    resp = client.post(
        f"/jobs/{job['id']}/orders", headers=headers,
        json={"quote_id": quote["id"], "customer_po": "PO-1001", "plan_name": "Model Home"},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["supplier"] == "everluxe"
    assert body["confirmation_status"] == "pending"
    assert body["has_file"] is True
    assert body["skipped_skus"] == ["RANGE1.30"]  # appliance never reaches the form

    # download works
    resp = client.get(f"/orders/{body['id']}/file", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    # and the workbook itself is right
    path = next(Path(tmp_path).glob("*.xlsx"))
    ws = load_workbook(path).active
    assert ws["B3"].value == "PO-1001"
    assert ws["B9"].value == "Townsend Dothan"
    assert ws["E14"].value == "Shaker"      # door style pulled from room selection
    assert ws["E15"].value == "Chestnut"    # door/color from finish
    # line row: qty, sku, net each (461.75*0.217=100.20), total (200.40)
    assert ws["A18"].value == 2
    assert ws["B18"].value == "B18"
    assert abs(ws["F18"].value - 100.20) < 0.001
    assert abs(ws["G18"].value - 200.40) < 0.001
    # excluded SKU not present anywhere in column B
    skus = [ws.cell(row=r, column=2).value for r in range(18, 25)]
    assert "RANGE1.30" not in skus
    # order total row (one line only, so row 20)
    assert ws["F20"].value == "Order Total"
    assert abs(ws["G20"].value - 200.40) < 0.001


def test_order_status_updates(client, db, tmp_path):
    get_settings().generated_dir = str(tmp_path)
    headers, job, quote = setup_accepted_quote(client, db)
    order = client.post(f"/jobs/{job['id']}/orders", headers=headers, json={"quote_id": quote["id"]}).json()

    resp = client.patch(
        f"/orders/{order['id']}", headers=headers,
        json={"confirmation_status": "confirmed", "ship_status": "scheduled", "po_number": "PO-9"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["confirmation_status"] == "confirmed"
    assert body["ship_status"] == "scheduled"

    listing = client.get(f"/jobs/{job['id']}/orders", headers=headers).json()
    assert len(listing) == 1
    assert listing[0]["po_number"] == "PO-9"
