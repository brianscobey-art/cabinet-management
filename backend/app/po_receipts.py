"""PO Receipts — the DOMO 'PO Receipt List' brought into CabinetTron, joined to
jobs through the tracker's POTracker table (DOMO 'Order #' = POTracker 'Our PO #').

Refresh order of preference:
  1. Live server-side DOMO pull (needs po_receipt_dataset_id + DOMO_ACCESS_TOKEN) —
     the cloud pulls it itself, no file, no browser.
  2. Newest 'PO Receipt List*.xlsx' export in po_receipt_dir (fallback / seed).

The POTracker PO->job map rides in on the normal tracker sync.
"""

import json
import logging
import re
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from sqlalchemy.orm import joinedload

from app.config import get_settings
from app.models import Job, JobPo, JobStatus, PoReceipt

logger = logging.getLogger("uvicorn.error")

# DOMO PO Receipt List columns.
COLS = ["Receipt #", "Receipt Date", "POS", "Supplier", "Supplier Cost", "Landed Cost", "Order #"]


def _as_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):  # DOMO epoch millis
        try:
            return datetime.utcfromtimestamp(float(v) / 1000.0).date()
        except Exception:  # noqa: BLE001
            return None
    s = str(v).strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s.split(" ")[0], fmt).date()
        except ValueError:
            continue
    return None


def _num(v):
    try:
        return round(float(str(v).replace(",", "").replace("$", "")), 2)
    except (TypeError, ValueError):
        return None


def _upsert_receipts(db, rows: list[dict]) -> int:
    """rows: dicts keyed by the COLS names. Full refresh (delete + insert)."""
    db.query(PoReceipt).delete()
    seen = set()
    n = 0
    for r in rows:
        num = str(r.get("Receipt #") or "").strip()
        if not num or num in seen:
            continue
        seen.add(num)
        db.add(PoReceipt(
            receipt_number=num,
            receipt_date=_as_date(r.get("Receipt Date")),
            pos=(str(r.get("POS")).strip() if r.get("POS") is not None else None),
            supplier=(str(r.get("Supplier")).strip() if r.get("Supplier") is not None else None),
            supplier_cost=_num(r.get("Supplier Cost")),
            landed_cost=_num(r.get("Landed Cost")),
            order_number=(str(r.get("Order #")).strip() if r.get("Order #") is not None else None),
        ))
        n += 1
    return n


# --- source 1: live DOMO pull --------------------------------------------------
def _domo_sql(sql: str, dataset_id: str) -> list[list]:
    s = get_settings()
    url = f"https://{s.domo_instance}/api/query/v1/execute/{dataset_id}"
    req = urllib.request.Request(
        url,
        data=json.dumps({"sql": sql}).encode(),
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-DOMO-Developer-Token": s.domo_access_token.strip(),
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read()).get("rows", [])


# The "PO Receipt List" is a card on the Purchase Receipt Details dataset:
# receipts are `transaction code`='RE', one row per receipt at the Dothan store
# (pos 750), costs summed. This grouped SQL reproduces the export's 7 columns.
_RECEIPT_SQL = (
    "SELECT `transaction number` AS `Receipt #`, MAX(`transaction date`) AS `Receipt Date`, "
    "MAX(`pos`) AS `POS`, MAX(`supplier name title`) AS `Supplier`, "
    "SUM(`reported cost`) AS `Supplier Cost`, SUM(`landed cost`) AS `Landed Cost`, "
    "MAX(`order number`) AS `Order #` "
    "FROM table WHERE `transaction code`='RE' AND `pos` LIKE '750%' "
    "GROUP BY `transaction number`"
)


def pull_receipts_domo(db) -> dict:
    s = get_settings()
    if not (s.po_receipt_dataset_id.strip() and s.domo_access_token.strip()):
        return {"error": "no PO_RECEIPT_DATASET_ID / DOMO_ACCESS_TOKEN configured"}
    try:
        raw = _domo_sql(_RECEIPT_SQL, s.po_receipt_dataset_id.strip())
    except urllib.error.HTTPError as e:
        return {"error": f"Domo returned {e.code} — token/dataset invalid?"}
    except urllib.error.URLError as e:
        return {"error": f"could not reach Domo: {e.reason}"}
    rows = [dict(zip(COLS, r)) for r in raw]
    n = _upsert_receipts(db, rows)
    return {"source": "Domo live pull", "receipts": n}


# --- source 2: export file -----------------------------------------------------
def _newest_file(folder: Path) -> Path | None:
    files = [
        f for f in folder.glob("PO Receipt List*")
        if f.suffix.lower() in (".xlsx", ".csv") and not f.name.startswith("~")
    ]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def import_receipt_file(db) -> dict:
    s = get_settings()
    folder = Path(s.po_receipt_folder)
    # Cloud: the export is uploaded to R2 by the PC — pull it down first.
    if s.r2_enabled:
        try:
            from app.storage import hydrate_feeds

            hydrate_feeds(s)
        except Exception as exc:  # noqa: BLE001 — fall through to whatever is on disk
            logger.warning("R2 hydrate before receipt import failed: %s", exc)
    if not folder.is_dir():
        return {"error": f"folder not found: {folder}"}
    f = _newest_file(folder)
    if f is None:
        return {"error": "no 'PO Receipt List*' export (.xlsx/.csv) found"}
    if f.suffix.lower() == ".csv":
        import csv as _csv

        with f.open(newline="", encoding="utf-8-sig") as fh:
            grid = [tuple(r) for r in _csv.reader(fh)]
    else:
        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb["data"] if "data" in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = list(ws.iter_rows(values_only=True))
        wb.close()
    hdr = [str(h).strip() if h is not None else "" for h in grid[0]]
    idx = {c: (hdr.index(c) if c in hdr else -1) for c in COLS}
    rows = [
        {c: (r[idx[c]] if idx[c] >= 0 and idx[c] < len(r) else None) for c in COLS}
        for r in grid[1:]
        if any(cell is not None for cell in r)
    ]
    n = _upsert_receipts(db, rows)
    return {"file": f.name, "receipts": n}


def refresh_receipts(db, with_potracker: bool = True) -> dict:
    """Live DOMO pull when configured, else the newest export file. Also refreshes
    the POTracker PO->job map when it's empty, so one button lights up the report."""
    s = get_settings()
    if s.po_receipt_dataset_id.strip() and s.domo_access_token.strip():
        res = pull_receipts_domo(db)
        if "error" in res:
            logger.warning("PO receipt live pull failed (%s) — trying file", res["error"])
            res = import_receipt_file(db)
    else:
        res = import_receipt_file(db)

    if with_potracker and db.query(JobPo).count() == 0:
        try:
            from app.feeds import _by_mtime

            files = _by_mtime(Path(s.tracker_dir), "3.0 Online Sales Tracker *.xlsm")
            if files:
                res["job_pos"] = ingest_potracker(db, files[0])
        except Exception as exc:  # noqa: BLE001
            res["job_pos"] = f"error: {exc}"
    return res


# --- POTracker PO->job map (rides in on the tracker sync) -----------------------
def ingest_potracker(db, tracker_path: Path) -> int:
    """Replace job_pos from the tracker's POTracker table (PO Tracking sheet)."""
    wb = load_workbook(tracker_path, data_only=True)
    if "PO Tracking" not in wb.sheetnames or "POTracker" not in wb["PO Tracking"].tables:
        wb.close()
        return 0
    ws = wb["PO Tracking"]
    m = re.match(r"[A-Z]+(\d+):([A-Z]+)(\d+)", ws.tables["POTracker"].ref)
    hr, mr = int(m.group(1)), int(m.group(3))
    grid = list(ws.iter_rows(min_row=hr, max_row=mr, values_only=True))
    wb.close()
    hdr = [str(h).replace("\n", " ").strip() if h else "" for h in grid[0]]

    def gi(*names):
        for name in names:
            if name in hdr:
                return hdr.index(name)
        return -1

    ji, oi, vi, pi = gi("Job Code"), gi("Our PO #"), gi("Vendor"), gi("Product")
    odi, tdi, csti = gi("Order Date", "Order  Date"), gi("Tent Due Date"), gi("Cost")
    db.query(JobPo).delete()
    n = 0
    for r in grid[1:]:
        po = r[oi] if oi >= 0 else None
        if po is None:
            continue
        db.add(JobPo(
            our_po=str(po).strip(),
            job_code=(str(r[ji]).strip() if ji >= 0 and r[ji] is not None else None),
            vendor=(str(r[vi]).strip() if vi >= 0 and r[vi] is not None else None),
            product=(str(r[pi]).strip()[:200] if pi >= 0 and r[pi] is not None else None),
            order_date=_as_date(r[odi]) if odi >= 0 else None,
            tent_due_date=_as_date(r[tdi]) if tdi >= 0 else None,
            cost=_num(r[csti]) if csti >= 0 else None,
        ))
        n += 1
    return n


# --- report --------------------------------------------------------------------
def _f(v):
    return float(v) if v is not None else None


def build_report(db, today: date | None = None) -> dict:
    """Receipts joined to jobs (via POTracker Our PO #), plus a 'still awaiting
    delivery' list of ordered POs with no receipt yet."""
    today = today or date.today()
    # our_po -> first POTracker line (job + product/vendor/dates)
    po_to_jp: dict[str, JobPo] = {}
    for jp in db.query(JobPo).all():
        if jp.our_po and jp.our_po not in po_to_jp:
            po_to_jp[jp.our_po] = jp
    jobs = {
        j.job_code: j
        for j in db.query(Job).options(joinedload(Job.account), joinedload(Job.community))
        .filter(Job.job_code.isnot(None)).all()
    }

    def job_bits(jc):
        j = jobs.get(jc)
        if j is None:
            return {"job_id": None, "address": None, "community_name": None,
                    "account_name": None, "status": None, "install_date": None}
        return {
            "job_id": j.id, "address": j.address,
            "community_name": j.community.name if j.community else None,
            "account_name": j.account.name if j.account else None,
            "status": j.status.value, "install_date": j.install_date.isoformat() if j.install_date else None,
        }

    all_receipts = db.query(PoReceipt).all()
    received_pos = set(r.order_number for r in all_receipts if r.order_number)
    # Scope to OUR receipts — those whose Order # is one of our K&B POs (in
    # POTracker). The raw DOMO list spans every Carter store/department.
    receipts = [r for r in all_receipts if r.order_number in po_to_jp]
    rows, matched = [], 0
    for r in receipts:
        jp = po_to_jp.get(r.order_number)
        jc = jp.job_code if jp else None
        if jc and jc in jobs:
            matched += 1
        rows.append({
            "receipt_number": r.receipt_number,
            "receipt_date": r.receipt_date.isoformat() if r.receipt_date else None,
            "pos": r.pos, "supplier": r.supplier,
            "supplier_cost": _f(r.supplier_cost), "landed_cost": _f(r.landed_cost),
            "order_number": r.order_number, "job_code": jc,
            "vendor": jp.vendor if jp else None, "product": jp.product if jp else None,
            "order_date": jp.order_date.isoformat() if jp and jp.order_date else None,
            **job_bits(jc),
        })
    rows.sort(key=lambda x: (x["receipt_date"] or ""), reverse=True)

    # Ordered but NOT received — active jobs only, most overdue first.
    inactive = (JobStatus.closed, JobStatus.void)
    outstanding = []
    for our_po, jp in po_to_jp.items():
        if our_po in received_pos:
            continue
        j = jobs.get(jp.job_code)
        if j is None or j.status in inactive:
            continue
        due = jp.tent_due_date
        outstanding.append({
            "order_number": our_po, "job_code": jp.job_code,
            "vendor": jp.vendor, "product": jp.product,
            "order_date": jp.order_date.isoformat() if jp.order_date else None,
            "tent_due_date": due.isoformat() if due else None,
            "days_overdue": (today - due).days if due and due < today else None,
            **job_bits(jp.job_code),
        })
    outstanding.sort(key=lambda x: (x["days_overdue"] is None, -(x["days_overdue"] or 0)))

    first_of_month = today.replace(day=1)
    this_month = sum(
        1 for r in receipts if r.receipt_date and r.receipt_date >= first_of_month
    )
    return {
        "as_of": today.isoformat(),
        "total_receipts": len(receipts),
        "received_this_month": this_month,
        "matched_to_job": matched,
        "outstanding_count": len(outstanding),
        "rows": rows,
        "outstanding": outstanding,
    }
