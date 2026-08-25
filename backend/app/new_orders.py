"""CabinetTron -> the New Orders Status workbook.

Brian's "New Orders Status.xlsx" is a flat 39-column table with no formulas on
its data sheet, and CabinetTron already carries ~35 of those columns. This
writes the app's values back into it so placing an order updates the sheet.

DESIGN: update in place, do NOT rebuild. Two columns (Zip, Folder Location)
have no home in the database, and a from-scratch rebuild would blank them. So
the workbook is opened as-is and only APP_OWNED columns are overwritten; every
other cell passes through untouched. That also keeps the Status Summary tab's
structured-reference formulas intact.

The writer runs on Brian's PC (the cloud can't see OneDrive) — see
scripts/write_new_orders.py.
"""

import logging
import re
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.orm import joinedload

from app.models import Job, OrderingChecklist

logger = logging.getLogger("uvicorn.error")

HEADER_ROW = 2   # row 1 is the title
FIRST_DATA_ROW = 3
SHEET = "New Orders"
TABLE = "NewOrders"


def _norm(h) -> str:
    """Headers carry stray double spaces ("Install  Pay $") and newlines."""
    return " ".join(str(h or "").split()).lower()


def _plan_name(plan: str | None) -> str | None:
    """Tracker plan is "DRH1 Madison STD" / "4EBF Cali"; the sheet wants the
    bare name. Done with token surgery rather than a regex — the escapes in
    this file have been mangled by the tooling twice.
    """
    if not plan:
        return None
    text = str(plan).strip()
    text = text.split(" / ")[0]          # "Palm STD w PWD / X / R" -> "Palm STD w PWD"
    parts = text.split()
    # Leading plan codes: short, mixing letters and digits ("DRH1", "4EBF").
    # There can be MORE THAN ONE — "EX4 4EBF Cali" — so strip while they last.
    while len(parts) > 1:
        head = parts[0]
        if 2 <= len(head) <= 5 and any(c.isdigit() for c in head) and any(c.isalpha() for c in head):
            parts = parts[1:]
        else:
            break
    parts = [w for w in parts if w.upper() != "STD"]   # a grade, not the name
    return " ".join(parts) or None


def _yes_no(v):
    return None if v is None else ("Yes" if v else "No")


# header (normalized) -> how to get it from the job + its checklist.
# Anything NOT listed here is left exactly as the workbook already has it.
#
# DELIBERATELY ABSENT, do not add without reading this:
#   "Status"         — in THIS sheet it means pipeline state ("Processed"),
#                      NOT job.status, which is the construction level
#                      ("4.0-Punch"). Mapping them would have rewritten 129
#                      rows into a different vocabulary.
#   "Superintendent" — the sheet carries DR Horton's own spelling
#                      ("Barrett, James A"); the app holds a display name
#                      ("Austin Barrett"). Different source of truth, and on
#                      some rows a different person entirely.
#   "Zip", "Folder Location" — no home in the database at all; they only
#                      survive because this writer leaves unlisted columns alone.
OWNED = {
    "setup date":                   lambda j, c: c.setup_date,
    "1. pos and selections":        lambda j, c: c.stage1_date,
    "2. orders and layouts":        lambda j, c: c.stage2_date,
    "3. sos and order comparison":  lambda j, c: c.stage3_date,
    "4. tbs pos attached":          lambda j, c: c.stage4_date,
    "5.0 moved to sold folder":     lambda j, c: c.moved_to_sold_date,
    "install pay $":                lambda j, c: c.install_pay,
    "install packet in folder":     lambda j, c: _yes_no(c.installer_pay_sheet),
    "carter po #":                  lambda j, c: c.carter_po_number,
    "carter so #":                  lambda j, c: c.carter_so_number,
    "carter po date":               lambda j, c: c.carter_po_date,
    "everluxe so #":                lambda j, c: c.so_number,
    "everluxe so $":                lambda j, c: c.so_amount,
    "2x4x8 syp #2":                 lambda j, c: c.lumber_2x4x8,
    "1x4x8 syp #2":                 lambda j, c: c.lumber_1x4x8,
    "1x6x8 syp #2":                 lambda j, c: c.lumber_1x6x8,
    '1/2" plywood':                 lambda j, c: c.plywood_half,
    "misc":                         lambda j, c: c.misc_materials,
    "po date":                      lambda j, c: c.po_date,
    "total po ($)":                 lambda j, c: c.po_total,
    "folder name":                  lambda j, c: c.folder_name,
}


# Columns the app fills when the sheet is BLANK but never overwrites. These
# are identity fields where the workbook is at least as authoritative, and a
# blind overwrite measurably made things worse in a dry run:
#   Swing        — the app holds "X" placeholders; the sheet has real L/R
#   Lot #        — the sheet is zero-padded ("0066"); the app stores 66
#   Subdivision  — 3 rows name a different community entirely
#   PO #         — 5 rows disagree outright on the builder PO
#   BUID         — the app is usually right, but 2 sheet rows hold a lot number
# Disagreements are reported as conflicts instead of silently resolved, so a
# human decides which side is wrong.
FILL_ONLY = {
    "buid":            lambda j, c: c.buid,
    "lot #":           lambda j, c: j.lot_number,
    "swing":           lambda j, c: c.swing,
    "subdivision":     lambda j, c: j.community.name if j.community else None,
    "plan name":       lambda j, c: _plan_name(j.plan),
    "elevation":       lambda j, c: c.elevation,
    "floorplan abbr":  lambda j, c: c.plan_abbr,
    "po #":            lambda j, c: j.builder_po,
}

RESOLVERS = {**OWNED, **FILL_ONLY}


def _clean(v):
    """openpyxl wants plain types; Decimal and date/datetime are fine, enums are not."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return v


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _checklists(db) -> dict[str, tuple[Job, OrderingChecklist]]:
    out = {}
    q = (
        db.query(Job, OrderingChecklist)
        .join(OrderingChecklist, OrderingChecklist.job_id == Job.id)
        .options(joinedload(Job.community))
        .filter(Job.job_code.isnot(None))
    )
    for job, cl in q.all():
        out[str(job.job_code).strip()] = (job, cl)
    return out


def export_values(db) -> dict:
    """{job_code: {normalized header: value}} — everything the app knows about
    this sheet, JSON-safe.

    Split out from the writer on purpose: the writer runs on Brian's PC (only
    it can see OneDrive) but the data lives in the cloud database, and shipping
    the production DB password to the PC to bridge that is not worth it. The PC
    fetches this over the API instead.
    """
    out = {}
    for code, (job, cl) in _checklists(db).items():
        row = {}
        for name, fn in RESOLVERS.items():
            v = fn(job, cl)
            if v is None:
                continue
            if isinstance(v, Decimal):
                v = float(v)
            elif isinstance(v, datetime):
                v = v.date().isoformat()
            elif isinstance(v, date):
                v = v.isoformat()
            row[name] = v
        if row:
            out[code] = row
    return out


def _revive(v):
    """ISO date strings come back from JSON as text — Excel wants datetimes."""
    if isinstance(v, str) and len(v) == 10 and v[4] == "-" and v[7] == "-":
        try:
            return datetime.strptime(v, "%Y-%m-%d")
        except ValueError:
            return v
    return v


def write_workbook(path, values: dict, dry_run: bool = False) -> dict:
    """Apply export_values() to the workbook in place.

    OWNED columns overwrite; FILL_ONLY columns fill blanks and report a
    conflict instead of overwriting. Nothing is written when dry_run, so a
    caller can show the diff first.
    """
    wb = load_workbook(path)  # keep formulas — Status Summary depends on them
    ws = wb[SHEET]
    headers = [c.value for c in ws[HEADER_ROW]]
    cols = {}
    for i, h in enumerate(headers, start=1):
        n = _norm(h)
        if n in RESOLVERS:
            cols[n] = i

    missing = sorted(set(RESOLVERS) - set(cols))
    known = values
    changes, conflicts, rows_touched, unknown = [], [], 0, []

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if not code:
            continue
        code = str(code).strip()
        vals = known.get(code)
        if vals is None:
            unknown.append(code)
            continue
        touched = False
        for name, ci in cols.items():
            new = _clean(_revive(vals.get(name)))
            if new is None:
                continue  # the app has nothing to say — leave the sheet alone
            cur = ws.cell(row=r, column=ci).value
            blank = cur is None or str(cur).strip() == ""
            if name in FILL_ONLY and not blank:
                if _s(cur) != _s(new):
                    conflicts.append((code, headers[ci - 1], cur, new))
                continue  # never overwrite an identity column that has a value
            if isinstance(cur, datetime) and isinstance(new, datetime):
                same = cur.date() == new.date()
            elif isinstance(cur, (int, float)) and isinstance(new, (int, float)):
                same = abs(float(cur) - float(new)) < 0.005
            else:
                # NOT `cur or ""` — that turns a real 0 into "" and reports a
                # phantom change on every zeroed lumber count.
                same = _s(cur) == _s(new)
            if not same:
                changes.append((code, headers[ci - 1], cur, new))
                if not dry_run:
                    ws.cell(row=r, column=ci).value = new
                touched = True
        rows_touched += 1 if touched else 0

    if not dry_run and changes:
        wb.save(path)
    wb.close()
    return {
        "rows_in_sheet": ws.max_row - HEADER_ROW,
        "rows_changed": rows_touched,
        "cells_changed": len(changes),
        "unknown_job_codes": unknown,
        "headers_not_found": missing,
        "changes": changes,
        "conflicts": conflicts,
        "written": bool(changes) and not dry_run,
    }
