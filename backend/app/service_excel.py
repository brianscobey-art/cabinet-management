"""Service Request Excel template that matches the app's on-screen/printed report,
plus a parser to round-trip a filled sheet back into the app.

Built on a fine 60-column base grid so each section (info, cabinet, hardware,
parts, service, signatures) merges cells to its own column widths — the online
form's proportions, not a single shared grid. Row 1 (above the print area)
carries the file-naming rule. Only the Job Code and the Parts / Service tables
are read on import.
"""

import io
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

PARTS_MARKER = "PARTS NEEDED"
SERVICE_MARKER = "SERVICE NEEDED"
PART_ROWS = 3
SERVICE_ROWS = 3
SPEC_ROWS = 3
NBASE = 60  # base columns; every section's spans sum to this

# (label, span) specs — each list sums to NBASE
# Parts: Item#/Qty/Part/Cabinet/Style/Color/Vendor/Vendor Order#/Order Date/Due Date/Notes/✓
PART_SPEC = [("Item #", 5), ("Qty", 4), ("Part", 6), ("Cabinet", 5), ("Style", 4),
             ("Color", 4), ("Vendor", 5), ("Vendor Order #", 5), ("Order Date", 6),
             ("Due Date", 6), ("Notes", 8), ("✓", 2)]  # sums to 60
# Cabinet specs take the left ~73%; two hardware boxes fill the right ~27%.
CAB_SPEC = [("Room / Zone", 8), ("Vendor", 7), ("Series", 7), ("Door Style", 8),
            ("Color", 7), ("Species", 7)]  # sums to 44
HW_BOXES = [("Door Hardware", 45, 8), ("Drawer Hardware", 53, 8)]  # (label, start col, span) → cols 45..60
SERVICE_SPEC = [("Part #", 5), ("Cabinet", 6), ("Description of Work", 33),
                ("✓", 3), ("Date", 7), ("Tech", 6)]  # sums to 60
INFO_SPEC = [("PROJECT", 6, 14), ("ADDRESS", 6, 14), ("LOT", 6, 14)]   # label span, value span
INFO_SPEC2 = [("JOB CODE", 6, 14), ("DATE", 6, 14), ("STATUS", 6, 14)]
SIGN_SPEC = [("Service Tech — Signature", 20), ("Date", 10),
             ("Customer — Signature", 20), ("Date", 10)]


def _starts(spec):
    """Base-column start for each field given (label, span[, ...]) specs."""
    out, c = [], 1
    for item in spec:
        out.append(c)
        c += item[1]
    return out


PART_STARTS = _starts(PART_SPEC)          # part=col10, cabinet=16, ... notes=51
SERVICE_STARTS = _starts(SERVICE_SPEC)    # part#=1, description=12

_thin = Side(style="thin", color="000000")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_GRAY_BORDER = Border(left=Side(style="medium", color="595959"), right=Side(style="medium", color="595959"),
                      top=Side(style="medium", color="595959"), bottom=Side(style="medium", color="595959"))
_GREEN = "125952"  # Carter dark green
_LABEL_FONT = Font(bold=True, size=9)
_SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")   # light gray, matches online
_HEADER_FILL = PatternFill("solid", fgColor="EDEDED")
_LABEL_FILL = PatternFill("solid", fgColor="EDEDED")
_GREEN_FILL = PatternFill("solid", fgColor=_GREEN)       # Door/Drawer hardware labels
_WHITE_BOLD = Font(bold=True, size=9, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _box(ws, row, start, span, value=None, *, kind="cell", center=False, border=_BORDER, datefmt=False):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=start + span - 1)
    cell = ws.cell(row=row, column=start, value=value)
    for c in range(start, start + span):
        ws.cell(row=row, column=c).border = border
    if kind == "section":
        cell.fill = _SECTION_FILL
        cell.font = Font(bold=True, size=10, color="222222")
        cell.alignment = _CENTER  # centered section title
    elif kind == "header":
        cell.fill = _HEADER_FILL
        cell.font = _LABEL_FONT
        cell.alignment = _CENTER
    elif kind == "label":
        cell.fill = _LABEL_FILL
        cell.font = _LABEL_FONT
        cell.alignment = _LEFT
    else:
        cell.alignment = _CENTER if center else _LEFT
    if datefmt:
        cell.number_format = "m/d/yy"
    return cell


def build_blank_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Service Request"
    ws.sheet_view.showGridLines = False
    for i in range(1, NBASE + 1):
        ws.column_dimensions[get_column_letter(i)].width = 1.4

    # --- row 1 (above print area): file-naming rule ----------------------
    _box(ws, 1, 1, 8, "FILE NAME", border=_GRAY_BORDER).fill = PatternFill("solid", fgColor="C00000")
    fn = ws.cell(row=1, column=1)
    fn.font = Font(bold=True, color="FFFFFF")
    fn.alignment = _CENTER
    rule = _box(ws, 1, 9, 34, "Service Request [Job Code] [Builder] [MMDDYY]", border=_GRAY_BORDER)
    rule.font = Font(bold=True)
    rule.alignment = _LEFT

    # --- title (left) + logo & brand (upper right) — print area starts here
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=30)
    t = ws.cell(row=3, column=1, value="SERVICE REQUEST")
    t.font = Font(size=17, bold=True, color=_GREEN)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    logo = Path(__file__).resolve().parent.parent.parent / "frontend" / "public" / "carter-logo.png"
    if logo.exists():
        img = XLImage(str(logo))
        img.height = 34
        img.width = int(34 * 946 / 228)
        ws.add_image(img, get_column_letter(44) + "3")
    ws.merge_cells(start_row=5, start_column=40, end_row=5, end_column=NBASE)
    b = ws.cell(row=5, column=40, value="Carter Kitchen and Bath")
    b.font = Font(size=11, bold=True, color=_GREEN)
    b.alignment = Alignment(horizontal="right", vertical="center")
    # green rule under the header block (matches the online header underline)
    for c in range(1, NBASE + 1):
        ws.cell(row=5, column=c).border = Border(bottom=Side(style="medium", color=_GREEN))

    # --- info grid (boxed, 2 rows) ---------------------------------------
    def info_row(r, spec):
        starts = _starts([(l, ls + vs) for l, ls, vs in spec])
        for (label, lspan, vspan), start in zip(spec, starts):
            _box(ws, r, start, lspan, label, kind="label")
            _box(ws, r, start + lspan, vspan)
    info_row(6, INFO_SPEC)
    info_row(7, INFO_SPEC2)

    r = 9

    def section(title, spec, blanks, *, center_idx=(), date_idx=()):
        nonlocal r
        _box(ws, r, 1, NBASE, title, kind="section")
        r += 1
        starts = _starts(spec)
        for (label, span), start in zip(spec, starts):
            _box(ws, r, start, span, label, kind="header")
        ws.row_dimensions[r].height = 24  # room for headers that wrap to 2 lines
        r += 1
        for _ in range(blanks):
            for idx, ((_l, span), start) in enumerate(zip(spec, starts)):
                _box(ws, r, start, span, center=idx in center_idx, datefmt=idx in date_idx)
            r += 1

    # --- CABINET SPECIFICATIONS & HARDWARE (specs table left, hardware boxes right) ---
    _box(ws, r, 1, NBASE, "CABINET SPECIFICATIONS & HARDWARE", kind="section")
    r += 1
    cab_starts = _starts(CAB_SPEC)
    for (label, span), start in zip(CAB_SPEC, cab_starts):      # cabinet headers (left)
        _box(ws, r, start, span, label, kind="header")
    for label, start, span in HW_BOXES:                        # green hardware labels (right)
        cell = _box(ws, r, start, span, label, center=True)
        cell.fill = _GREEN_FILL
        cell.font = _WHITE_BOLD
    ws.row_dimensions[r].height = 26  # fit "Drawer Hardware" without clipping
    for i in range(SPEC_ROWS):
        r += 1
        for (label, span), start in zip(CAB_SPEC, cab_starts):  # blank cabinet rows
            _box(ws, r, start, span)
        if i == 0:                                              # SKU value under each green label
            for label, start, span in HW_BOXES:
                _box(ws, r, start, span, center=True)
    r += 2

    section(PARTS_MARKER, PART_SPEC, PART_ROWS, center_idx={0, 1, 7, 8, 9, 11}, date_idx={8, 9})
    r += 1
    section(SERVICE_MARKER, SERVICE_SPEC, SERVICE_ROWS, center_idx={0, 3, 4}, date_idx={4})

    # --- signatures on one line ------------------------------------------
    r += 1
    starts = _starts(SIGN_SPEC)
    for (_l, span), start in zip(SIGN_SPEC, starts):
        ws.merge_cells(start_row=r, start_column=start, end_row=r, end_column=start + span - 1)
        for c in range(start, start + span):
            ws.cell(row=r, column=c).border = Border(bottom=_thin)
    r += 1
    for (label, span), start in zip(SIGN_SPEC, starts):
        c = ws.cell(row=r, column=start, value=label)
        c.font = Font(size=8, color="555555")
    # footer rule (matches the thin top border on the online footer)
    r += 2
    for c in range(1, NBASE + 1):
        ws.cell(row=r, column=c).border = Border(top=Side(style="thin", color="999999"))
    ws.cell(row=r, column=1,
            value="Fill in the Job Code and the Parts / Service tables, save, and import into "
                  "Carter Kitchen and Bath. Part # in Service refers to the Item # in Parts.").font = Font(
        italic=True, size=8, color="888888")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.15, footer=0.15)
    ws.print_area = f"A3:{get_column_letter(NBASE)}{r}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Parser
# --------------------------------------------------------------------------

def _find_marker(ws, marker: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and str(v).strip().upper() == marker:
            return row
    return None


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None and str(v).strip() != "" else None


def _label_value(ws, label: str):
    """Find a cell whose text == label near the top and read the next non-empty cell to its right."""
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, min(ws.max_column, NBASE + 2) + 1):
            v = ws.cell(row=row, column=col).value
            if v and str(v).strip().lower() == label:
                for c2 in range(col + 1, min(ws.max_column, NBASE + 2) + 1):
                    got = _cell(ws, row, c2)
                    if got:
                        return got
                return None
    return None


def _pdate(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _int(s, default=None):
    return int(float(s)) if s and str(s).replace(".", "").isdigit() else default


def parse_import(data: bytes) -> dict:
    """Read a filled template into {job_code, title, parts[], lines[]}."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    job_code = _label_value(ws, "job code")
    title = _label_value(ws, "title")

    parts_head = _find_marker(ws, PARTS_MARKER)
    service_head = _find_marker(ws, SERVICE_MARKER)
    if parts_head is None or service_head is None:
        raise ValueError("This does not look like a Service Request template (missing section headers).")

    ps = PART_STARTS  # base-col start of each parts field
    parts = []
    for row in range(parts_head + 2, service_head):
        part = _cell(ws, row, ps[2])
        if not part:
            continue
        parts.append({
            "item_num": _int(_cell(ws, row, ps[0])),
            "qty": _int(_cell(ws, row, ps[1]), 1) or 1,
            "part": part,
            "cabinet": _cell(ws, row, ps[3]),
            "style": _cell(ws, row, ps[4]),
            "color": _cell(ws, row, ps[5]),
            "vendor": _cell(ws, row, ps[6]),
            "order_number": _cell(ws, row, ps[7]),
            "order_date": _pdate(ws.cell(row=row, column=ps[8]).value),
            "due_date": _pdate(ws.cell(row=row, column=ps[9]).value),
            "notes": _cell(ws, row, ps[10]),
        })

    ss = SERVICE_STARTS
    lines = []
    for row in range(service_head + 2, ws.max_row + 1):
        desc = _cell(ws, row, ss[2])
        if not desc:
            continue
        lines.append({"part_num": _int(_cell(ws, row, ss[0])), "instruction": desc})

    return {"job_code": job_code, "title": title, "parts": parts, "lines": lines}
