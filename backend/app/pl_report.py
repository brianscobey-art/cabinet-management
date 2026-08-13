"""Read NET SALES year-to-date from the newest K&B P&L workbook.

The monthly P&L files live in settings.pl_reports_dir (a OneDrive share locally;
uploaded to R2 so the cloud can see them). Each is the TNKB cost-center sheet
with month columns and a "2026 Total" column; we pull the NET SALES row.

Fails soft: returns None if the folder/file/row isn't found, so the report still
renders (it just shows "P&L not available").
"""

import logging
from pathlib import Path

from app.config import Settings, get_settings

logger = logging.getLogger("uvicorn.error")

SHEET = "TNKB"
# Row labels that carry the net-sales figure, best first.
NET_SALES_LABELS = ("NET SALES BUDGET", "NET SALES GOAL", "NET SALES")


def _newest_file(folder: Path) -> Path | None:
    files = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~")]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def read_net_sales(settings: Settings | None = None) -> dict | None:
    s = settings or get_settings()
    folder = Path(s.pl_reports_dir)
    if not folder.is_dir():
        return None
    f = _newest_file(folder)
    if f is None:
        return None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:  # noqa: BLE001
        logger.warning("P&L read failed for %s: %s", f.name, exc)
        return None

    # Find the "…Total" column (year total) from the header rows.
    total_col = None
    for r in grid[:10]:
        for ci, val in enumerate(r):
            if isinstance(val, str) and "total" in val.lower():
                total_col = ci
                break
        if total_col is not None:
            break
    # Find a NET SALES row (label lives in column B / index 1).
    for label in NET_SALES_LABELS:
        for r in grid:
            cell = r[1] if len(r) > 1 else None
            if isinstance(cell, str) and cell.strip().upper() == label:
                val = r[total_col] if (total_col is not None and total_col < len(r)) else None
                if isinstance(val, (int, float)):
                    return {"value": float(val), "source_file": f.name, "label": cell.strip()}
    return {"value": None, "source_file": f.name, "label": None}
