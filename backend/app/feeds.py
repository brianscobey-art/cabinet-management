"""Daily feed sync — pulls the outputs of Brian's scheduled cloud reports into the app.

Two feeds, both written to OneDrive by cloud tasks each morning:
- Vendor Suite merge (~3:45 AM): DRH_Cabinets_Combined_{MMDDYY}[.Rn].xlsx —
  every open DRH cabinet PO + schedule dates across all 5 regions.
- Century SupplyPro refresh (~6:00 AM): Century Production Report-{MMDDYY}.xlsx —
  all Century cabinet jobs with status and delivery dates.

Sync is idempotent: jobs match on community + lot; existing jobs get their
feed segment in notes refreshed (marked VS:/SP:), install dates updated, and
early statuses bumped to 'ordered' when an open PO exists. Unknown lots become
new jobs, with job codes derived from coded siblings in the same community
(e.g. Links Crossing siblings DRLICR-#### -> DRLICR-0135).
"""

import os
import re
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models import Account, AccountType, Community, Job, JobStatus, JobType
from app.sales import resolve_salesperson

DEFAULT_SALES_CONTACT = ("Brian Scobey", "850-890-0482", "Brian.Scobey@TownsendBuildingSupply.com")

VS_REGION_ACCOUNTS = {
    "Mont": "DR Horton Montgomery",
    "PC West": "DR Horton Panama City West",
    "PC East": "DR Horton Panama City East",
    "Pensa East": "DR Horton Pensacola East",
    "Pensa West": "DR Horton Pensacola West",
}

CENTURY_ACCOUNT = "Century PC"


def _canon_lot(lot) -> str | None:
    if lot is None:
        return None
    s = str(lot).strip().lstrip("0")
    return s or "0"


def _as_date(value) -> date | None:
    return value.date() if isinstance(value, datetime) else None


def _fmt(d: date | None) -> str:
    return f"{d.month}/{d.day}/{d:%y}" if d else ""


def _money(value):
    from decimal import Decimal, InvalidOperation

    if value is None or isinstance(value, str):
        return None
    try:
        d = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return d if d != 0 else None


def _po_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    return s or None


def latest_file(directory: Path, pattern: str) -> Path | None:
    files = [f for f in directory.glob(pattern) if not f.name.startswith("~")]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def _load_wb(path: Path):
    """Open a workbook read-only; if Excel/OneDrive holds a lock, read a shadow copy.

    Returns (workbook, cleanup) — call cleanup() after wb.close().
    """
    try:
        return load_workbook(path, read_only=True, data_only=True), (lambda: None)
    except PermissionError:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copyfile(path, tmp.name)
        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        return wb, (lambda: os.unlink(tmp.name))


def _get_or_create_account(db: Session, name: str) -> Account:
    account = db.query(Account).filter(Account.name == name).first()
    if account is None:
        account = Account(name=name, type=AccountType.builder)
        db.add(account)
        db.flush()
    return account


def _get_or_create_community(db: Session, account: Account, name: str) -> Community:
    community = (
        db.query(Community)
        .filter(Community.account_id == account.id, Community.name == name)
        .first()
    )
    if community is None:
        community = Community(account_id=account.id, name=name)
        db.add(community)
        db.flush()
    return community


def _find_job(db: Session, community: Community, lot) -> Job | None:
    want = _canon_lot(lot)
    if want is None:
        return None
    for job in db.query(Job).filter(Job.community_id == community.id).all():
        if _canon_lot(job.lot_number) == want:
            return job
    return None


def _derive_job_code(db: Session, community: Community, lot4: str) -> str | None:
    """New lots inherit the code prefix their coded siblings use (DRLICR-0135)."""
    prefixes: dict[str, int] = {}
    for (code,) in (
        db.query(Job.job_code).filter(Job.community_id == community.id, Job.job_code.isnot(None)).all()
    ):
        if "-" in code:
            prefix = code.rsplit("-", 1)[0]
            prefixes[prefix] = prefixes.get(prefix, 0) + 1
    if not prefixes:
        return None
    prefix = max(prefixes, key=lambda p: prefixes[p])
    candidate = f"{prefix}-{lot4}"
    if db.query(Job).filter(Job.job_code == candidate).first():
        return None  # collision — leave uncoded rather than mislabel
    return candidate


def _set_feed_note(job: Job, tag: str, text: str) -> None:
    """Keep exactly one 'VS:'/'SP:' segment in notes, refreshed each sync."""
    segment = f"{tag} {text}".strip()
    notes = job.notes or ""
    pattern = re.compile(rf"{tag}[^|]*")
    if pattern.search(notes):
        job.notes = pattern.sub(segment, notes).strip(" |")
    else:
        job.notes = f"{notes} | {segment}".strip(" |")


def _bump_to_ordered(job: Job) -> bool:
    if job.status in (JobStatus.track, JobStatus.preord, JobStatus.ndord):
        job.status = JobStatus.ord
        return True
    return False


def sync_vendorsuite(db: Session, path: Path) -> dict:
    wb, cleanup = _load_wb(path)
    tab = next((n for n in wb.sheetnames if n.startswith("DRH Cabinets Combined")), None)
    if tab is None:
        wb.close()
        cleanup()
        raise ValueError(f"No 'DRH Cabinets Combined' tab in {path.name}")
    rows = wb[tab].iter_rows(min_row=2, values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    counts = {"created": 0, "updated": 0, "skipped": 0}
    for values in rows:
        row = {h: values[i] for h, i in idx.items() if i < len(values)}
        region = row.get("Region")
        account_name = VS_REGION_ACCOUNTS.get(str(region).strip()) if region else None
        project = row.get("Project")
        job_number = row.get("Job Number")
        if not account_name or not project or job_number is None:
            counts["skipped"] += 1
            continue

        account = _get_or_create_account(db, account_name)
        community = _get_or_create_community(db, account, str(project).strip())
        lot4 = str(job_number)[-4:]

        po_bits = []
        if row.get("PO Number"):
            amount = row.get("PO Amount")
            amount_s = f" ${amount:,.2f}" if isinstance(amount, (int, float)) else ""
            po_bits.append(f"PO# {row['PO Number']}{amount_s} ({row.get('PO Status') or '?'})")
        measure = _as_date(row.get("Cabinet Measure/Order"))
        install = _as_date(row.get("Cabinet Install"))
        punch = _as_date(row.get("Cabinet Trim/Punch"))
        for label, d in (("measure", measure), ("install", install), ("punch", punch)):
            if d:
                po_bits.append(f"{label} {_fmt(d)}")
        feed_text = ", ".join(po_bits) or "on report, no PO detail"

        # builder-side PO facts, straight from the authoritative VS report
        po_number = _po_str(row.get("PO Number"))
        po_amount = _money(row.get("PO Amount"))
        po_status = str(row.get("PO Status")).strip() if row.get("PO Status") else None

        job = _find_job(db, community, lot4)
        if job is None:
            plan = " ".join(str(row.get(k) or "") for k in ("Plan", "Elevation", "Swing")).strip()
            job = Job(
                job_code=_derive_job_code(db, community, lot4),
                account_id=account.id,
                community_id=community.id,
                lot_number=lot4.lstrip("0") or "0",
                address=str(row.get("Street Address") or f"{project} Lot {lot4}").title(),
                job_type=JobType.tract,
                plan=plan[:100] or None,
                status=JobStatus.ord if row.get("PO Number") else JobStatus.track,
                measure_date=measure,
                install_date=install,
                warranty_start_date=install,
                builder_po=po_number,
                po_amount=po_amount,
                po_status=po_status,
                salesperson=resolve_salesperson(account_name),
                sales_contact_name=DEFAULT_SALES_CONTACT[0],
                sales_contact_phone=DEFAULT_SALES_CONTACT[1],
                sales_contact_email=DEFAULT_SALES_CONTACT[2],
                field_contact_name="DRH Superintendent",
                notes=f"Plan: {plan} | VS BUID {job_number}" if plan else f"VS BUID {job_number}",
            )
            db.add(job)
            db.flush()
            _set_feed_note(job, "VS:", feed_text)
            counts["created"] += 1
        else:
            changed = False
            # Scheduling is app-managed: feeds never touch install_date on existing
            # jobs (the VS: note still carries the builder's schedule for reference).
            # Measure date fills in only when the app has none yet.
            if measure and job.measure_date is None:
                job.measure_date = measure
                changed = True
            # PO facts are builder-authoritative — refresh from the daily VS report
            for attr, value in (("builder_po", po_number), ("po_amount", po_amount), ("po_status", po_status)):
                if value is not None and getattr(job, attr) != value:
                    setattr(job, attr, value)
                    changed = True
            if row.get("PO Number") and str(row.get("PO Status")).strip().lower() == "open":
                changed = _bump_to_ordered(job) or changed
            before = job.notes
            _set_feed_note(job, "VS:", feed_text)
            counts["updated" if (changed or job.notes != before) else "skipped"] += 1
    wb.close()
    cleanup()
    db.commit()
    return counts


CENTURY_PLACEHOLDER_CONTACTS = {"TBD", "Century Superintendent"}
_FAKE_PHONES = {"555-555-5555", "5555555555"}


def _century_subdivision(name) -> str:
    """SupplyPro lists some communities twice ('Grand Oaks at Callaway' and
    'LCA - Grand Oaks at Callaway') — the LCA prefix is the same place."""
    s = str(name).strip()
    return s[6:].strip() if s.upper().startswith("LCA - ") else s


def century_candidates(directory: Path) -> list[Path]:
    """SupplyPro Century files, newest first by the MMDDYY in the filename
    (revision R.N breaks ties). Files without a date code are ignored."""
    dated = []
    for f in directory.glob("Century Cabinet Jobs - SupplyPro*.xlsx"):
        if f.name.startswith("~"):
            continue
        m = re.search(r"(\d{2})(\d{2})(\d{2})(?:\s+R\.?(\d+))?", f.name)
        if not m:
            continue
        mm, dd, yy, rev = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4) or 0)
        try:
            dated.append(((date(2000 + yy, mm, dd), rev), f))
        except ValueError:
            continue
    return [f for _, f in sorted(dated, key=lambda t: t[0], reverse=True)]


def sync_century(db: Session, path: Path) -> dict:
    """SupplyPro 'Century Cabinet Jobs' workbook -> Century jobs.

    'Cabinet Jobs' tab: subdivision/lot/address, builder contact, and the
    measure/order/deliver dates. 'Cabinet POs' tab adds plan and PO number.
    """
    wb, cleanup = _load_wb(path)
    if "Cabinet Jobs" not in wb.sheetnames:
        wb.close()
        cleanup()
        raise ValueError(f"No 'Cabinet Jobs' tab in {path.name}")

    def read_tab(name: str) -> list[dict]:
        rows = wb[name].iter_rows(min_row=2, values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        return [
            {h: values[i] for h, i in idx.items() if i < len(values)}
            for values in rows
        ]

    records: dict[tuple[str, str], dict] = {}
    for row in read_tab("Cabinet Jobs"):
        sub, lot = row.get("Subdivision"), row.get("Lot")
        if not sub or lot is None:
            continue
        records[(_century_subdivision(sub), _canon_lot(lot))] = {"jobs": row}
    if "Cabinet POs" in wb.sheetnames:
        for row in read_tab("Cabinet POs"):
            sub, lot = row.get("Subdivision"), row.get("Lot")
            if not sub or lot is None:
                continue
            records.setdefault((_century_subdivision(sub), _canon_lot(lot)), {})["pos"] = row
    wb.close()
    cleanup()

    account = _get_or_create_account(db, CENTURY_ACCOUNT)
    counts = {"created": 0, "updated": 0, "skipped": 0}
    for (sub, _), rec in records.items():
        jobs_row = rec.get("jobs", {})
        pos_row = rec.get("pos", {})
        lot = jobs_row.get("Lot") or pos_row.get("Lot")
        community = _get_or_create_community(db, account, sub)

        measure = _as_date(jobs_row.get("Measure Cabinets")) or _as_date(pos_row.get("Measure Date"))
        ordered = _as_date(jobs_row.get("Order Cabinets"))
        deliver = _as_date(jobs_row.get("Deliver Cabinets")) or _as_date(pos_row.get("Deliver / Install Date"))
        po_number = pos_row.get("PO Number")
        plan = " ".join(
            str(pos_row.get(k) or "") for k in ("Plan", "Elevation", "Swing") if pos_row.get(k)
        ).replace("See Start", "").strip()

        bits = [b for b in (
            f"PO {po_number}" if po_number else None,
            f"measure {_fmt(measure)}" if measure else None,
            f"order {_fmt(ordered)}" if ordered else None,
            f"deliver {_fmt(deliver)}" if deliver else None,
        ) if b]
        feed_text = ", ".join(bits) or "on report"

        contact = str(jobs_row.get("Builder Contact") or "").strip()
        phone = str(jobs_row.get("Contact Phone") or "").strip()
        if phone.replace("-", "") in {p.replace("-", "") for p in _FAKE_PHONES}:
            phone = ""

        job = _find_job(db, community, lot)
        if job is None:
            address = str(jobs_row.get("Address") or f"{sub} Lot {lot}").title()
            city, state = jobs_row.get("City"), jobs_row.get("State")
            if city:
                address = f"{address}, {city}, {state or ''}".strip(", ")
            job = Job(
                account_id=account.id,
                community_id=community.id,
                lot_number=str(lot).strip(),
                address=address,
                job_type=JobType.tract,
                plan=plan[:100] or None,
                status=JobStatus.ord if ordered and ordered <= date.today() else JobStatus.track,
                measure_date=measure,
                builder_po=_po_str(po_number),
                salesperson=resolve_salesperson(CENTURY_ACCOUNT),
                sales_contact_name=DEFAULT_SALES_CONTACT[0],
                sales_contact_phone=DEFAULT_SALES_CONTACT[1],
                sales_contact_email=DEFAULT_SALES_CONTACT[2],
                field_contact_name=contact.title() if contact else "Century Superintendent",
                field_contact_phone=phone or None,
            )
            db.add(job)
            db.flush()
            _set_feed_note(job, "SP:", feed_text)
            counts["created"] += 1
        else:
            changed = False
            if measure and job.measure_date is None:
                job.measure_date = measure
                changed = True
            if po_number and job.builder_po != _po_str(po_number):
                job.builder_po = _po_str(po_number)
                changed = True
            if plan and not job.plan:
                job.plan = plan[:100]
                changed = True
            if contact and job.field_contact_name in CENTURY_PLACEHOLDER_CONTACTS:
                job.field_contact_name = contact.title()
                if phone:
                    job.field_contact_phone = phone
                changed = True
            if ordered and ordered <= date.today():
                changed = _bump_to_ordered(job) or changed
            before = job.notes
            _set_feed_note(job, "SP:", feed_text)
            counts["updated" if (changed or job.notes != before) else "skipped"] += 1
    db.commit()
    return counts


def sync_new_orders(db: Session, path: Path) -> dict:
    """Brian's New Orders Status workbook -> the 4-stage ordering checklists.

    Fill-forward only: a dated stage in the file checks the stage with that
    date; nothing already checked in the app is ever touched or un-checked.
    """
    from app.models import OrderingChecklist

    wb, cleanup = _load_wb(path)
    if "New Orders" not in wb.sheetnames:
        wb.close()
        cleanup()
        raise ValueError(f"No 'New Orders' tab in {path.name}")
    rows = wb["New Orders"].iter_rows(min_row=2, values_only=True)
    headers = [str(h).replace("\n", " ").strip() if h else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    def col(prefix: str) -> str | None:
        return next((h for h in headers if h.startswith(prefix)), None)

    stage_cols = {
        "stage1": col("1."),
        "stage2": col("2."),
        "stage3": col("3."),
        "stage4": col("4."),
    }
    counts = {"updated": 0, "unchanged": 0, "no_job": 0}
    for values in rows:
        row = {h: values[i] for h, i in idx.items() if i < len(values)}
        code = row.get("Job Code")
        if not code or str(code).startswith("#"):
            continue
        job = db.query(Job).filter(Job.job_code == str(code).strip()).first()
        if job is None:
            counts["no_job"] += 1
            continue

        checklist = db.query(OrderingChecklist).filter(OrderingChecklist.job_id == job.id).first()
        if checklist is None:
            checklist = OrderingChecklist(job_id=job.id)
            db.add(checklist)
            db.flush()

        changed = False
        for stage, header in stage_cols.items():
            done_date = _as_date(row.get(header)) if header else None
            if done_date and not getattr(checklist, f"{stage}_done"):
                setattr(checklist, f"{stage}_done", True)
                setattr(checklist, f"{stage}_date", done_date)
                changed = True

        bits = []
        if row.get(col("Carter PO")):
            bits.append(f"Carter PO# {row[col('Carter PO')]}")
        if row.get(col("Carter SO")):
            bits.append(f"Carter SO# {row[col('Carter SO')]}")
        if _as_date(row.get(col("5.0"))):
            bits.append(f"moved to sold folder {_fmt(_as_date(row[col('5.0')]))}")
        note = " | ".join(str(b) for b in bits)
        if note and note not in (checklist.notes or ""):
            checklist.notes = note
            changed = True

        counts["updated" if changed else "unchanged"] += 1
    wb.close()
    cleanup()
    db.commit()
    return counts


def _sync_newest_readable(db: Session, files: list[Path], sync_fn) -> dict:
    """Try files newest-first — a workbook open in Excel is unreadable, so fall
    back to the next-most-recent copy rather than failing the whole sync."""
    if not files:
        return {"error": "no file found"}
    skipped = []
    for f in files[:5]:
        try:
            return {"file": f.name, "skipped_locked": skipped, **sync_fn(db, f)}
        except PermissionError:
            skipped.append(f.name)
    return {"error": "all recent files locked", "skipped_locked": skipped}


def _by_mtime(directory: Path, pattern: str) -> list[Path]:
    return sorted(
        (f for f in directory.glob(pattern) if not f.name.startswith("~")),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )


def sync_all(db: Session) -> dict:
    """Run all feeds against the newest readable file in each OneDrive location."""
    settings = get_settings()
    result = {
        "vendorsuite": _sync_newest_readable(
            db, _by_mtime(Path(settings.vendorsuite_dir), "DRH_Cabinets_Combined_*.xlsx"), sync_vendorsuite
        ),
        "century": _sync_newest_readable(
            db, century_candidates(Path(settings.century_dir)), sync_century
        ),
    }
    new_orders = Path(settings.new_orders_file)
    if new_orders.is_file():
        try:
            result["new_orders"] = {"file": new_orders.name, **sync_new_orders(db, new_orders)}
        except PermissionError:
            result["new_orders"] = {"error": "file locked (open in Excel)"}
    else:
        result["new_orders"] = {"error": "file not found"}
    return result
