"""Everluxe order form generator — replaces the manual Excel→Everluxe form step.

Layout mirrors the dealer's existing order form: order meta, Dealer Information,
Shipping Information, then a QTY/SKU/Product Code/FIN End/Color/Total Each/
Total Cost/Notes line table with an order total. Excluded SKUs (appliance
placeholders, see app.pricing) are never written to the order.
"""

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Job, Quote
from app.pricing import is_excluded, line_total, money, net_each

AQUA = "24B3C6"
GRAY = "595959"

_thin = Side(style="thin", color=GRAY)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_header_font = Font(name="Arial", size=9, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor=AQUA)
_label_font = Font(name="Arial", size=9, bold=True)
_value_font = Font(name="Arial", size=9)
_center = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass
class OrderFormInfo:
    """Everything on the form that isn't a line item. Defaults come from settings."""

    dealer_name: str
    dealer_contact: str
    dealer_phone: str
    dealer_email: str
    ship_to_name: str
    delivery_address: str
    delivery_city_st_zip: str
    delivery_type: str = "Everluxe Truck"
    assembly: bool = True
    customer_po: str = ""
    job_code: str = ""
    plan_name: str = ""
    area: str = ""
    door_style: str = ""
    door_color: str = ""
    freight: str = ""
    order_date: str = ""  # filled by caller (route knows "today")
    skipped_skus: list[str] = field(default_factory=list)


def _label(ws, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = _label_font


def _value(ws, cell: str, text) -> None:
    ws[cell] = text
    ws[cell].font = _value_font
    ws[cell].border = _border


def build_order_workbook(job: Job, quote: Quote, info: OrderFormInfo) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Everluxe Order"

    # Title + order meta
    ws.merge_cells("A1:H1")
    ws["A1"] = f"EVERLUXE ORDER — {job.account.name.upper()}"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A1"].fill = _header_fill
    ws["A1"].alignment = _center

    field_pairs = [
        # (label cell, label, value cell, value)
        ("A3", "Customer PO", "B3", info.customer_po),
        ("A4", "Order Date", "B4", info.order_date),
        ("A5", "Pricing Model", "B5", f"Multip. {quote.multiplier}"),
        ("A6", "Freight", "B6", info.freight),
        ("D3", "Plan Name", "E3", info.plan_name),
        ("D4", "Area", "E4", info.area),
        # Dealer block
        ("A9", "Dealer Name", "B9", info.dealer_name),
        ("A10", "Dealer Contact", "B10", info.dealer_contact),
        ("A11", "Phone Number", "B11", info.dealer_phone),
        ("A12", "Email Address", "B12", info.dealer_email),
        ("A13", "Job Code", "B13", info.job_code),
        # Shipping block
        ("D9", "Ship To Name", "E9", info.ship_to_name),
        ("D10", "Delivery Address", "E10", info.delivery_address),
        ("D11", "Delivery City, ST, Zip", "E11", info.delivery_city_st_zip),
        ("D12", "Delivery Type", "E12", info.delivery_type),
        ("D13", "Assembly", "E13", "Yes" if info.assembly else "No"),
        ("D14", "Door Style", "E14", info.door_style),
        ("D15", "Door/Color", "E15", info.door_color),
    ]
    _label(ws, "A8", "Dealer Information:")
    _label(ws, "D8", "Shipping Information:")
    for label_cell, label, value_cell, value in field_pairs:
        _label(ws, label_cell, label)
        _value(ws, value_cell, value)

    # Line-item table
    header_row = 17
    headers = ["QTY", "SKU", "Product Code", "FIN End", "Color", "Total Each", "Total Cost", "Notes"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _center
        cell.border = _border

    row = header_row + 1
    order_total = Decimal("0")
    for line in quote.lines:
        if is_excluded(line.sku):
            info.skipped_skus.append(line.sku)
            continue
        each = net_each(line.list_price, quote.multiplier)
        total = line_total(line.list_price, line.qty, quote.multiplier)
        order_total += total
        values = [line.qty, line.sku, line.product_code, line.fin_end, line.color,
                  float(each), float(total), line.notes]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = _value_font
            cell.border = _border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).number_format = "$#,##0.00"
        ws.cell(row=row, column=7).number_format = "$#,##0.00"
        row += 1

    total_cell = ws.cell(row=row + 1, column=7, value=float(money(order_total)))
    total_cell.number_format = "$#,##0.00"
    total_cell.font = _label_font
    ws.cell(row=row + 1, column=6, value="Order Total").font = _label_font

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["H"].width = 28

    return wb


def write_order_file(job: Job, quote: Quote, info: OrderFormInfo, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_job = "".join(c if c.isalnum() or c in " -_" else "_" for c in job.address)[:60].strip()
    path = out_dir / f"Everluxe Order - Job {job.id} - {safe_job}.xlsx"
    build_order_workbook(job, quote, info).save(path)
    return path
