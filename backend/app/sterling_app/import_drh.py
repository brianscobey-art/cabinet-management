"""One-shot importer for Brian's "DRH All Division Pricing" workbook.

Pulls in:
- Everluxe Grand Inventory  -> catalog items with price-group prices G1-G5
- door style/color header    -> DoorStyle rows (style -> price group)
- Install Pricing            -> PlanInstall (per division+plan install rates)
- Floorplan SKU              -> PlanTemplateItem (per-plan SKU list with doors/drawers)
                                + per-SKU doors/drawers onto the catalog
- Hardware-Accessories       -> catalog items (vendor Hardware Resources, multiplier 1)

Safe to re-run: everything upserts.
"""

import warnings
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.sterling_app.models import CatalogItem, DoorStyle, PlanInstall, PlanTemplateItem

# 021126 (2/11/26) supersedes the OneDrive Trackers 021125 copy — 194 plans vs 135.
DEFAULT_PATH = r"C:\Users\Brian SE6\Downloads\DRH All Division Pricing 021126.xlsm"


def _dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def import_workbook(db: Session, path: str | None = None) -> dict:
    src = Path(path or DEFAULT_PATH)
    if not src.exists():
        raise FileNotFoundError(f"Workbook not found: {src}")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(src, read_only=True, data_only=True)
    counts = {}

    # --- Everluxe Grand Inventory: SKU rows x price-group columns ---
    ws = wb["Everluxe Grand Inventory"]
    rows = list(ws.iter_rows(values_only=True, max_col=11))
    # Row 3 (idx 2) = style names, row 4 (idx 3) = codes, row 5 (idx 4) = group numbers.
    style_names, style_codes, group_row = rows[2], rows[3], rows[4]
    col_groups = {}  # column index -> group number
    styles = 0
    for col in range(1, 11):
        group = _int(group_row[col])
        name = str(style_names[col] or "").strip()
        if not group or not name:
            continue
        col_groups[col] = group
        code = str(style_codes[col] or "").strip() or None
        existing = db.query(DoorStyle).filter(DoorStyle.name == name).first()
        if existing:
            existing.code, existing.price_group = code, group
        else:
            db.add(DoorStyle(name=name, code=code, price_group=group))
            styles += 1
    counts["door_styles"] = styles

    def get_item(vendor: str, sku: str) -> CatalogItem:
        item = (
            db.query(CatalogItem)
            .filter(CatalogItem.vendor == vendor, CatalogItem.sku == sku)
            .first()
        )
        if item is None:
            item = CatalogItem(vendor=vendor, sku=sku)
            db.add(item)
        return item

    skus = 0
    for row in rows[5:]:
        sku = str(row[0] or "").strip().upper()
        if not sku or sku.lower() in ("sku", "none"):
            continue
        groups: dict[int, Decimal] = {}
        for col, group in col_groups.items():
            price = _dec(row[col])
            if price is not None and group not in groups:
                groups[group] = price
        if not groups:
            continue
        item = get_item("Everluxe", sku)
        for g, price in groups.items():
            setattr(item, f"price_g{g}", price)
        item.list_price = groups.get(1) or next(iter(groups.values()))
        item.category = item.category or "Cabinet"
        skus += 1
    counts["everluxe_skus"] = skus

    # --- Install Pricing: plan / division / rates ---
    ws = wb["Install Pricing"]
    installs = 0
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=8):
        plan = str(row[0] or "").strip()
        division = str(row[2] or "").strip()
        if not plan or not division:
            continue
        rec = (
            db.query(PlanInstall)
            .filter(PlanInstall.division == division, PlanInstall.plan == plan)
            .first()
        )
        if rec is None:
            rec = PlanInstall(division=division, plan=plan)
            db.add(rec)
        rec.cabinet_install = _dec(row[3]) or Decimal("0")
        rec.knob_install = _dec(row[4]) or Decimal("0")
        rec.handle_install = _dec(row[5]) or Decimal("0")
        installs += 1
    counts["plan_installs"] = installs

    # --- Floorplan SKU: per-plan SKU lists + doors/drawers ---
    ws = wb["Floorplan SKU"]
    db.query(PlanTemplateItem).delete()  # full refresh — the sheet is the truth
    template_rows = 0
    sku_pieces: dict[str, tuple[int, int]] = {}  # sku -> (doors_each, drawers_each)
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=10):
        division = str(row[0] or "").strip()
        plan = str(row[1] or "").strip()
        sku = str(row[3] or "").strip().upper()
        if not division or not plan or not sku:
            continue
        qty = _int(row[2]) or 1
        doors, drawers = _int(row[8]), _int(row[9])
        db.add(
            PlanTemplateItem(
                division=division, plan=plan, sku=sku, qty=qty,
                area=(str(row[4]).strip() if row[4] else None),
                doors=doors, drawers=drawers,
            )
        )
        template_rows += 1
        # Doors/Drawers columns are line totals; derive per-cabinet counts.
        if qty > 0 and (doors or drawers) and sku not in sku_pieces:
            if doors % qty == 0 and drawers % qty == 0:
                sku_pieces[sku] = (doors // qty, drawers // qty)
    counts["plan_template_rows"] = template_rows

    db.flush()  # session has autoflush=False — make pending catalog rows queryable
    pieces = 0
    for sku, (doors, drawers) in sku_pieces.items():
        item = (
            db.query(CatalogItem)
            .filter(CatalogItem.vendor == "Everluxe", CatalogItem.sku == sku)
            .first()
        )
        if item:
            item.doors, item.drawers = doors, drawers
            pieces += 1
    counts["skus_with_hardware_counts"] = pieces

    # --- Table8 (Floorplan SKU AE:AO): per-plan Assembly/Install/Hardware units.
    # This is what the live Pricing Sheet computes labor from (units x rate) —
    # it covers plans the old Install Pricing sheet hasn't priced yet.
    db.flush()  # make the Install Pricing rows above queryable (autoflush=False)
    division_by_plan: dict[str, str] = {}
    for row in wb["Floorplan SKU"].iter_rows(min_row=4, values_only=True, max_col=2):
        if row[0] and row[1]:
            division_by_plan.setdefault(str(row[1]).strip(), str(row[0]).strip())

    units = 0
    for row in wb["Floorplan SKU"].iter_rows(min_row=5, values_only=True, min_col=31, max_col=41):
        plan = str(row[0] or "").strip()
        if not plan or plan.lower() == "floorplan":
            continue
        assembly, install, hw_count = _int(row[8]), _int(row[9]), _int(row[10])
        division = division_by_plan.get(plan)
        if division is None:
            continue
        rec = (
            db.query(PlanInstall)
            .filter(PlanInstall.division == division, PlanInstall.plan == plan)
            .first()
        )
        if rec is None:
            rec = PlanInstall(division=division, plan=plan)
            db.add(rec)
        rec.assembly_units = assembly
        rec.install_units = install
        # Keep the dollar columns matrix-consistent: install units x $25, hw x $1/$2.
        rec.cabinet_install = Decimal(install * 25)
        rec.knob_install = Decimal(hw_count * 1)
        rec.handle_install = Decimal(hw_count * 2)
        units += 1
    counts["plan_labor_units"] = units

    # --- Hardware-Accessories: knobs/pulls at cost ---
    ws = wb["Hardware-Accessories"]
    hardware = 0
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=6):
        sku = str(row[0] or "").strip().upper()
        cost = _dec(row[5])
        if not sku or cost is None:
            continue
        item = get_item(str(row[4] or "Hardware Resources").strip(), sku)
        item.description = str(row[3] or "").strip() or item.description
        item.category = "Hardware"
        item.list_price = cost
        item.multiplier = Decimal("1")  # cost-each pricing, no dealer multiplier
        hardware += 1
    counts["hardware_items"] = hardware

    wb.close()
    db.commit()
    return counts
