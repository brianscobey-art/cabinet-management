"""Sterling reports — one definition, three renderings (screen, Excel, PowerPoint).

A report is a REPORTS entry: a title, a subtitle, columns, and a build function
returning {"meta": {...}, "rows": [...]}. The screen, the .xlsx and the .pptx all
read that one structure, so a column added here shows up in every export and the
three can never drift apart.

Column spec: (key, label, kind) where kind is
    text | num | money | pct | pill   — kind drives alignment and formatting.
"""

from __future__ import annotations

import collections
import statistics
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Callable

CARTER_GREEN = "125952"
NEG = "9E2B25"
POS = "2F6F5E"
WARN = "B07D21"


@dataclass(frozen=True)
class Report:
    key: str
    title: str
    blurb: str
    columns: list[tuple[str, str, str]]
    build: Callable
    notes: str = ""
    inputs: list[dict] = field(default_factory=list)
    group_by: str | None = None   # column to split into per-group tabs


# --------------------------------------------------------------------------
# Plan margin vs actual PO
# --------------------------------------------------------------------------
MARGIN_COLUMNS = [
    ("plan", "Plan", "text"),
    ("division", "Division", "text"),
    ("n", "Houses", "num"),
    ("po_price", "PO price", "pill"),
    ("avg", "Avg PO", "money"),
    ("cogs", "Our cost", "money"),
    ("sale", "Our price", "money"),
    ("gap", "Gap / house", "money"),
    ("exp", "12-month", "money"),
    ("margin", "Margin", "pct"),
    ("target", "Target", "pct"),
    ("status", "Status", "text"),
]


ADJUST_BELOW = 0.10   # a division under this needs its pricing revisited


def rollup(rows: list[dict], group_key: str, target_default: float = 0.15) -> list[dict]:
    """Group plan rows into a per-division summary.

    Margin is DOLLAR-WEIGHTED, not the mean of the plan margins: a plan built
    once should not move a division as much as one built twenty times, and the
    mean would let a single bad one-off drag the whole division under.
    """
    groups: dict[str, list[dict]] = collections.defaultdict(list)
    for r in rows:
        groups[r.get(group_key) or "—"].append(r)

    out = []
    for name, rs in groups.items():
        houses = sum(r["n"] for r in rs)
        revenue = sum(r["avg"] * r["n"] for r in rs)
        cost = sum(r["cogs"] * r["n"] for r in rs)
        margin = (revenue - cost) / revenue if revenue else 0.0
        target = max((r.get("target") or target_default) for r in rs) if rs else target_default
        below = sum(1 for r in rs if r["margin"] < ADJUST_BELOW)
        action = ("Adjust pricing" if margin < ADJUST_BELOW
                  else "Watch" if margin < target else "OK")
        out.append({
            "division": name, "plans": len(rs), "houses": houses,
            "revenue": round(revenue, 2), "cost": round(cost, 2),
            "avg_po": round(revenue / houses, 2) if houses else 0.0,
            "margin": round(margin, 4), "target": round(target, 4),
            "below10": below,
            "exposure": round(sum(r["exp"] for r in rs), 2),
            "action": action,
        })
    out.sort(key=lambda g: g["margin"])
    return out


DIVISION_COLUMNS = [
    ("division", "Division", "text"),
    ("plans", "Plans", "num"),
    ("houses", "Houses", "num"),
    ("avg_po", "Avg PO", "money"),
    ("revenue", "Revenue", "money"),
    ("cost", "Cost", "money"),
    ("margin", "Avg margin", "pct"),
    ("below10", "Plans under 10%", "num"),
    ("exposure", "12-month", "money"),
    ("action", "Action", "text"),
]


def _vs_po_by_job(path: str) -> tuple[dict, list]:
    """{job number: summed PO amount} from the Vendor Suite combined report.

    Summed per JOB, not per row: a house can carry a change order or backcharge
    as a second PO line, and counting those as separate houses turns a -$100
    adjustment into a whole house at -$100.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    name = next((s for s in wb.sheetnames if s.startswith("DRH Cabinets Combined")), None)
    if name is None:
        wb.close()
        raise ValueError("no 'DRH Cabinets Combined*' sheet in that workbook")
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    by_job: dict[str, float] = collections.defaultdict(float)
    dates = []
    for r in rows[1:]:
        job, amt = r[ix["Job Number"]], r[ix["PO Amount"]]
        if job and isinstance(amt, (int, float)):
            by_job[str(job).strip()] += float(amt)
        d = r[ix.get("PO Date", 0)]
        if hasattr(d, "year"):
            dates.append(d)
    wb.close()
    return dict(by_job), dates


def build_plan_margin(db, vs_path: str | None = None, **_) -> dict:
    """Priced plans against what DR Horton actually paid."""
    from app.sterling_app.compute import national_pricing_rows

    if not vs_path:
        vs_path = _newest_vs()
    if not vs_path:
        raise ValueError("No Vendor Suite combined report found — pick one to run this report.")

    by_job, dates = _vs_po_by_job(vs_path)

    # BUID -> plan comes from CabinetTron, which is a separate database.
    from app.database import SessionLocal as CtSession
    from app.models import Job as CtJob, OrderingChecklist

    with CtSession() as ct:
        plan_of = {
            str(c.buid).strip(): (j.plan or "").strip()
            for j, c in ct.query(CtJob, OrderingChecklist)
            .join(OrderingChecklist, OrderingChecklist.job_id == CtJob.id)
            .all()
            if c.buid
        }

    priced = {r["plan"]: r for r in national_pricing_rows(db, None, None)}

    agg: dict[str, list[float]] = collections.defaultdict(list)
    skipped = collections.Counter()
    for buid, total in by_job.items():
        plan = plan_of.get(buid)
        if plan is None:
            skipped["no job in CabinetTron"] += 1
        elif not plan:
            skipped["job has no plan"] += 1
        elif plan not in priced:
            skipped["plan not priced in Sterling"] += 1
        else:
            agg[plan].append(total)

    rows = []
    for plan, amounts in agg.items():
        p = priced[plan]
        cogs, sale = float(p["cogs"]), float(p["sale"])
        target = float(p["margin_pct"])
        target = target / 100 if target > 1 else target
        avg = statistics.mean(amounts)
        spread = max(amounts) - min(amounts)
        margin = (avg - cogs) / avg if avg else 0.0
        status = ("Below cost" if margin < 0
                  else "Below target" if margin < target - 0.005 else "OK")
        rows.append({
            "plan": plan, "division": p["division"], "n": len(amounts),
            "po_price": "Fixed" if spread < 1 else f"Varies ${spread:,.0f}",
            "avg": round(avg, 2), "cogs": round(cogs, 2), "sale": round(sale, 2),
            "gap": round(avg - sale, 2), "exp": round((avg - sale) * len(amounts), 2),
            "margin": round(margin, 4), "target": round(target, 4), "status": status,
        })
    rows.sort(key=lambda r: r["exp"])

    under = -sum(r["exp"] for r in rows if r["exp"] < 0)
    over = sum(r["exp"] for r in rows if r["exp"] > 0)
    meta = {
        "source": Path(vs_path).name,
        "period": (f"{min(dates).month}/{min(dates).day}/{min(dates):%y} – "
                   f"{max(dates).month}/{max(dates).day}/{max(dates):%y}" if dates else ""),
        "houses": sum(r["n"] for r in rows),
        "plans": len(rows),
        "under": round(under, 2),
        "over": round(over, 2),
        "net": round(under - over, 2),
        "below_cost": sum(1 for r in rows if r["margin"] < 0),
        "skipped": dict(skipped),
        "headline": [
            ("Net, 12 months", f"-${under - over:,.0f}", "under-contract minus over"),
            ("Priced below", f"-${under:,.0f}",
             f"{sum(1 for r in rows if r['exp'] < 0)} plans"),
            ("Priced above", f"+${over:,.0f}",
             f"{sum(1 for r in rows if r['exp'] > 0)} plans"),
            ("Below cost", str(sum(1 for r in rows if r["margin"] < 0)),
             "plans under our cost"),
        ],
    }
    return {"meta": meta, "rows": rows}


def _newest_vs() -> str | None:
    """Newest DRH_Cabinets_Combined_*.xlsx from the feed folder."""
    from app.config import get_settings

    folder = Path(get_settings().vendorsuite_dir)
    if not folder.is_dir():
        return None
    files = sorted(folder.glob("DRH_Cabinets_Combined_*.xlsx"),
                   key=lambda f: f.stat().st_mtime, reverse=True)
    return str(files[0]) if files else None


# --------------------------------------------------------------------------
# Reprice compliance — POs written after a price change, at the old price
# --------------------------------------------------------------------------
REPRICE_COLUMNS = [
    ("job_code", "Job", "text"),
    ("division", "Division", "text"),
    ("plan", "Plan", "text"),
    ("po_date", "PO date", "text"),
    ("po_number", "PO #", "text"),
    ("po_amount", "PO amount", "money"),
    ("expected", "Should be", "money"),
    ("prior", "Old price", "money"),
    ("short", "Short by", "money"),
    ("verdict", "Verdict", "text"),
]

TOL = 1.0   # a dollar — these are whole-dollar contract prices


def build_reprice_check(db, vs_path: str | None = None, **_) -> dict:
    """Every PO dated on or after a reprice, checked against the agreed price.

    Matching the SUPERSEDED price to the dollar is the tell: it means the old
    number was used, not that the house was unusual. Anything else is reported
    as "check" rather than accused of anything.
    """
    from app.sterling_app.models import ContractPrice

    prices = collections.defaultdict(list)
    for cp in db.query(ContractPrice).all():
        prices[(cp.division, cp.plan)].append(cp)
    if not prices:
        raise ValueError("No contract prices loaded — import a pricing matrix first.")
    for v in prices.values():
        v.sort(key=lambda c: c.effective_from)

    if not vs_path:
        vs_path = _newest_vs()
    if not vs_path:
        raise ValueError("No Vendor Suite combined report found.")

    from openpyxl import load_workbook

    wb = load_workbook(vs_path, data_only=True, read_only=True)
    name = next((n for n in wb.sheetnames if n.startswith("DRH Cabinets Combined")), None)
    ws = wb[name]
    raw = list(ws.iter_rows(min_row=2, values_only=True))
    hdr = [str(h).strip() if h else "" for h in raw[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    # Keep the date PER LINE. Using the job's earliest date hides exactly the
    # case this report exists for: a base PO written after a reprice with an
    # earlier-dated change order beside it, which drags the job's min date back
    # before the cutoff and drops it from the check.
    jobs = collections.defaultdict(list)
    for r in raw[1:]:
        b = r[ix["Job Number"]]
        if b:
            jobs[str(b).strip()].append(
                (r[ix["PO Number"]], r[ix["PO Amount"]], r[ix["PO Date"]]))
    wb.close()

    from app.database import SessionLocal as CtSession
    from app.models import Job as CtJob, OrderingChecklist

    with CtSession() as ct:
        info = {
            str(c.buid).strip(): (j.job_code or "", (j.plan or "").strip())
            for j, c in ct.query(CtJob, OrderingChecklist)
            .join(OrderingChecklist, OrderingChecklist.job_id == CtJob.id).all()
            if c.buid
        }

    rows = []
    checked = 0
    for buid, lines in jobs.items():
        code, plan = info.get(buid, ("", ""))
        if not plan:
            continue
        match = next((k for k in prices if k[1] == plan), None)
        if match is None:
            continue

        # The base PO is the largest line; the rest are change orders and
        # backcharges, which should not be judged against a contract price.
        priced_lines = [(n, float(a), d) for n, a, d in lines
                        if isinstance(a, (int, float))]
        if not priced_lines:
            continue
        base_po, base, base_date = max(priced_lines, key=lambda x: x[1])
        if not hasattr(base_date, "year"):
            continue
        po_date = base_date.date() if hasattr(base_date, "date") else base_date

        in_force = [c for c in prices[match] if c.effective_from <= po_date]
        if not in_force:
            continue
        cp = in_force[-1]
        checked += 1
        expected = float(cp.price)
        upch = float(cp.color_upcharge or 0)
        prior = float(cp.prior_price) if cp.prior_price is not None else None

        if abs(base - expected) <= TOL or (upch and abs(base - (expected + upch)) <= TOL):
            continue                                   # correct, nothing to report
        if prior is not None and abs(base - prior) <= TOL:
            verdict = "Old price used"
        else:
            verdict = "Check — matches neither"
        rows.append({
            "job_code": code or buid, "division": cp.division, "plan": plan,
            "po_date": f"{po_date.month}/{po_date.day}/{po_date:%y}",
            "po_number": str(base_po or ""), "po_amount": round(base, 2),
            "expected": round(expected, 2),
            "prior": round(prior, 2) if prior is not None else None,
            "short": round(base - expected, 2), "verdict": verdict,
        })

    rows.sort(key=lambda r: r["short"])
    short_total = -sum(r["short"] for r in rows if r["short"] < 0)
    old_used = sum(1 for r in rows if r["verdict"] == "Old price used")
    eff = {f"{k[0]}": max(c.effective_from for c in v) for k, v in prices.items()}
    meta = {
        "source": Path(vs_path).name,
        "checked": checked,
        "flagged": len(rows),
        "effective": ", ".join(f"{d} from {e}" for d, e in sorted(set(eff.items()))),
        "headline": [
            ("POs checked", str(checked), "dated on/after a reprice"),
            ("Old price used", str(old_used), "matches the superseded price exactly"),
            ("Needs a look", str(len(rows) - old_used), "matches neither price"),
            ("Under-billed", f"-${short_total:,.0f}", "against the agreed price"),
        ],
    }
    return {"meta": meta, "rows": rows}


REPORTS: dict[str, Report] = {
    "reprice-check": Report(
        key="reprice-check",
        title="Reprice Compliance",
        blurb="POs raised after a price change that were not written at the new price.",
        columns=REPRICE_COLUMNS,
        build=build_reprice_check,
        notes=(
            "Only the base PO is judged — the largest line on a job. Change orders "
            "and backcharges are separate and are not compared to a contract price. "
            "A PO matching the SUPERSEDED price to the dollar is reported as the old "
            "price being used; anything else is flagged for a look rather than "
            "assumed wrong."
        ),
    ),
    "plan-margin": Report(
        key="plan-margin",
        title="Plan Margin vs Actual PO",
        blurb="Every priced house plan against what DR Horton actually paid, "
              "ranked by 12-month exposure.",
        columns=MARGIN_COLUMNS,
        build=build_plan_margin,
        group_by="division",
        notes=(
            "PO amounts are summed per job — a house can carry a change order or "
            "backcharge as a second PO line. Margin is against cabinet cost, not "
            "total: the DRH cabinet PO covers cabinets and install, while "
            "countertops are billed separately."
        ),
    ),
}


def run(key: str, db, **kwargs) -> dict:
    rep = REPORTS.get(key)
    if rep is None:
        raise KeyError(key)
    data = rep.build(db, **kwargs)
    data["report"] = {"key": rep.key, "title": rep.title, "blurb": rep.blurb,
                      "notes": rep.notes, "group_by": rep.group_by,
                      "columns": [{"key": k, "label": lb, "kind": kd}
                                  for k, lb, kd in rep.columns],
                      "group_columns": [{"key": k, "label": lb, "kind": kd}
                                        for k, lb, kd in DIVISION_COLUMNS]}
    if rep.group_by:
        data["groups"] = rollup(data["rows"], rep.group_by)
    today = date.today()
    # %-m is glibc-only; this app runs on Windows too.
    data["meta"]["generated"] = f"{today.month}/{today.day}/{today:%y}"
    return data


def catalog() -> list[dict]:
    return [{"key": r.key, "title": r.title, "blurb": r.blurb} for r in REPORTS.values()]
