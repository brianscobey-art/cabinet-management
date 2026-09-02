"""Sales Order Cover Sheet as a fillable .xlsx — same layout as the web form.

Built to the "make-excel-template" rules: one narrow base column, every field a
merged span, borders across whole merges, live formulas for the totals, and
print setup that fits one portrait letter page.

Geometry: 40 base columns at width 2.0 (~19 px each) = 7.9in, i.e. the full
printable width of letter at 0.3in margins. Field spans below are sized to the
characters each one actually holds (addresses/emails ~30, money ~10).

`build_cover_workbook(sheet)` returns BytesIO; pass None for a blank form.
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

GREEN = "125952"
BAND_TXT = "FFFFFF"
LABEL_FILL = "EAF2F0"
THIN = Side(style="thin", color="7F918D")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '"$"#,##0.00;;""'          # blank instead of $0.00
MONEY_HARD = '"$"#,##0.00'
PCT = '0.0%;;""'

COLS = 40          # base grid; N, Z, AM are the wide ones
# Column widths are CALIBRATED against Excel itself (openpyxl's width value is
# not the pixel count — Excel subtracts padding and scales by the Normal font).
# Verified via COM: written 2.10 -> 15.2px, written 7.77 -> 56.0px, which read
# as 19px and 70px in Excel on a 125%-scaled display. Total 7.61in of an 8.00in
# page, so it still prints at 100% with no fit-to-page shrink.
COL_W = 2.10                                       # reads 19px
WIDE_W = 7.77                                      # reads 70px
WIDE_COLS = {14: WIDE_W, 26: WIDE_W, 39: WIDE_W}   # N, Z, AM
ROW_ENTRY = 18     # typing rows — roomy enough to read and click
ROW_TABLE = 16
ROW_BAND = 19

# Header blocks (label span, value span) — values sized to ~24-30 characters
SALE_L, SALE_V = 5, 9          # "Ashley's Code" is the longest label here
JOB_L, JOB_V = 3, 10           # short labels, long values (addresses, emails)
CUST_L, CUST_V = 3, 10         # 14 + 13 + 13 = 40


def _cell(ws, row, col, span, value=None, *, bold=False, fill=None, align="left",
          wrap=False, fmt=None, size=10, color=None, border=True, shrink=False):
    """One merged field; returns the next free column.

    shrink=True lets Excel scale long entries down instead of clipping them —
    every cell someone types into uses it (wrap and shrink are exclusive).
    """
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=size, bold=bold,
                  color=color or (BAND_TXT if fill == GREEN else "000000"))
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap,
                            shrink_to_fit=shrink and not wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if border:
        for cc in range(col, col + span):
            ws.cell(row=row, column=cc).border = BORDER
    return col + span


def _band(ws, row, text, col=1, span=COLS):
    _cell(ws, row, col, span, text, bold=True, fill=GREEN, align="center", size=10.5)
    ws.row_dimensions[row].height = ROW_BAND


def _pair(ws, row, col, label, value, lab_span, val_span, fmt=None):
    """Label cell + fillable value cell."""
    nxt = _cell(ws, row, col, lab_span, label, bold=True, fill=LABEL_FILL, size=9, shrink=True)
    return _cell(ws, row, nxt, val_span, value, fmt=fmt, size=10, shrink=True)


def build_cover_workbook(s: dict | None = None) -> io.BytesIO:
    s = s or {}
    pos = s.get("pos", [])
    products = [p for p in pos if p.get("kind") != "labor"]
    labor = [p for p in pos if p.get("kind") == "labor"]
    PROD_ROWS, LAB_ROWS = 8, 4          # blank PO lines on an empty form

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Order Cover Sheet"
    for c in range(1, COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = WIDE_COLS.get(c, COL_W)

    r = 1
    ws.row_dimensions[r].height = 24
    _cell(ws, r, 1, 26, "SALES ORDER COVER SHEET", bold=True, size=15,
          color=GREEN, border=False)
    _cell(ws, r, 27, 14, s.get("job_code") or "", bold=True, size=12, align="right",
          border=False)
    r += 1
    ws.row_dimensions[r].height = 15
    _cell(ws, r, 1, 26, "Carter Kitchen & Bath", size=9.5, color="555555", border=False)
    _cell(ws, r, 27, 14, f"Printed {date.today():%m/%d/%y}", size=9, align="right",
          color="555555", border=False)
    ws.row_dimensions[r + 1].height = 7      # spacer
    r += 2

    # ---- three header blocks: Sale | Job Information | Customer ----
    sale_c, job_c, cust_c = 1, 1 + SALE_L + SALE_V, 1 + SALE_L + SALE_V + JOB_L + JOB_V
    _band(ws, r, "Sale", sale_c, SALE_L + SALE_V)
    _band(ws, r, "Job Information", job_c, JOB_L + JOB_V)
    _band(ws, r, "Customer", cust_c, CUST_L + CUST_V)
    r += 1
    sale_rows = [
        ("Sale Date", s.get("sale_date")), ("Plan Type", s.get("plan_type")),
        ("Ashley's Code", s.get("customer_account")), ("G-Code", s.get("job_number")),
        ("I-Code", s.get("install_code")), ("Cab Job Code", s.get("job_code")),
        ("Scope", s.get("scope")),
    ]
    job_rows = [
        ("Name", s.get("ji_name")), ("Contact", s.get("ji_contact")),
        ("Address", s.get("ji_address")), ("City", s.get("ji_city")),
        ("State", s.get("ji_state")), ("Zip", s.get("ji_zip")),
        ("Phone", s.get("ji_phone")), ("Email", s.get("ji_email")),
    ]
    cust_rows = [
        ("Company", s.get("cu_company")), ("Name", s.get("cu_name")),
        ("Address", s.get("cu_address")), ("City", s.get("cu_city")),
        ("State", s.get("cu_state")), ("Zip", s.get("cu_zip")),
        ("Phone", s.get("cu_phone")), ("Email", s.get("cu_email")),
    ]
    block_rows = max(len(sale_rows), len(job_rows), len(cust_rows))
    for i in range(block_rows):
        ws.row_dimensions[r + i].height = ROW_ENTRY
        if i < len(sale_rows):
            _pair(ws, r + i, sale_c, sale_rows[i][0], sale_rows[i][1], SALE_L, SALE_V)
        else:  # keep the block's box square
            _cell(ws, r + i, sale_c, SALE_L + SALE_V, None)
        if i < len(job_rows):
            _pair(ws, r + i, job_c, job_rows[i][0], job_rows[i][1], JOB_L, JOB_V)
        if i < len(cust_rows):
            _pair(ws, r + i, cust_c, cust_rows[i][0], cust_rows[i][1], CUST_L, CUST_V)
    r += block_rows

    # ---- superintendent: name 12 / phone 8 / email 8 wide ----
    _band(ws, r, "Superintendent")
    r += 1
    ws.row_dimensions[r].height = ROW_ENTRY
    col = _cell(ws, r, 1, 4, "Name", bold=True, fill=LABEL_FILL, size=9, shrink=True)
    col = _cell(ws, r, col, 11, s.get("super_name"), shrink=True)
    col = _cell(ws, r, col, 4, "Phone", bold=True, fill=LABEL_FILL, size=9, shrink=True)
    col = _cell(ws, r, col, 7, s.get("super_phone"), shrink=True)
    col = _cell(ws, r, col, 4, "Email", bold=True, fill=LABEL_FILL, size=9, shrink=True)
    _cell(ws, r, col, COLS - col + 1, s.get("super_email"), shrink=True)
    ws.row_dimensions[r + 1].height = 7      # spacer
    r += 2

    # PO columns sized to content: PO number 19ch, vendor 27ch, code 13ch,
    # type 16ch, money 10ch each.
    # = 40. Spans chosen so the wide columns (14, 26, 39) land inside Vendor,
    # Type and Total; Cost gets a 5th column so a full figure isn't squeezed.
    PO_SPANS = [5, 11, 5, 6, 5, 4, 4]
    MONEY_COL_1 = 1 + sum(PO_SPANS[:4])          # first money column
    MONEY_COL_2 = MONEY_COL_1 + PO_SPANS[4]
    TOTAL_COL = MONEY_COL_2 + PO_SPANS[5]

    def po_table(start, title, rows, count, c1, c2):
        _band(ws, start, title)
        hr = start + 1
        ws.row_dimensions[hr].height = ROW_TABLE
        col = 1
        for label, span, align in zip(
            ("Job / PO Number", "Vendor", "Vendor Code", "Type", c1, c2, "Total"),
            PO_SPANS,
            ("center",) * 7,          # column headers centered across their span
        ):
            col = _cell(ws, hr, col, span, label, bold=True, fill=GREEN, align=align, size=9, shrink=True)
        first = hr + 1
        for i in range(count):
            rr = first + i
            ws.row_dimensions[rr].height = ROW_TABLE
            p = rows[i] if i < len(rows) else {}
            col = _cell(ws, rr, 1, PO_SPANS[0], p.get("po_number") or "", size=10, shrink=True)
            col = _cell(ws, rr, col, PO_SPANS[1], p.get("vendor"), size=10, shrink=True)
            col = _cell(ws, rr, col, PO_SPANS[2], p.get("vendor_code"), size=10, shrink=True)
            col = _cell(ws, rr, col, PO_SPANS[3], p.get("po_type"), size=10, shrink=True)
            col = _cell(ws, rr, col, PO_SPANS[4],
                        float(p["amount1"]) if p.get("amount1") else None,
                        align="center", fmt=MONEY, size=10, shrink=True)
            col = _cell(ws, rr, col, PO_SPANS[5],
                        float(p["amount2"]) if p.get("amount2") else None,
                        align="center", fmt=MONEY, size=10, shrink=True)
            a1, a2 = get_column_letter(MONEY_COL_1), get_column_letter(MONEY_COL_2)
            _cell(ws, rr, col, PO_SPANS[6], f"={a1}{rr}+{a2}{rr}",
                  align="center", fmt=MONEY, size=10, shrink=True)
        last = first + count - 1
        tr = last + 1
        ws.row_dimensions[tr].height = ROW_TABLE
        _cell(ws, tr, 1, TOTAL_COL - 1, f"{title} total", bold=True, fill=LABEL_FILL,
              align="right", size=9.5, shrink=True)
        tcol = get_column_letter(TOTAL_COL)
        _cell(ws, tr, TOTAL_COL, PO_SPANS[6], f"=SUM({tcol}{first}:{tcol}{last})",
              bold=True, align="center", fmt=MONEY_HARD, size=10)
        return tr

    prod_total_row = po_table(r, "PO's Needed — Products", products,
                              max(PROD_ROWS, len(products)), "Cost", "Freight")
    ws.row_dimensions[prod_total_row + 1].height = 7   # spacer
    r = prod_total_row + 2
    lab_total_row = po_table(r, "PO's Needed — Labor", labor,
                             max(LAB_ROWS, len(labor)), "Assemble", "Install")
    ws.row_dimensions[lab_total_row + 1].height = 7    # spacer
    r = lab_total_row + 2

    # ---- summary: cost | contract | margin (15 + 13 + 12 = 40) ----
    COST_L, COST_V = 8, 7
    CONTRACT_L, CONTRACT_V = 6, 7
    MARGIN_L, MARGIN_V = 5, 7
    cost_c = 1
    contract_c = cost_c + COST_L + COST_V
    margin_c = contract_c + CONTRACT_L + CONTRACT_V
    _band(ws, r, "Cost Summary", cost_c, COST_L + COST_V)
    _band(ws, r, "Contract", contract_c, CONTRACT_L + CONTRACT_V)
    _band(ws, r, "Margin", margin_c, MARGIN_L + MARGIN_V)
    r += 1

    T = get_column_letter(TOTAL_COL)
    mat_ref, lab_ref = f"{T}{prod_total_row}", f"{T}{lab_total_row}"
    cost_val_col = cost_c + COST_L
    CV = get_column_letter(cost_val_col)
    tax_row = r + 4
    for i, label in enumerate(("Total Materials", "Total Labor", "Total Tax", "Total COGS")):
        rr = r + i
        ws.row_dimensions[rr].height = ROW_ENTRY
        _cell(ws, rr, cost_c, COST_L, label, bold=(i == 3), fill=LABEL_FILL, size=9.5, shrink=True)
        formula = (
            f"={mat_ref}", f"={lab_ref}",
            f"={mat_ref}*{CV}{tax_row}",
            f"={CV}{r}+{CV}{r + 1}+{CV}{r + 2}",
        )[i]
        _cell(ws, rr, cost_val_col, COST_V, formula, align="center",
              fmt=MONEY_HARD, bold=(i == 3), size=10, shrink=True)
    ws.row_dimensions[tax_row].height = ROW_TABLE
    _cell(ws, tax_row, cost_c, COST_L, "Tax rate (on materials)", fill=LABEL_FILL, size=9, shrink=True)
    _cell(ws, tax_row, cost_val_col, COST_V, float(s.get("tax_pct") or 7) / 100,
          align="center", fmt="0.0%", size=10)

    contract_val_col = contract_c + CONTRACT_L
    KV = get_column_letter(contract_val_col)
    for i, (label, val) in enumerate((
        ("Cabinets", s.get("sale_cabinets")), ("Countertops", s.get("sale_countertops")),
        ("Other", s.get("sale_other")), ("Total Sale", None),
    )):
        rr = r + i
        _cell(ws, rr, contract_c, CONTRACT_L, label, bold=(i == 3), fill=LABEL_FILL, size=9.5, shrink=True)
        if i < 3:
            v = float(val) if val not in (None, "") and float(val) else None
            _cell(ws, rr, contract_val_col, CONTRACT_V, v, align="center", fmt=MONEY,
                  size=10, shrink=True)
        else:
            _cell(ws, rr, contract_val_col, CONTRACT_V, f"=SUM({KV}{r}:{KV}{r + 2})",
                  bold=True, align="center", fmt=MONEY_HARD, size=10)
    sale_cell = f"{KV}{r + 3}"
    cogs_cell = f"{CV}{r + 3}"

    margin_val_col = margin_c + MARGIN_L
    for i, label in enumerate(("Dollars", "Percent")):
        rr = r + i
        _cell(ws, rr, margin_c, MARGIN_L, label, fill=LABEL_FILL, size=9.5, shrink=True)
        formula = (
            f"={sale_cell}-{cogs_cell}",
            f'=IF({sale_cell}=0,"",({sale_cell}-{cogs_cell})/{sale_cell})',
        )[i]
        _cell(ws, rr, margin_val_col, MARGIN_V, formula, align="center",
              fmt=(MONEY_HARD if i == 0 else PCT), bold=True, size=10, shrink=True)
    # square off the margin block against the taller cost/contract columns
    for i in (2, 3):
        _cell(ws, r + i, margin_c, MARGIN_L + MARGIN_V, None)
    r += 5

    _band(ws, r, "Notes")
    r += 1
    ws.row_dimensions[r].height = 32
    _cell(ws, r, 1, COLS, s.get("notes"), wrap=True, align="left", size=10)
    last_row = r

    # ---- print setup: one portrait letter page ----
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(COLS)}{last_row}"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.25)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
