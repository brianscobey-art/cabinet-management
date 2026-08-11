"""Excel workbook persistence — the .xlsx IS the database.

`CKB Pricing Data.xlsx` (next to the app) holds every table as a styled Excel
Table: Settings, Catalog, PlanBids, Jobs, Rooms, LineItems. SQLite is only a
runtime cache: the workbook is loaded into it at startup (and via Reload), and
every committed change writes the workbook back (debounced ~1.5 s).

Excel lock gotcha (same as Brian's other workbooks): while the file is open in
Excel the save fails with PermissionError — we keep the dirty flag and retry
every few seconds until Excel lets go, and report status to the UI.
"""

import enum
import shutil
import tempfile
import threading
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.sterling_app.database import SessionLocal
from app.sterling_app.models import (
    CatalogItem,
    CostModel,
    DoorStyle,
    InstallMode,
    Job,
    JobType,
    LineItem,
    PlanBid,
    PlanInstall,
    PlanTemplateItem,
    PlanTops,
    PriceSnapshot,
    PriceSnapshotRow,
    PricingMode,
    Room,
    Setting,
    Stage,
    TopPiece,
)

from app.sterling_app.database import APP_DIR, DATA_DIR

WORKBOOK_PATH = DATA_DIR / "CKB Pricing Data.xlsx"
# The workbook committed with the app seeds a fresh deployment's disk.
SEED_PATH = APP_DIR / "CKB Pricing Data.xlsx"

# (sheet, table name, model, [(header, attr)...]) — IDs included so relations survive round-trips.
SHEETS = [
    ("Settings", "Settings", Setting, [("Key", "key"), ("Value", "value")]),
    ("Catalog", "Catalog", CatalogItem, [
        ("ID", "id"), ("SKU", "sku"), ("Description", "description"), ("Vendor", "vendor"),
        ("Category", "category"), ("List Price", "list_price"), ("Multiplier", "multiplier"),
        ("G1", "price_g1"), ("G2", "price_g2"), ("G3", "price_g3"),
        ("G4", "price_g4"), ("G5", "price_g5"), ("Doors", "doors"), ("Drawers", "drawers"),
        ("Install Group", "install_group"), ("Assemble Value", "assemble_value"),
        ("Install Value", "install_value"),
    ]),
    ("DoorStyles", "DoorStyles", DoorStyle, [
        ("ID", "id"), ("Name", "name"), ("Code", "code"),
        ("Price Group", "price_group"), ("Vendor", "vendor"),
    ]),
    ("PlanInstalls", "PlanInstalls", PlanInstall, [
        ("ID", "id"), ("Division", "division"), ("Plan", "plan"),
        ("Cabinet Install", "cabinet_install"), ("Knob Install", "knob_install"),
        ("Handle Install", "handle_install"),
        ("Assembly Units", "assembly_units"), ("Install Units", "install_units"),
        ("Margin %", "margin_pct"),
    ]),
    ("PlanTemplates", "PlanTemplates", PlanTemplateItem, [
        ("ID", "id"), ("Division", "division"), ("Plan", "plan"), ("SKU", "sku"),
        ("Qty", "qty"), ("Area", "area"), ("Doors", "doors"), ("Drawers", "drawers"),
    ]),
    ("PlanTops", "PlanTops", PlanTops, [
        ("ID", "id"), ("Division", "division"), ("Plan", "plan"), ("Job ID", "job_id"),
        ("Material", "material"),
        ("Rate SqFt", "rate_sqft"), ("K Sinks", "k_sinks"), ("V Sinks", "v_sinks"),
    ]),
    ("TopPieces", "TopPieces", TopPiece, [
        ("ID", "id"), ("Tops ID", "tops_id"), ("Area", "area"), ("Kind", "kind"),
        ("Qty", "qty"), ("Width", "width"), ("Depth", "depth"),
    ]),
    ("Snapshots", "Snapshots", PriceSnapshot, [
        ("ID", "id"), ("Division", "division"), ("Door Style", "door_style"),
        ("Label", "label"), ("Created", "created_at"),
    ]),
    ("SnapshotRows", "SnapshotRows", PriceSnapshotRow, [
        ("ID", "id"), ("Snapshot ID", "snapshot_id"), ("Division", "division"),
        ("Plan", "plan"), ("COGS", "cogs"), ("Margin %", "margin_pct"),
        ("Sale", "sale"), ("Tops", "tops"), ("Total", "total"),
    ]),
    ("PlanBids", "PlanBids", PlanBid, [
        ("ID", "id"), ("Builder", "builder"), ("Plan", "plan"),
        ("Bid Price", "bid_price"), ("Notes", "notes"),
    ]),
    ("Jobs", "Jobs", Job, [
        ("ID", "id"), ("Name", "name"), ("Builder", "builder"), ("Community", "community"),
        ("Lot #", "lot_number"), ("Address", "address"), ("Plan", "plan"),
        ("Job Type", "job_type"), ("Stage", "stage"), ("Pricing Mode", "pricing_mode"),
        ("Margin %", "margin_pct"), ("Plan Price", "plan_price"),
        ("Multiplier", "multiplier_override"), ("Freight %", "freight_pct_override"),
        ("Assembly Boxes", "assembly_boxes"), ("Install Rate", "install_rate"),
        ("Hardware Rate", "hardware_rate"), ("Assembly Rate", "assembly_rate"),
        ("PIA", "pia_amount"), ("KSR", "ksr"),
        ("Cost Model", "cost_model"), ("Install Mode", "install_mode"), ("Install Price", "install_price"),
        ("Hardware SKU", "hardware_sku"), ("Hardware Qty", "hardware_qty_override"),
        ("Sales Contact", "sales_contact_name"), ("Sales Phone", "sales_contact_phone"),
        ("Sales Email", "sales_contact_email"), ("Field Contact", "field_contact_name"),
        ("Field Phone", "field_contact_phone"), ("Field Email", "field_contact_email"),
        ("Notes", "notes"), ("CT Job ID", "exported_job_id"), ("Exported At", "exported_at"),
        ("Created", "created_at"), ("Updated", "updated_at"),
    ]),
    ("Rooms", "Rooms", Room, [
        ("ID", "id"), ("Job ID", "job_id"), ("Room", "name"), ("Zone", "zone"),
        ("Brand", "cabinet_brand"), ("Series", "series"), ("Door Style", "door_style"),
        ("Finish", "finish"), ("Wood Species", "wood_species"), ("Notes", "notes"),
    ]),
    ("LineItems", "LineItems", LineItem, [
        ("ID", "id"), ("Room ID", "room_id"), ("SKU", "sku"), ("Description", "description"),
        ("Qty", "qty"), ("List Price", "list_price"), ("Multiplier", "multiplier"),
        ("Notes", "notes"),
    ]),
]

# Brian's standard table styling (Carter green headers for this app).
HEADER_FILL = PatternFill("solid", fgColor="125952")
HEADER_FONT = Font(name="Arial", size=9, bold=False, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="595959")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_ENUMS = {
    "job_type": JobType, "stage": Stage,
    "pricing_mode": PricingMode, "install_mode": InstallMode, "cost_model": CostModel,
}

state = {
    "dirty": False,
    "locked": False,       # True while the workbook is open in Excel and blocking saves
    "last_saved": None,    # ISO string
    "last_error": None,
    "suppress": False,     # True while loading the workbook into SQLite
}
_lock = threading.Lock()
_timer: threading.Timer | None = None


def _cell_value(obj, attr):
    v = getattr(obj, attr)
    if v is None:
        return None
    if isinstance(v, enum.Enum):
        return v.value
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.astimezone(timezone.utc).replace(tzinfo=None) if v.tzinfo else v
    return v


def _build_workbook(db) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, table_name, model, cols in SHEETS:
        ws = wb.create_sheet(sheet)
        ws.append([h for h, _ in cols])
        rows = db.query(model).order_by(*model.__table__.primary_key.columns).all()
        for obj in rows:
            ws.append([_cell_value(obj, attr) for _, attr in cols])
        last_row = max(2, ws.max_row)
        ref = f"A1:{get_column_letter(len(cols))}{last_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.add_table(table)
        widths = [len(h) for h, _ in cols]
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value is not None:
                    widths[i] = max(widths[i], len(str(cell.value)))
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(30, max(10, w + 2))
        ws.freeze_panes = "A2"
    return wb


def _daily_backup():
    """First save of the day copies the workbook to backups/ (30-day retention)."""
    if not WORKBOOK_PATH.exists():
        return
    bdir = WORKBOOK_PATH.parent / "backups"
    bdir.mkdir(exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    target = bdir / f"CKB Pricing Data {today}.xlsx"
    if target.exists():
        return
    try:
        shutil.copyfile(WORKBOOK_PATH, target)
        cutoff = datetime.now().timestamp() - 30 * 86400
        for old in bdir.glob("CKB Pricing Data *.xlsx"):
            if old.stat().st_mtime < cutoff:
                old.unlink(missing_ok=True)
    except OSError:
        pass  # backup must never block a save


def save_now() -> bool:
    """Write the workbook. Returns True on success; on Excel lock keeps dirty."""
    with _lock:
        _daily_backup()
        db = SessionLocal()
        try:
            wb = _build_workbook(db)
        finally:
            db.close()
        try:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", dir=WORKBOOK_PATH.parent, delete=False
            ) as tmp:
                tmp_path = Path(tmp.name)
            wb.save(tmp_path)
            try:
                tmp_path.replace(WORKBOOK_PATH)
            except PermissionError:
                tmp_path.unlink(missing_ok=True)
                raise
            state.update(
                dirty=False, locked=False, last_error=None,
                last_saved=datetime.now(timezone.utc).isoformat(),
            )
            return True
        except PermissionError:
            state.update(
                locked=True,
                last_error="Workbook is open in Excel — close it so changes can save",
            )
            _schedule(5.0)  # retry until Excel releases the file
            return False
        except Exception as exc:  # keep serving; surface the error
            state.update(last_error=f"Save failed: {exc}")
            return False


def _schedule(delay: float):
    global _timer
    if _timer:
        _timer.cancel()
    _timer = threading.Timer(delay, save_now)
    _timer.daemon = True
    _timer.start()


def mark_dirty():
    """Called after every DB commit — debounce then write the workbook."""
    if state["suppress"]:
        return
    state["dirty"] = True
    _schedule(1.5)


def _parse(attr, v):
    if v is None or v == "":
        return None
    if attr in _ENUMS:
        text = str(v).strip()
        if "." in text:  # legacy "InstallMode.with_knobs" form from an earlier save
            text = text.split(".", 1)[1]
        return _ENUMS[attr](text)
    if attr in (
        "list_price", "multiplier", "bid_price", "margin_pct", "plan_price",
        "price_g1", "price_g2", "price_g3", "price_g4", "price_g5",
        "cabinet_install", "knob_install", "handle_install", "install_price",
        "multiplier_override", "freight_pct_override", "install_rate",
        "hardware_rate", "assembly_rate", "pia_amount",
        "rate_sqft", "width", "depth", "cogs", "sale", "tops", "total",
    ):
        return Decimal(str(v))
    if attr in (
        "id", "job_id", "room_id", "qty", "exported_job_id",
        "price_group", "doors", "drawers", "hardware_qty_override",
        "assembly_units", "install_units", "assembly_boxes",
        "assemble_value", "install_value",
        "tops_id", "k_sinks", "v_sinks", "snapshot_id",
    ):
        return int(v)
    if attr in ("exported_at", "created_at", "updated_at"):
        if isinstance(v, datetime):
            return v.replace(tzinfo=timezone.utc)
        return datetime.fromisoformat(str(v))
    return str(v).strip() if isinstance(v, str) else v


def load_into_db() -> dict:
    """Replace SQLite contents with the workbook's. Returns row counts."""
    if not WORKBOOK_PATH.exists():
        return {}
    # Read from a copy so an Excel lock never blocks (and we see a consistent file).
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    shutil.copyfile(WORKBOOK_PATH, tmp_path)
    wb = None
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        counts = {}
        state["suppress"] = True
        db = SessionLocal()
        try:
            # Children first on delete; parents first on insert.
            for _, _, model, _ in reversed(SHEETS):
                db.query(model).delete()
            for sheet, _, model, cols in SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if header is None:
                    continue
                # Map by header text so column reordering in Excel is harmless.
                by_header = {h: attr for h, attr in cols}
                attrs = [by_header.get(str(h).strip() if h else "") for h in header]
                required = {
                    Setting: "value", CatalogItem: "sku", PlanBid: "builder",
                    Job: "name", Room: "name", LineItem: "sku",
                    DoorStyle: "name", PlanInstall: "plan", PlanTemplateItem: "sku",
                    PlanTops: "plan", TopPiece: "area",
                    PriceSnapshot: "division", PriceSnapshotRow: "plan",
                }[model]
                n = 0
                for row in rows:
                    # Blank cells are omitted so model defaults apply on insert.
                    data = {}
                    for attr, v in zip(attrs, row):
                        if attr is None:
                            continue
                        parsed = _parse(attr, v)
                        if parsed is not None:
                            data[attr] = parsed
                    key_field = "key" if model is Setting else "id"
                    if data.get(key_field) in (None, "") or not data.get(required):
                        continue  # blank or garbage row
                    db.add(model(**data))
                    n += 1
                counts[sheet] = n
            db.commit()
        finally:
            db.close()
            state["suppress"] = False
        return counts
    finally:
        if wb is not None:
            wb.close()
        tmp_path.unlink(missing_ok=True)


def startup():
    """Workbook exists -> it is the source of truth. Otherwise bootstrap it."""
    if not WORKBOOK_PATH.exists() and SEED_PATH != WORKBOOK_PATH and SEED_PATH.exists():
        WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(SEED_PATH, WORKBOOK_PATH)  # first boot on a fresh disk
    if WORKBOOK_PATH.exists():
        load_into_db()
    else:
        save_now()
