import io
import os
import re
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload, selectinload

from app.sterling_app import compute, schemas
from app.sterling_app.database import get_db
from app.sterling_app.export_cabinettron import ExportError, export_job
from app.sterling_app.models import (
    CatalogItem,
    CoverSheet,
    CoverSheetPO,
    CoverVendor,
    DoorStyle,
    Job,
    LineItem,
    PlanBid,
    PlanInstall,
    PlanTemplateItem,
    Room,
    Setting,
    Stage,
    Superintendent,
)

router = APIRouter(prefix="/api")


def _sku(s: str | None) -> str | None:
    """SKUs are always stored uppercase, however they're typed."""
    return s.strip().upper() if s else s


def _get_or_404(db: Session, model, obj_id: int, label: str):
    obj = db.get(model, obj_id)
    if obj is None:
        raise HTTPException(status_code=404, detail=f"{label} not found")
    return obj


# --- Job documents: files kept on disk under docs/job_<id>/<category>/ (survives DB rebuilds) ---

DOCS_ROOT = Path(__file__).resolve().parent.parent / "docs"
# "2020" is the on-disk name for the 20/20 folder (slashes are illegal in paths)
DOC_CATEGORIES = ["Pricing", "Layouts", "Plans", "2020", "Contract"]


def _doc_dir(job_id: int, category: str) -> Path:
    if category not in DOC_CATEGORIES:
        raise HTTPException(status_code=422, detail=f"Category must be one of {DOC_CATEGORIES}")
    return DOCS_ROOT / f"job_{job_id}" / category


def _safe_doc_name(name: str) -> str:
    name = Path(name or "document").name
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip().strip(".")
    return name or "document"


def _save_job_doc(job_id: int, category: str, filename: str | None, data: bytes) -> str:
    d = _doc_dir(job_id, category)
    d.mkdir(parents=True, exist_ok=True)
    name = _safe_doc_name(filename or "document")
    if (d / name).exists():
        stem, ext = Path(name).stem, Path(name).suffix
        i = 2
        while (d / f"{stem} ({i}){ext}").exists():
            i += 1
        name = f"{stem} ({i}){ext}"
    (d / name).write_bytes(data)
    return name


# --- Settings ---

@router.get("/settings", response_model=schemas.SettingsOut)
def get_settings(db: Session = Depends(get_db)):
    return schemas.SettingsOut(
        default_multiplier=compute.default_multiplier(db),
        default_margin_pct=compute.default_margin_pct(db),
        cabinettron_url=compute.get_setting(
            db, "cabinettron_url", os.environ.get("CABINETTRON_URL", "http://127.0.0.1:8000")
        ),
        cabinettron_email=compute.get_setting(db, "cabinettron_email", ""),
        cabinettron_password_set=bool(compute.get_setting(db, "cabinettron_password", "")),
        ksr_list=compute.get_setting(db, "ksr_list", ""),
        national_margin=compute.matrix_rate(db, "national_margin"),
    )


@router.put("/settings", response_model=schemas.SettingsOut)
def update_settings(payload: schemas.SettingsUpdate, db: Session = Depends(get_db)):
    if payload.default_multiplier is not None:
        compute.set_setting(db, "default_multiplier", str(payload.default_multiplier))
    if payload.default_margin_pct is not None:
        compute.set_setting(db, "default_margin_pct", str(payload.default_margin_pct))
    if payload.cabinettron_url is not None:
        compute.set_setting(db, "cabinettron_url", payload.cabinettron_url.strip())
    if payload.cabinettron_email is not None:
        compute.set_setting(db, "cabinettron_email", payload.cabinettron_email.strip())
    if payload.cabinettron_password:
        compute.set_setting(db, "cabinettron_password", payload.cabinettron_password)
    if payload.ksr_list is not None:
        compute.set_setting(db, "ksr_list", payload.ksr_list.strip())
    if payload.national_margin is not None:
        compute.set_setting(db, "national_margin", str(payload.national_margin))
    db.commit()
    return get_settings(db)


# --- Catalog ---

@router.get("/catalog", response_model=list[schemas.CatalogOut])
def list_catalog(q: str | None = None, vendor: str | None = None, db: Session = Depends(get_db)):
    query = db.query(CatalogItem)
    if vendor:
        query = query.filter(CatalogItem.vendor == vendor)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(CatalogItem.sku.ilike(like), CatalogItem.description.ilike(like))
        )
    return query.order_by(CatalogItem.vendor, CatalogItem.sku).limit(3000).all()


@router.post("/catalog", response_model=schemas.CatalogOut, status_code=201)
def create_catalog_item(payload: schemas.CatalogCreate, db: Session = Depends(get_db)):
    payload.sku = _sku(payload.sku)
    clash = (
        db.query(CatalogItem)
        .filter(CatalogItem.vendor == payload.vendor, CatalogItem.sku == payload.sku)
        .first()
    )
    if clash:
        raise HTTPException(status_code=409, detail=f"{payload.vendor} already has SKU {payload.sku}")
    item = CatalogItem(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/catalog/{item_id}", response_model=schemas.CatalogOut)
def update_catalog_item(item_id: int, payload: schemas.CatalogUpdate, db: Session = Depends(get_db)):
    item = _get_or_404(db, CatalogItem, item_id, "Catalog item")
    if payload.sku:
        payload.sku = _sku(payload.sku)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/catalog/{item_id}", status_code=204)
def delete_catalog_item(item_id: int, db: Session = Depends(get_db)):
    db.delete(_get_or_404(db, CatalogItem, item_id, "Catalog item"))
    db.commit()


CATALOG_HEADER_ALIASES = {
    "sku": {"sku", "item", "item code", "itemcode", "code", "nomenclature", "item#", "item #"},
    "description": {"description", "desc", "item description", "name"},
    "list_price": {"list price", "list", "listprice", "price", "msrp", "list_price"},
    "vendor": {"vendor", "supplier", "brand", "manufacturer"},
    "category": {"category", "cat", "type", "group"},
    "multiplier": {"multiplier", "mult", "mx"},
}


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = str(header or "").strip().lower()
        for field, aliases in CATALOG_HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


@router.post("/catalog/import")
async def import_catalog(file: UploadFile, vendor: str = "Everluxe", db: Session = Depends(get_db)):
    """Import SKUs from .xlsx or .csv. Needs at least SKU + list price columns."""
    raw = await file.read()
    name = (file.filename or "").lower()
    rows: list[list] = []
    if name.endswith(".csv"):
        import csv

        text = raw.decode("utf-8-sig", errors="replace")
        rows = [row for row in csv.reader(io.StringIO(text))]
    elif name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise HTTPException(status_code=422, detail="Upload a .xlsx or .csv file")

    if not rows:
        raise HTTPException(status_code=422, detail="File is empty")

    # Find the header row within the first 10 rows (price sheets often have title rows).
    header_idx, mapping = None, {}
    for i, row in enumerate(rows[:10]):
        m = _map_headers([str(c) if c is not None else "" for c in row])
        if "sku" in m and "list_price" in m:
            header_idx, mapping = i, m
            break
    if header_idx is None:
        raise HTTPException(
            status_code=422,
            detail="Could not find SKU + List Price columns in the first 10 rows",
        )

    added = updated = skipped = 0
    for row in rows[header_idx + 1:]:
        def cell(field):
            idx = mapping.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            return str(row[idx]).strip()

        sku = cell("sku")
        sku = sku.upper() if sku else sku
        raw_price = cell("list_price")
        if not sku or raw_price in (None, ""):
            skipped += 1
            continue
        try:
            price = Decimal(raw_price.replace("$", "").replace(",", ""))
        except InvalidOperation:
            skipped += 1
            continue
        row_vendor = cell("vendor") or vendor
        item = (
            db.query(CatalogItem)
            .filter(CatalogItem.vendor == row_vendor, CatalogItem.sku == sku)
            .first()
        )
        mult = None
        if cell("multiplier"):
            try:
                mult = Decimal(cell("multiplier"))
            except InvalidOperation:
                mult = None
        if item:
            item.list_price = price
            if cell("description"):
                item.description = cell("description")
            if cell("category"):
                item.category = cell("category")
            if mult is not None:
                item.multiplier = mult
            updated += 1
        else:
            db.add(
                CatalogItem(
                    sku=sku,
                    description=cell("description"),
                    vendor=row_vendor,
                    category=cell("category"),
                    list_price=price,
                    multiplier=mult,
                )
            )
            added += 1
    db.commit()
    return {"added": added, "updated": updated, "skipped": skipped}


# --- Plan bids ---

@router.get("/plan-bids", response_model=list[schemas.PlanBidOut])
def list_plan_bids(q: str | None = None, db: Session = Depends(get_db)):
    query = db.query(PlanBid)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(or_(PlanBid.builder.ilike(like), PlanBid.plan.ilike(like)))
    return query.order_by(PlanBid.builder, PlanBid.plan).all()


@router.post("/plan-bids", response_model=schemas.PlanBidOut, status_code=201)
def create_plan_bid(payload: schemas.PlanBidCreate, db: Session = Depends(get_db)):
    clash = (
        db.query(PlanBid)
        .filter(PlanBid.builder == payload.builder, PlanBid.plan == payload.plan)
        .first()
    )
    if clash:
        raise HTTPException(status_code=409, detail=f"{payload.builder} / {payload.plan} already has a bid")
    bid = PlanBid(**payload.model_dump())
    db.add(bid)
    db.commit()
    db.refresh(bid)
    return bid


@router.put("/plan-bids/{bid_id}", response_model=schemas.PlanBidOut)
def update_plan_bid(bid_id: int, payload: schemas.PlanBidUpdate, db: Session = Depends(get_db)):
    bid = _get_or_404(db, PlanBid, bid_id, "Plan bid")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(bid, field, value)
    db.commit()
    db.refresh(bid)
    return bid


@router.delete("/plan-bids/{bid_id}", status_code=204)
def delete_plan_bid(bid_id: int, db: Session = Depends(get_db)):
    db.delete(_get_or_404(db, PlanBid, bid_id, "Plan bid"))
    db.commit()


# --- Jobs ---

def _job_query(db: Session):
    return db.query(Job).options(selectinload(Job.rooms).selectinload(Room.lines))


@router.get("/jobs", response_model=list[schemas.JobListItem])
def list_jobs(
    stage: Stage | None = None,
    q: str | None = None,
    ksr: str | None = None,
    db: Session = Depends(get_db),
):
    query = _job_query(db)
    if stage is not None:
        query = query.filter(Job.stage == stage)
    if ksr:
        query = query.filter(Job.ksr == ksr)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Job.name.ilike(like),
                Job.builder.ilike(like),
                Job.community.ilike(like),
                Job.lot_number.ilike(like),
                Job.plan.ilike(like),
            )
        )
    jobs = query.order_by(Job.updated_at.desc()).limit(1000).all()
    return [compute.job_list_item(db, job) for job in jobs]


@router.post("/jobs", response_model=schemas.JobDetail, status_code=201)
def create_job(payload: schemas.JobCreate, db: Session = Depends(get_db)):
    payload.hardware_sku = _sku(payload.hardware_sku)
    job = Job(**payload.model_dump())
    db.add(job)
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


@router.get("/jobs/{job_id}", response_model=schemas.JobDetail)
def get_job(job_id: int, db: Session = Depends(get_db)):
    return compute.job_detail(db, _get_or_404(db, Job, job_id, "Job"))


@router.put("/jobs/{job_id}", response_model=schemas.JobDetail)
def update_job(job_id: int, payload: schemas.JobUpdate, db: Session = Depends(get_db)):
    job = _get_or_404(db, Job, job_id, "Job")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(job, field, _sku(value) if field == "hardware_sku" else value)
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


@router.delete("/jobs/{job_id}", status_code=204)
def delete_job(job_id: int, db: Session = Depends(get_db)):
    db.delete(_get_or_404(db, Job, job_id, "Job"))
    db.commit()


@router.post("/jobs/{job_id}/duplicate", response_model=schemas.JobDetail, status_code=201)
def duplicate_job(job_id: int, db: Session = Depends(get_db)):
    """Copy a job with all rooms + lines — for repeat plans and Option B pricing."""
    src = _get_or_404(db, Job, job_id, "Job")
    copy = Job(
        name=f"{src.name} (copy)",
        builder=src.builder,
        community=src.community,
        lot_number=src.lot_number,
        address=src.address,
        plan=src.plan,
        job_type=src.job_type,
        stage=Stage.lead,
        pricing_mode=src.pricing_mode,
        margin_pct=src.margin_pct,
        plan_price=src.plan_price,
        multiplier_override=src.multiplier_override,
        freight_pct_override=src.freight_pct_override,
        assembly_boxes=src.assembly_boxes,
        cost_model=src.cost_model,
        install_mode=src.install_mode,
        install_price=src.install_price,
        hardware_sku=src.hardware_sku,
        hardware_qty_override=src.hardware_qty_override,
        sales_contact_name=src.sales_contact_name,
        sales_contact_phone=src.sales_contact_phone,
        sales_contact_email=src.sales_contact_email,
        field_contact_name=src.field_contact_name,
        field_contact_phone=src.field_contact_phone,
        field_contact_email=src.field_contact_email,
        notes=src.notes,
    )
    for room in src.rooms:
        new_room = Room(
            name=room.name,
            zone=room.zone,
            cabinet_brand=room.cabinet_brand,
            series=room.series,
            door_style=room.door_style,
            finish=room.finish,
            wood_species=room.wood_species,
            notes=room.notes,
        )
        for line in room.lines:
            new_room.lines.append(
                LineItem(
                    sku=line.sku,
                    description=line.description,
                    qty=line.qty,
                    list_price=line.list_price,
                    multiplier=line.multiplier,
                    notes=line.notes,
                )
            )
        copy.rooms.append(new_room)
    db.add(copy)
    db.commit()
    db.refresh(copy)
    return compute.job_detail(db, copy)


# --- Rooms ---

@router.post("/jobs/{job_id}/rooms", response_model=schemas.JobDetail, status_code=201)
def add_room(job_id: int, payload: schemas.RoomCreate, db: Session = Depends(get_db)):
    job = _get_or_404(db, Job, job_id, "Job")
    job.rooms.append(Room(**payload.model_dump()))
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


@router.put("/rooms/{room_id}", response_model=schemas.JobDetail)
def update_room(room_id: int, payload: schemas.RoomUpdate, db: Session = Depends(get_db)):
    room = _get_or_404(db, Room, room_id, "Room")
    data = payload.model_dump(exclude_unset=True)
    style_changed = "door_style" in data and data["door_style"] != room.door_style
    for field, value in data.items():
        setattr(room, field, value)
    if style_changed:
        # Re-price catalog lines to the new door style's price group.
        group = compute.resolve_price_group(db, room.door_style)
        for line in room.lines:
            match = (
                db.query(CatalogItem)
                .filter(CatalogItem.vendor == "Everluxe", CatalogItem.sku.ilike(line.sku.strip()))
                .first()
            )
            if match:
                line.list_price = compute.group_price(match, group)
    db.commit()
    return compute.job_detail(db, room.job)


@router.delete("/rooms/{room_id}", response_model=schemas.JobDetail)
def delete_room(room_id: int, db: Session = Depends(get_db)):
    room = _get_or_404(db, Room, room_id, "Room")
    job = room.job
    db.delete(room)
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


# --- Line items ---

@router.post("/rooms/{room_id}/lines", response_model=schemas.JobDetail, status_code=201)
def add_line(room_id: int, payload: schemas.LineCreate, db: Session = Depends(get_db)):
    room = _get_or_404(db, Room, room_id, "Room")
    data = payload.model_dump()
    data["sku"] = _sku(data["sku"])
    # No price given -> pull from the catalog by SKU, using the room's door
    # style to pick the Everluxe price group (G1-G5); list_price is the fallback.
    # Prefer priced rows so a $0 trim/placeholder row never shadows a real one.
    if data["list_price"] is None:
        match = (
            db.query(CatalogItem)
            .filter(CatalogItem.sku.ilike(data["sku"].strip()))
            .order_by(CatalogItem.list_price == 0, CatalogItem.id)
            .first()
        )
        if match:
            group = compute.resolve_price_group(db, room.door_style)
            data["list_price"] = compute.group_price(match, group)
            if data["multiplier"] is None and match.multiplier is not None:
                data["multiplier"] = match.multiplier
            if not data["description"]:
                data["description"] = match.description
        else:
            data["list_price"] = Decimal("0")
    room.lines.append(LineItem(**data))
    db.commit()
    return compute.job_detail(db, room.job)


@router.put("/lines/{line_id}", response_model=schemas.JobDetail)
def update_line(line_id: int, payload: schemas.LineUpdate, db: Session = Depends(get_db)):
    line = _get_or_404(db, LineItem, line_id, "Line item")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(line, field, _sku(value) if field == "sku" else value)
    db.commit()
    return compute.job_detail(db, line.room.job)


@router.delete("/lines/{line_id}", response_model=schemas.JobDetail)
def delete_line(line_id: int, db: Session = Depends(get_db)):
    line = _get_or_404(db, LineItem, line_id, "Line item")
    job = line.room.job
    db.delete(line)
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


# --- Door styles / install rates / plan templates (DRH workbook data) ---

@router.get("/door-styles", response_model=list[schemas.DoorStyleOut])
def list_door_styles(db: Session = Depends(get_db)):
    return db.query(DoorStyle).order_by(DoorStyle.price_group, DoorStyle.name).all()


@router.get("/plan-installs", response_model=list[schemas.PlanInstallOut])
def list_plan_installs(q: str | None = None, db: Session = Depends(get_db)):
    query = db.query(PlanInstall)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(or_(PlanInstall.plan.ilike(like), PlanInstall.division.ilike(like)))
    return query.order_by(PlanInstall.division, PlanInstall.plan).all()


@router.get("/plan-templates", response_model=list[schemas.PlanTemplateSummary])
def list_plan_templates(db: Session = Depends(get_db)):
    from sqlalchemy import func

    rows = (
        db.query(
            PlanTemplateItem.division,
            PlanTemplateItem.plan,
            func.count(PlanTemplateItem.id),
            func.sum(PlanTemplateItem.qty),
        )
        .group_by(PlanTemplateItem.division, PlanTemplateItem.plan)
        .order_by(PlanTemplateItem.division, PlanTemplateItem.plan)
        .all()
    )
    return [
        schemas.PlanTemplateSummary(
            division=d, plan=p, line_count=c, total_qty=int(q or 0)
        )
        for d, p, c, q in rows
    ]


@router.post("/jobs/from-plan", response_model=schemas.JobDetail, status_code=201)
def create_job_from_plan(payload: schemas.JobFromPlan, db: Session = Depends(get_db)):
    """New job pre-filled from a builder plan's standard SKU list."""
    items = (
        db.query(PlanTemplateItem)
        .filter(
            PlanTemplateItem.division == payload.division,
            PlanTemplateItem.plan == payload.plan,
        )
        .order_by(PlanTemplateItem.id)
        .all()
    )
    if not items:
        raise HTTPException(status_code=404, detail="Plan template not found")

    job = Job(
        name=payload.name or f"{payload.plan} — new",
        builder=payload.builder or payload.division,
        community=payload.community,
        lot_number=payload.lot_number,
        plan=payload.plan,
        cost_model="matrix",  # DRH plan jobs use the full pricing-matrix cost model
        install_mode="with_knobs",
        # National tier, editable per quote on the job page.
        multiplier_override=compute.matrix_rate(db, "drh_multiplier"),
        freight_pct_override=compute.matrix_rate(db, "freight_pct"),
    )
    group = compute.resolve_price_group(db, payload.door_style)
    catalog = {
        i.sku: i
        for i in db.query(CatalogItem).filter(CatalogItem.vendor == "Everluxe").all()
    }
    # One room per area; the common "All" area becomes a single Kitchen room.
    by_area: dict[str, list[PlanTemplateItem]] = {}
    for item in items:
        area = (item.area or "All").strip()
        by_area.setdefault("Kitchen" if area.lower() == "all" else area, []).append(item)
    for area, area_items in by_area.items():
        room = Room(name=area, door_style=payload.door_style)
        for it in area_items:
            match = catalog.get(it.sku)
            room.lines.append(
                LineItem(
                    sku=it.sku,
                    description=match.description if match else None,
                    qty=it.qty,
                    list_price=compute.group_price(match, group) if match else Decimal("0"),
                    # Blank -> the job's per-quote multiplier; SKU-specific ones stick.
                    multiplier=match.multiplier if match and match.multiplier is not None else None,
                )
            )
        job.rooms.append(room)
    db.add(job)
    db.commit()
    db.refresh(job)
    return compute.job_detail(db, job)


@router.post("/import/drh-workbook")
def import_drh_workbook(path: str | None = None, db: Session = Depends(get_db)):
    from app.sterling_app.import_drh import import_workbook

    try:
        return import_workbook(db, path)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Import failed: {exc}")


# --- Dashboard ---

@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db)):
    jobs = _job_query(db).all()
    stages: dict[str, dict] = {
        s.value: {"count": 0, "sell": Decimal("0"), "margin": Decimal("0")} for s in Stage
    }
    for job in jobs:
        t = compute.job_totals(db, job)
        bucket = stages[job.stage.value]
        bucket["count"] += 1
        bucket["sell"] += t["sell"]
        bucket["margin"] += t["sell"] - t["cost"]
    active = [s.value for s in Stage if s not in (Stage.sold, Stage.lost)]
    pipeline_sell = sum((stages[s]["sell"] for s in active), Decimal("0"))
    sold = stages[Stage.sold.value]
    return {
        "stages": {k: {kk: str(vv) if isinstance(vv, Decimal) else vv for kk, vv in v.items()} for k, v in stages.items()},
        "pipeline_sell": str(pipeline_sell),
        "sold_sell": str(sold["sell"]),
        "sold_margin": str(sold["margin"]),
        "job_count": len(jobs),
        "catalog_count": db.query(CatalogItem).count(),
        "plan_bid_count": db.query(PlanBid).count(),
    }


INSTALL_HW_PATH = (
    r"C:\Users\Brian SE6\OneDrive - carterlumber.com"
    r"\Townsend Kitchen and Bath - Master Plans & Pricing\Pricing\AI Pricing Info"
    r"\Everluxe Install-Hardware Pricing 073126.xlsx"
)


@router.post("/import/install-hardware")
def import_install_hardware(path: str = INSTALL_HW_PATH, db: Session = Depends(get_db)):
    """Per-SKU labor units: Assemble/Install Value (= boxes) + Doors/Drawers
    (= hardware pieces per cabinet) from the Everluxe Install-Hardware file."""
    from pathlib import Path as _P

    from openpyxl import load_workbook

    if not _P(path).exists():
        raise HTTPException(status_code=404, detail=f"File not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    def _i(v):
        if v is None or str(v).strip() == "":
            return None
        try:
            return int(Decimal(str(v).strip()))
        except InvalidOperation:
            return None

    existing = {
        i.sku.upper(): i
        for i in db.query(CatalogItem).filter(CatalogItem.vendor == "Everluxe").all()
    }
    updated = created = 0
    for row in ws.iter_rows(min_row=3, values_only=True, max_col=6):
        sku = str(row[0] or "").strip().upper()
        if not sku:
            continue
        group = str(row[1] or "").strip() or None
        assemble, install = _i(row[2]), _i(row[3])
        doors, drawers = _i(row[4]), _i(row[5])
        item = existing.get(sku.upper())
        if item is None:
            item = CatalogItem(sku=sku, vendor="Everluxe", list_price=Decimal("0"))
            db.add(item)
            existing[sku.upper()] = item
            created += 1
        else:
            updated += 1
        item.install_group = group
        item.assemble_value = assemble
        item.install_value = install
        if doors is not None:
            item.doors = doors
        if drawers is not None:
            item.drawers = drawers
    wb.close()
    db.commit()
    return {"updated": updated, "created": created}


# --- National builder pricing (the workbook's Pricing Sheet, live) ---

@router.get("/national-pricing")
def national_pricing(
    division: str | None = None,
    door_style: str | None = None,
    db: Session = Depends(get_db),
):
    from app.sterling_app.models import PlanTemplateItem

    rows = compute.national_pricing_rows(db, division, door_style)
    divisions = [
        d for (d,) in db.query(PlanTemplateItem.division).distinct().order_by(PlanTemplateItem.division)
    ]
    return {
        "default_margin": str(compute.matrix_rate(db, "national_margin")),
        "divisions": divisions,
        "rows": [
            {k: (str(v) if isinstance(v, Decimal) else v) for k, v in r.items()} for r in rows
        ],
    }


class NationalMarginUpdate(BaseModel):
    division: str
    plan: str
    margin_pct: Decimal | None = Field(default=None, ge=0, lt=100)


@router.put("/national-pricing/margin")
def set_national_margin(payload: NationalMarginUpdate, db: Session = Depends(get_db)):
    rec = (
        db.query(PlanInstall)
        .filter(PlanInstall.division == payload.division, PlanInstall.plan == payload.plan)
        .first()
    )
    if rec is None:
        rec = PlanInstall(division=payload.division, plan=payload.plan)
        db.add(rec)
    rec.margin_pct = payload.margin_pct
    db.commit()
    return {"ok": True}


@router.post("/national-pricing/save-bids")
def save_national_bids(
    division: str | None = None,
    door_style: str | None = None,
    db: Session = Depends(get_db),
):
    """Write the computed sale prices into Plan Bids (builder = division)."""
    rows = compute.national_pricing_rows(db, division, door_style)
    saved = 0
    for r in rows:
        if r["sale"] <= 0:
            continue
        bid = (
            db.query(PlanBid)
            .filter(PlanBid.builder == r["division"], PlanBid.plan == r["plan"])
            .first()
        )
        if bid is None:
            bid = PlanBid(builder=r["division"], plan=r["plan"], bid_price=r["sale"])
            db.add(bid)
        else:
            bid.bid_price = r["sale"]
        bid.notes = (
            f"National pricing {door_style or 'base group'} @ {r['margin_pct']}% margin"
        )
        saved += 1
    db.commit()
    return {"saved": saved}


class TopPieceIn(BaseModel):
    area: str = Field(default="Kitchen", max_length=20)
    kind: str = Field(default="top", max_length=20)
    qty: int = Field(default=1, ge=1)
    width: Decimal = Field(default=Decimal("0"), ge=0)
    depth: Decimal = Field(default=Decimal("0"), ge=0)


class PlanTopsIn(BaseModel):
    division: str
    plan: str
    job_id: int | None = None  # set for job-level tops (division/plan become Jobs / Job {id})
    material: str = "Laminate"
    rate_sqft: Decimal | None = Field(default=None, ge=0)
    k_sinks: int = Field(default=0, ge=0)
    v_sinks: int = Field(default=0, ge=0)
    pieces: list[TopPieceIn] = []


@router.get("/national-pricing/tops")
def get_plan_tops(division: str, plan: str, db: Session = Depends(get_db)):
    from app.sterling_app.models import PlanTops

    t = (
        db.query(PlanTops)
        .filter(PlanTops.division == division, PlanTops.plan == plan)
        .first()
    )
    if t is None:
        return {"division": division, "plan": plan, "material": "Laminate",
                "rate_sqft": None, "k_sinks": 0, "v_sinks": 0, "pieces": [], "totals": None}
    totals = compute.tops_total(db, t)
    return {
        "division": t.division, "plan": t.plan, "material": t.material,
        "rate_sqft": str(t.rate_sqft) if t.rate_sqft is not None else None,
        "k_sinks": t.k_sinks, "v_sinks": t.v_sinks,
        "pieces": [
            {"area": p.area, "kind": p.kind, "qty": p.qty,
             "width": str(p.width), "depth": str(p.depth)} for p in t.pieces
        ],
        "totals": {k: str(v) for k, v in totals.items()},
    }


@router.put("/national-pricing/tops")
def save_plan_tops(payload: PlanTopsIn, db: Session = Depends(get_db)):
    from app.sterling_app.models import PlanTops, TopPiece

    if payload.job_id is not None:
        payload.division, payload.plan = "Jobs", f"Job {payload.job_id}"
    t = (
        db.query(PlanTops)
        .filter(PlanTops.division == payload.division, PlanTops.plan == payload.plan)
        .first()
    )
    if not payload.pieces and not payload.k_sinks and not payload.v_sinks:
        if t is not None:  # emptied out -> plan no longer prices tops
            db.delete(t)
            db.commit()
        return {"deleted": True}
    if t is None:
        t = PlanTops(division=payload.division, plan=payload.plan, job_id=payload.job_id)
        db.add(t)
    t.material = payload.material
    t.rate_sqft = payload.rate_sqft
    t.k_sinks = payload.k_sinks
    t.v_sinks = payload.v_sinks
    t.pieces = [TopPiece(**p.model_dump()) for p in payload.pieces]
    db.commit()
    totals = compute.tops_total(db, t)
    return {"totals": {k: str(v) for k, v in totals.items()}}


@router.post("/national-pricing/snapshot")
def create_snapshot(
    division: str | None = None,
    door_style: str | None = None,
    label: str | None = None,
    db: Session = Depends(get_db),
):
    """Freeze the current national pricing — the copy of what was sent."""
    from app.sterling_app.models import PriceSnapshot, PriceSnapshotRow

    rows = compute.national_pricing_rows(db, division, door_style)
    if not rows:
        raise HTTPException(status_code=422, detail="Nothing to snapshot")
    snap = PriceSnapshot(division=division or "All", door_style=door_style, label=label)
    for r in rows:
        snap.rows.append(PriceSnapshotRow(
            division=r["division"], plan=r["plan"], cogs=r["cogs"],
            margin_pct=r["margin_pct"], sale=r["sale"], tops=r["tops"], total=r["total"],
        ))
    db.add(snap)
    db.commit()
    return {"id": snap.id, "rows": len(snap.rows)}


@router.get("/national-pricing/snapshots")
def list_snapshots(db: Session = Depends(get_db)):
    from app.sterling_app.models import PriceSnapshot

    return [
        {"id": s.id, "division": s.division, "door_style": s.door_style,
         "label": s.label, "created_at": s.created_at.isoformat(), "rows": len(s.rows)}
        for s in db.query(PriceSnapshot).order_by(PriceSnapshot.id.desc()).all()
    ]


@router.get("/national-pricing/snapshots/{snap_id}")
def get_snapshot(snap_id: int, db: Session = Depends(get_db)):
    from app.sterling_app.models import PriceSnapshot

    s = _get_or_404(db, PriceSnapshot, snap_id, "Snapshot")
    return {
        "id": s.id, "division": s.division, "door_style": s.door_style,
        "label": s.label, "created_at": s.created_at.isoformat(),
        "rows": [
            {"division": r.division, "plan": r.plan, "cogs": str(r.cogs),
             "margin_pct": str(r.margin_pct), "sale": str(r.sale),
             "tops": str(r.tops), "total": str(r.total)}
            for r in s.rows
        ],
    }


# --- Everluxe SKU pricer + demand projections ---

class EvxLine(BaseModel):
    sku: str = Field(min_length=1, max_length=100)
    qty: int = Field(default=1, ge=1)
    area: str | None = Field(default=None, max_length=60)
    fin_end: str | None = Field(default=None, max_length=10)
    notes: str | None = Field(default=None, max_length=300)


class EvxPriceIn(BaseModel):
    door_style: str | None = None
    multiplier: Decimal = Field(default=Decimal("0.21"), gt=0, le=1)
    margin_pct: Decimal = Field(default=Decimal("15"), ge=0, lt=100)
    fuel_pct: Decimal = Field(default=Decimal("0"), ge=0, lt=100)  # optional fuel surcharge % of dealer cost
    hardware_sku: str | None = None  # 3910*=knob ($1 labor), 156*=handle ($2); material from the item
    lines: list[EvxLine]


@router.post("/everluxe/price")
def everluxe_price(payload: EvxPriceIn, db: Session = Depends(get_db)):
    """Price an ad-hoc Everluxe SKU list with the full matrix (like the workbook):
    list@group x mult + freight + knob hardware + tax + assembly + install -> sale."""
    group = compute.resolve_price_group(db, payload.door_style)
    freight_pct = (
        compute.matrix_rate(db, "freight_pct")
        if payload.multiplier <= Decimal("0.22")
        else compute.matrix_rate(db, "retail_freight_pct")
    )
    tax_pct = compute.matrix_rate(db, "tax_pct")
    knob_mat = compute.matrix_rate(db, "knob_material")
    knob_labor = compute.matrix_rate(db, "knob_labor")
    assem_rate = compute.matrix_rate(db, "assem_rate")
    install_rate = compute.matrix_rate(db, "install_rate")
    money = compute.money

    from decimal import ROUND_HALF_UP as _RH

    # Hardware: material from the chosen item; labor $1 knob / $2 handle by family
    hw_unit, hw_labor_rate = knob_mat, knob_labor
    hw_kind = compute.hardware_kind(payload.hardware_sku)
    if payload.hardware_sku:
        hw_item = (
            db.query(CatalogItem)
            .filter(CatalogItem.sku.ilike(payload.hardware_sku.strip()))
            .order_by(CatalogItem.list_price == 0, CatalogItem.id)
            .first()
        )
        if hw_item:
            hw_unit = money(
                Decimal(hw_item.list_price)
                * (Decimal(hw_item.multiplier) if hw_item.multiplier is not None else Decimal("1"))
            )
        if hw_kind == "handle":
            hw_labor_rate = compute.matrix_rate(db, "handle_labor")

    def price_subset(lines):
        """Full matrix over a set of lines -> (out_lines, totals dict)."""
        out, list_total = [], Decimal("0")
        hw_qty = boxes = units = 0
        for ln in lines:
            item = (
                db.query(CatalogItem)
                .filter(CatalogItem.vendor == "Everluxe", CatalogItem.sku.ilike(ln.sku.strip()))
                .order_by(CatalogItem.list_price == 0, CatalogItem.id)
                .first()
            )
            each = compute.group_price(item, group) if item else Decimal("0")
            ext = money(each * ln.qty)
            list_total += ext
            if item:
                hw_qty += (item.doors + item.drawers) * ln.qty
                boxes += (item.assemble_value or 0) * ln.qty
                units += (item.install_value or 0) * ln.qty
            out.append({
                "sku": ln.sku.strip().upper(), "qty": ln.qty, "found": item is not None,
                "description": item.description if item else None,
                "area": ln.area or "All", "fin_end": ln.fin_end, "notes": ln.notes,
                "list_each": str(each), "ext_list": str(ext),
                "dealer_each": str(money(each * payload.multiplier)),
                "dealer_ext": str(money(ext * payload.multiplier)),
            })
        cabinets = money(list_total * payload.multiplier)
        # Assembly ($10/box) belongs with the Everluxe materials; freight is on
        # the cabinet cost ONLY — never on assembly. Install is a flat $25/box.
        assembly = money(Decimal(boxes) * assem_rate)
        freight = money(cabinets * freight_pct)
        fuel = money(cabinets * payload.fuel_pct / 100)
        hw_material = money(hw_qty * hw_unit)
        hw_labor = money(hw_qty * hw_labor_rate)
        tax = money((cabinets + hw_material) * tax_pct)
        install = money(Decimal(units) * install_rate)
        cogs = money(cabinets + assembly + freight + fuel + hw_material + tax + hw_labor + install)
        sale = (
            money(compute.sell_from_margin(cogs, payload.margin_pct).quantize(Decimal("1"), rounding=_RH))
            if cogs else Decimal("0.00")
        )
        return out, {
            "list": str(money(list_total)), "cabinets": str(cabinets), "freight": str(freight),
            "freight_pct": str(freight_pct), "fuel": str(fuel), "fuel_pct": str(payload.fuel_pct),
            "hardware_qty": hw_qty, "hardware_material": str(hw_material),
            "hardware_labor": str(hw_labor), "hardware_kind": hw_kind or "knob",
            "hardware_labor_rate": str(hw_labor_rate), "tax": str(tax),
            "assembly_boxes": boxes, "assembly": str(assembly),
            "install_units": units, "install": str(install),
            "cogs": str(cogs), "margin_pct": str(payload.margin_pct), "sale": str(sale),
        }

    out_lines, totals = price_subset(payload.lines)
    # Per-area pricing: each area computed COMPLETELY separately (own rounding,
    # own whole-dollar sale). Grand = sum of area sales — can differ from merged.
    area_names = []
    for ln in payload.lines:
        a = (ln.area or "All").strip() or "All"
        if a not in area_names:
            area_names.append(a)
    areas = []
    for a in area_names:
        subset = [ln for ln in payload.lines if ((ln.area or "All").strip() or "All") == a]
        _, at = price_subset(subset)
        areas.append({"area": a, "totals": at})
    grand_sale = money(sum((Decimal(a["totals"]["sale"]) for a in areas), Decimal("0")))
    grand_cogs = money(sum((Decimal(a["totals"]["cogs"]) for a in areas), Decimal("0")))
    return {
        "lines": out_lines,
        "totals": totals,
        "areas": areas,
        "grand": {"sale": str(grand_sale), "cogs": str(grand_cogs)},
    }


@router.get("/everluxe/projections")
def everluxe_projections(stages: str | None = None, db: Session = Depends(get_db)):
    """Cabinet demand: total qty per SKU across job line items (by stage)."""
    wanted = [s.strip() for s in (stages or "Lead,Design,Quoted,Revision,Sold").split(",") if s.strip()]
    q = (
        db.query(LineItem, Room, Job)
        .join(Room, LineItem.room_id == Room.id)
        .join(Job, Room.job_id == Job.id)
        .filter(Job.stage.in_(wanted))
    )
    agg: dict[str, dict] = {}
    catalog = {
        i.sku.upper(): i
        for i in db.query(CatalogItem).filter(CatalogItem.vendor == "Everluxe").all()
    }
    for line, room, job in q.all():
        if room.name.strip().lower() == "lumber":
            continue
        key = line.sku.strip().upper()
        item = catalog.get(key)
        rec = agg.setdefault(key, {
            "sku": line.sku.strip(), "description": item.description if item else None,
            "in_catalog": item is not None, "qty": 0, "jobs": set(),
        })
        rec["qty"] += line.qty
        rec["jobs"].add(job.id)
    rows = sorted(agg.values(), key=lambda r: -r["qty"])
    for r in rows:
        r["job_count"] = len(r.pop("jobs"))
    return {"stages": wanted, "rows": rows}


# --- 2020 Design quote import ---

@router.post("/import/quote2020", response_model=schemas.JobDetail, status_code=201)
async def import_quote2020(
    file: UploadFile,
    name: str | None = None,
    customer: str | None = None,
    room: str | None = None,
    cabinet_brand: str | None = None,
    series: str | None = None,
    door_style: str | None = None,
    finish: str | None = None,
    wood_species: str | None = None,
    db: Session = Depends(get_db),
):
    """Upload a 2020 Design .xls quote -> new priced job (one room per section)."""
    from app.sterling_app.import_2020 import parse_2020_xls

    raw = await file.read()
    try:
        parsed = parse_2020_xls(raw)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read 2020 export: {exc}")
    if not parsed["items"]:
        raise HTTPException(status_code=422, detail="No item lines found in this file")

    base = (file.filename or "2020 Quote").rsplit(".", 1)[0]
    national = compute.is_national(customer)
    job = Job(
        name=name or base,
        builder=customer or None,
        job_type="tract" if national else "custom",
        cost_model="matrix",  # freight + sales tax apply; install/assembly added manually
        install_mode="none",
        multiplier_override=compute.matrix_rate(
            db, "drh_multiplier" if national else "retail_multiplier"
        ),
        freight_pct_override=compute.matrix_rate(
            db, "freight_pct" if national else "retail_freight_pct"
        ),
        notes=f"Imported from 2020 Design export '{file.filename}'."
        + (f" Design file: {parsed['file_name']}" if parsed["file_name"] else ""),
    )
    from app.sterling_app.import_2020 import combine_items

    # The Cabinets section takes the user's room name + selections; other
    # sections (Accessories, Mouldings) keep their 2020 names.
    cab_room_name = (room or "").strip() or "Cabinets"
    rooms: dict[str, Room] = {}
    for it in combine_items(parsed["items"]):
        section = it["section"]
        new_room = rooms.get(section)
        if new_room is None:
            if section == "Cabinets":
                new_room = Room(
                    name=cab_room_name,
                    cabinet_brand=cabinet_brand,
                    series=series,
                    door_style=door_style,
                    finish=finish,
                    wood_species=wood_species,
                )
            else:
                new_room = Room(name=section)
            rooms[section] = new_room
            job.rooms.append(new_room)
        match = (
            db.query(CatalogItem)
            .filter(CatalogItem.sku.ilike(it["sku"]))
            .order_by(CatalogItem.list_price == 0, CatalogItem.id)
            .first()
        )
        new_room.lines.append(
            LineItem(
                sku=_sku(it["sku"]),
                description=(match.description if match else None)
                or ("non-plan item" if it["non_plan"] else None),
                qty=it["qty"],
                list_price=it["list_each"],
                notes=f"2020 line {it['line_no']}",
            )
        )
    db.add(job)
    db.commit()
    db.refresh(job)
    _save_job_doc(job.id, "2020", file.filename or "2020 Quote.xls", raw)
    return compute.job_detail(db, job)


@router.post("/jobs/{job_id}/import-2020", response_model=schemas.JobDetail)
async def import_2020_into_job(
    job_id: int,
    file: UploadFile,
    room: str | None = None,
    cabinet_brand: str | None = None,
    series: str | None = None,
    door_style: str | None = None,
    finish: str | None = None,
    wood_species: str | None = None,
    db: Session = Depends(get_db),
):
    """Upload a 2020 quote (.xls or .pdf) INTO an existing job.

    The Cabinets section lands in the room named by `room` (with the given
    selections). Like items are combined — within the file, and with lines
    already on the job (same room + SKU + list each -> qty bumped).
    """
    from app.sterling_app.import_2020 import combine_items, parse_2020_pdf, parse_2020_xls

    job = _get_or_404(db, Job, job_id, "Job")
    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith(".pdf"):
            parsed = parse_2020_pdf(raw)
        elif name.endswith((".xls", ".xlsx")):
            parsed = parse_2020_xls(raw)
        else:
            raise HTTPException(status_code=422, detail="Upload a 2020 .xls or .pdf quote")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read 2020 file: {exc}")
    if not parsed["items"]:
        raise HTTPException(
            status_code=422,
            detail="No item lines found — if this is a scanned/drawing PDF, use the 2020 quote report",
        )

    cab_room_name = (room or "").strip() or "Cabinets"
    rooms = {r.name: r for r in job.rooms}
    added = merged = 0
    for it in combine_items(parsed["items"]):
        target_name = cab_room_name if it["section"] == "Cabinets" else it["section"]
        target = rooms.get(target_name)
        if target is None:
            target = Room(name=target_name)
            if it["section"] == "Cabinets":
                target.cabinet_brand = cabinet_brand
                target.series = series
                target.door_style = door_style
                target.finish = finish
                target.wood_species = wood_species
            rooms[target_name] = target
            job.rooms.append(target)
        existing = next(
            (
                l for l in target.lines
                if l.sku.upper() == it["sku"].upper() and l.list_price == it["list_each"]
            ),
            None,
        )
        if existing:
            existing.qty += it["qty"]
            merged += 1
        else:
            match = (
                db.query(CatalogItem)
                .filter(CatalogItem.sku.ilike(it["sku"]))
                .order_by(CatalogItem.list_price == 0, CatalogItem.id)
                .first()
            )
            target.lines.append(
                LineItem(
                    sku=_sku(it["sku"]),
                    description=(match.description if match else None)
                    or ("non-plan item" if it["non_plan"] else None),
                    qty=it["qty"],
                    list_price=it["list_each"],
                    notes=f"2020 import ({file.filename})",
                )
            )
            added += 1
    job.notes = ((job.notes + "\n") if job.notes else "") + f"2020 import: {file.filename}"
    db.commit()
    db.refresh(job)
    _save_job_doc(job.id, "2020", file.filename or "2020 Quote", raw)
    return compute.job_detail(db, job)


# --- Job documents ---

@router.get("/jobs/{job_id}/docs")
def list_job_docs(job_id: int, db: Session = Depends(get_db)):
    _get_or_404(db, Job, job_id, "Job")
    docs = []
    for category in DOC_CATEGORIES:
        d = DOCS_ROOT / f"job_{job_id}" / category
        if not d.is_dir():
            continue
        for p in sorted(d.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
            if p.is_file():
                docs.append(
                    {
                        "category": category,
                        "name": p.name,
                        "size": p.stat().st_size,
                        "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
                    }
                )
    return docs


@router.post("/jobs/{job_id}/docs", status_code=201)
async def upload_job_doc(
    job_id: int, file: UploadFile, category: str = "2020", db: Session = Depends(get_db)
):
    _get_or_404(db, Job, job_id, "Job")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="Empty file")
    return {"category": category, "name": _save_job_doc(job_id, category, file.filename, raw)}


def _doc_path_or_404(job_id: int, category: str, name: str) -> Path:
    if name != _safe_doc_name(name):  # blocks path traversal
        raise HTTPException(status_code=404, detail="Document not found")
    path = _doc_dir(job_id, category) / name
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Document not found")
    return path


@router.get("/jobs/{job_id}/docs/{category}/{name}")
def download_job_doc(job_id: int, category: str, name: str, db: Session = Depends(get_db)):
    _get_or_404(db, Job, job_id, "Job")
    path = _doc_path_or_404(job_id, category, name)
    return FileResponse(path, filename=path.name)


@router.delete("/jobs/{job_id}/docs/{category}/{name}", status_code=204)
def delete_job_doc(job_id: int, category: str, name: str, db: Session = Depends(get_db)):
    _get_or_404(db, Job, job_id, "Job")
    _doc_path_or_404(job_id, category, name).unlink()


# --- Sales Order Cover Sheet ---

COVER_FIELDS = [
    "job_code", "sale_date", "plan_type", "customer_account", "job_number", "install_code", "scope",
    "ji_name", "ji_contact", "ji_address", "ji_city", "ji_state", "ji_zip", "ji_phone", "ji_email",
    "cu_company", "cu_name", "cu_address", "cu_city", "cu_state", "cu_zip", "cu_phone", "cu_email",
    "super_name", "super_phone", "super_email", "notes",
]
COVER_MONEY = ["tax_pct", "sale_cabinets", "sale_countertops", "sale_other"]


class CoverPOIn(BaseModel):
    kind: str = Field(default="product", max_length=10)
    po_type: str | None = None
    po_abb: str | None = None
    vendor: str | None = None
    vendor_code: str | None = None
    amount1: Decimal = Field(default=Decimal("0"))
    amount2: Decimal = Field(default=Decimal("0"))
    total_override: Decimal | None = None


class CoverSheetIn(BaseModel):
    job_id: int | None = None
    job_code: str | None = None
    sale_date: str | None = None
    plan_type: str | None = None
    customer_account: str | None = None
    job_number: str | None = None
    install_code: str | None = None
    scope: str | None = None
    ji_name: str | None = None
    ji_contact: str | None = None
    ji_address: str | None = None
    ji_city: str | None = None
    ji_state: str | None = None
    ji_zip: str | None = None
    ji_phone: str | None = None
    ji_email: str | None = None
    cu_company: str | None = None
    cu_name: str | None = None
    cu_address: str | None = None
    cu_city: str | None = None
    cu_state: str | None = None
    cu_zip: str | None = None
    cu_phone: str | None = None
    cu_email: str | None = None
    super_name: str | None = None
    super_phone: str | None = None
    super_email: str | None = None
    notes: str | None = None
    tax_pct: Decimal = Field(default=Decimal("7"), ge=0, lt=100)   # sales tax is always 7%
    sale_cabinets: Decimal = Field(default=Decimal("0"), ge=0)
    sale_countertops: Decimal = Field(default=Decimal("0"), ge=0)
    sale_other: Decimal = Field(default=Decimal("0"), ge=0)
    pos: list[CoverPOIn] = []


def _cover_out(db: Session, s: CoverSheet) -> dict:
    money = compute.money
    rows = []
    materials = labor = Decimal("0")
    for p in s.pos:
        total = (
            Decimal(p.total_override)
            if p.total_override is not None
            else money(Decimal(p.amount1) + Decimal(p.amount2))
        )
        if p.kind == "labor":
            labor += total
        else:
            materials += total
        rows.append({
            "id": p.id, "kind": p.kind, "po_type": p.po_type, "po_abb": p.po_abb,
            "vendor": p.vendor, "vendor_code": p.vendor_code,
            "amount1": str(p.amount1), "amount2": str(p.amount2),
            "total_override": str(p.total_override) if p.total_override is not None else None,
            "total": str(money(total)),
            "po_number": f"{s.job_code or ''} {p.po_abb or ''}".strip(),
        })
    materials, labor = money(materials), money(labor)
    tax = money(materials * Decimal(s.tax_pct) / 100)
    cogs = money(materials + labor + tax)
    sale = money(Decimal(s.sale_cabinets) + Decimal(s.sale_countertops) + Decimal(s.sale_other))
    margin = money(sale - cogs)
    data = {f: getattr(s, f) for f in COVER_FIELDS}
    data.update({m: str(getattr(s, m)) for m in COVER_MONEY})
    data.update({
        "id": s.id, "job_id": s.job_id, "pos": rows,
        "totals": {
            "materials": str(materials), "labor": str(labor), "tax": str(tax), "cogs": str(cogs),
            "sale": str(sale), "margin": str(margin),
            "margin_pct": str(money(margin / sale * 100)) if sale else None,
        },
        "updated_at": s.updated_at.isoformat(),
    })
    return data


@router.get("/cover-sheets")
def list_cover_sheets(db: Session = Depends(get_db)):
    out = []
    for s in db.query(CoverSheet).order_by(CoverSheet.id.desc()).all():
        d = _cover_out(db, s)
        out.append({
            "id": s.id, "job_id": s.job_id, "job_code": s.job_code, "sale_date": s.sale_date,
            "ji_name": s.ji_name, "ji_address": s.ji_address, "cu_company": s.cu_company,
            "sale": d["totals"]["sale"], "margin_pct": d["totals"]["margin_pct"],
            "updated_at": d["updated_at"],
        })
    return out


@router.get("/cover-sheets/refs")
def cover_refs(customer: str | None = None, db: Session = Depends(get_db)):
    """PO-type presets, C-Codes, and superintendent suggestions.

    Superintendents we've actually worked with for this customer come first
    (from prior cover sheets), then the general roster.
    """
    seen, supers = set(), []
    if customer and customer.strip():
        like = f"%{customer.strip()}%"
        prior = (
            db.query(CoverSheet)
            .filter(CoverSheet.super_name.isnot(None), CoverSheet.super_name != "")
            .filter(or_(CoverSheet.cu_company.ilike(like), CoverSheet.cu_name.ilike(like)))
            .order_by(CoverSheet.id.desc())
            .all()
        )
        for s in prior:
            key = (s.super_name or "").strip().lower()
            if key and key not in seen:
                seen.add(key)
                supers.append({"name": s.super_name, "phone": s.super_phone,
                               "email": s.super_email, "source": "prior job"})
    for s in db.query(Superintendent).order_by(Superintendent.name).all():
        if s.name.strip().lower() not in seen:
            seen.add(s.name.strip().lower())
            supers.append({"name": s.name, "phone": s.phone, "email": s.email,
                           "source": s.company or "roster"})
    # Vendors we've actually used: newest first, one entry per name+code pair.
    used: dict[tuple[str, str], dict] = {}
    for p in (
        db.query(CoverSheetPO)
        .filter(CoverSheetPO.vendor.isnot(None), CoverSheetPO.vendor != "")
        .order_by(CoverSheetPO.id.desc())
        .all()
    ):
        key = (p.kind, (p.vendor or "").strip().lower())
        if key not in used:
            used[key] = {"kind": p.kind, "vendor": p.vendor.strip(),
                         "vendor_code": (p.vendor_code or "").strip() or None}
    return {
        "po_types": [
            {"kind": v.kind, "po_type": v.po_type, "po_abb": v.po_abb}
            for v in db.query(CoverVendor).order_by(CoverVendor.kind, CoverVendor.id).all()
        ],
        "used_vendors": list(used.values()),
        "superintendents": supers,
    }



@router.get("/cover-sheets/blank.xlsx")
def blank_cover_xlsx():
    """Empty fillable Excel cover sheet (one portrait page)."""
    from fastapi.responses import StreamingResponse

    from app.sterling_app.cover_xlsx import build_cover_workbook

    return StreamingResponse(
        build_cover_workbook(None),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Sales Order Cover Sheet.xlsx"'},
    )


@router.get("/cover-sheets/{sheet_id}")
def get_cover_sheet(sheet_id: int, db: Session = Depends(get_db)):
    return _cover_out(db, _get_or_404(db, CoverSheet, sheet_id, "Cover sheet"))


@router.get("/cover-sheets/{sheet_id}/xlsx")
def cover_sheet_xlsx(sheet_id: int, db: Session = Depends(get_db)):
    """This cover sheet as a fillable Excel workbook (formulas live)."""
    from fastapi.responses import StreamingResponse

    from app.sterling_app.cover_xlsx import build_cover_workbook

    s = _get_or_404(db, CoverSheet, sheet_id, "Cover sheet")
    data = _cover_out(db, s)
    name = f"Sales Order Cover Sheet {s.job_code or s.id}.xlsx".replace("/", "-")
    return StreamingResponse(
        build_cover_workbook(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.post("/cover-sheets", status_code=201)
def create_cover_sheet(payload: CoverSheetIn, db: Session = Depends(get_db)):
    data = payload.model_dump(exclude={"pos"})
    s = CoverSheet(**data)
    s.pos = [CoverSheetPO(**p.model_dump()) for p in payload.pos]
    db.add(s)
    db.commit()
    db.refresh(s)
    return _cover_out(db, s)


@router.put("/cover-sheets/{sheet_id}")
def update_cover_sheet(sheet_id: int, payload: CoverSheetIn, db: Session = Depends(get_db)):
    s = _get_or_404(db, CoverSheet, sheet_id, "Cover sheet")
    for field, value in payload.model_dump(exclude={"pos"}).items():
        setattr(s, field, value)
    s.pos = [CoverSheetPO(**p.model_dump()) for p in payload.pos]
    db.commit()
    db.refresh(s)
    return _cover_out(db, s)


@router.delete("/cover-sheets/{sheet_id}", status_code=204)
def delete_cover_sheet(sheet_id: int, db: Session = Depends(get_db)):
    db.delete(_get_or_404(db, CoverSheet, sheet_id, "Cover sheet"))
    db.commit()


@router.post("/cover-sheets/from-job/{job_id}", status_code=201)
def cover_sheet_from_job(job_id: int, db: Session = Depends(get_db)):
    """Prefill a cover sheet from a Sterling job: addresses, contacts, and the
    priced cabinets/tops/COGS so only the PO details need typing."""
    job = _get_or_404(db, Job, job_id, "Job")
    t = compute.job_totals(db, job)
    existing = db.query(CoverSheet).filter(CoverSheet.job_id == job_id).first()
    s = existing or CoverSheet(job_id=job_id)
    if existing is None:
        db.add(s)
    s.job_code = s.job_code or job.name
    s.plan_type = job.plan
    s.ji_name = job.name
    s.ji_address = job.address
    s.ji_contact = job.field_contact_name or job.sales_contact_name
    s.ji_phone = job.field_contact_phone or job.sales_contact_phone
    s.ji_email = job.field_contact_email or job.sales_contact_email
    s.cu_company = job.builder
    s.cu_name = job.sales_contact_name
    s.cu_phone = job.sales_contact_phone
    s.cu_email = job.sales_contact_email
    tops = t["tops"]
    s.sale_cabinets = compute.money(t["sell"] - tops)
    s.sale_countertops = tops
    if not s.pos:
        # seed the PO lines from the job's own cost build-up
        cab = compute.money(t["cabinets_cost"] + t["lumber_cost"] + t["hardware_material"])
        s.pos.append(CoverSheetPO(
            kind="product", po_type="Cabinets", po_abb="CAB",
            amount1=cab, amount2=t["freight"],
        ))
        s.pos.append(CoverSheetPO(
            kind="labor", po_type="Install", po_abb="INS",
            amount1=t["assembly"], amount2=t["install_cost"],
        ))
    db.commit()
    db.refresh(s)
    return _cover_out(db, s)


# --- Excel workbook (the database) ---

@router.get("/excel/status")
def excel_status():
    from app.sterling_app import xlsx_store

    return {
        "path": str(xlsx_store.WORKBOOK_PATH),
        "exists": xlsx_store.WORKBOOK_PATH.exists(),
        "dirty": xlsx_store.state["dirty"],
        "locked": xlsx_store.state["locked"],
        "last_saved": xlsx_store.state["last_saved"],
        "last_error": xlsx_store.state["last_error"],
    }


@router.post("/excel/save")
def excel_save():
    from app.sterling_app import xlsx_store

    ok = xlsx_store.save_now()
    if not ok:
        raise HTTPException(status_code=423, detail=xlsx_store.state["last_error"] or "Save failed")
    return excel_status()


@router.post("/excel/reload")
def excel_reload():
    from app.sterling_app import xlsx_store

    if not xlsx_store.WORKBOOK_PATH.exists():
        raise HTTPException(status_code=404, detail="Workbook not found")
    try:
        counts = xlsx_store.load_into_db()
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Reload failed: {exc}")
    return {"reloaded": counts}


# --- Export to CabinetTron ---

@router.post("/jobs/{job_id}/export")
def export_to_cabinettron(job_id: int, db: Session = Depends(get_db)):
    job = _get_or_404(db, Job, job_id, "Job")
    if job.stage != Stage.sold:
        raise HTTPException(status_code=422, detail="Only Sold jobs export to CabinetTron")
    try:
        result = export_job(db, job)
    except ExportError as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    job.exported_job_id = result["job_id"]
    job.exported_at = datetime.now(timezone.utc)
    db.commit()
    return result
