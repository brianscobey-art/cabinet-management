"""Prove app.pricing_engine still reproduces the workbook, then show a scenario.

Phase 1 verified the calculation by reimplementing it inline. This runs the same
check through the REAL engine, so the price-group and stocking-scenario work
cannot quietly break the arithmetic Phase 1 signed off on.

Legacy settings = base multiplier 0.21, price group 1, nothing stocked. Under
those the engine must land on the workbook's numbers to the cent.

    python -m scripts.verify_pricing_engine --pricing <xlsm> [--stocked-mult 0.185] [--top 25]
"""

import argparse
import collections

from openpyxl import load_workbook

from app.pricing_engine import PlanLine, Rates, Scenario, compare, price_plan

TOL = 0.011


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def load(path):
    """Everything the engine needs, straight out of the workbook."""
    wb = load_workbook(path, data_only=True)
    inv = wb["Grand Inventory By Price Group"]

    prices = collections.defaultdict(dict)
    for r in range(3, 2147):
        sku, pg, val = (inv.cell(row=r, column=c).value for c in (1, 2, 3))
        if not sku or not pg:
            continue
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue                      # some price cells hold "N/A"
        head = str(pg).split("-")[0].strip()
        if head.isdigit():
            prices[str(sku).strip()][int(head)] = val

    meta = {}
    for r in range(3, 394):
        sku = inv.cell(row=r, column=21).value
        if sku:
            meta[str(sku).strip()] = (
                inv.cell(row=r, column=22).value,
                num(inv.cell(row=r, column=25).value),
                num(inv.cell(row=r, column=26).value),
            )

    from app.pricing_engine import SkuInfo
    skus = {}
    for sku, by_group in prices.items():
        g, d, dr = meta.get(sku, (None, 0.0, 0.0))
        skus[sku] = SkuInfo(sku=sku, prices=by_group, install_group=g, doors=d, drawers=dr)

    fs = wb["Floorplan SKU"]
    lines = collections.defaultdict(list)
    for r in range(4, 3203):
        plan, qty, sku = (fs.cell(row=r, column=c).value for c in (2, 3, 4))
        if plan and sku:
            lines[str(plan).strip()].append(
                PlanLine(qty=num(qty), sku=str(sku).strip(),
                         area=str(fs.cell(row=r, column=5).value or "All"))
            )

    counts, hard_count = {}, {}
    for r in range(5, 345):
        p = fs.cell(row=r, column=31).value
        if p:
            counts.setdefault(str(p).strip(), {
                "assembly": num(fs.cell(row=r, column=39).value),
                "install": num(fs.cell(row=r, column=40).value),
                "hardware": num(fs.cell(row=r, column=41).value),
            })
    for r in range(1, 1010):
        p = fs.cell(row=r, column=43).value
        if p:
            hard_count.setdefault(str(p).strip(), num(fs.cell(row=r, column=46).value))

    ps = wb["Pricing Sheet"]
    expected = {}
    for r in range(10, 185):
        plan = ps.cell(row=r, column=2).value
        if not plan:
            continue
        expected[str(plan).strip()] = {
            "hard_sel": str(ps.cell(row=r, column=6).value or "KNB").strip(),
            "rates": Rates(
                tax=num(ps["D4"].value), freight_pct=num(ps["G2"].value),
                delivery=num(ps.cell(row=r, column=19).value),
                assem_rate=num(ps.cell(row=r, column=12).value),
                install_rate=num(ps.cell(row=r, column=13).value),
                hardware_rate=num(ps.cell(row=r, column=14).value),
                knob_cost=num(ps["E5"].value), handle_cost=num(ps["E6"].value),
                margin=num(ps.cell(row=r, column=21).value),
            ),
            "cabinet_cost": num(ps.cell(row=r, column=5).value, None),
            "total_labor": num(ps.cell(row=r, column=18).value, None),
            "total_cogs": num(ps.cell(row=r, column=20).value, None),
            "sale_price": num(ps.cell(row=r, column=22).value, None),
        }
    wb.close()
    return skus, lines, counts, hard_count, expected


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pricing", required=True)
    ap.add_argument("--base-mult", type=float, default=0.21)
    ap.add_argument("--stocked-mult", type=float, default=0.185)
    ap.add_argument("--top", type=int, default=25, help="how many SKUs to model as stocked")
    args = ap.parse_args()

    skus, lines, counts, hard_count, expected = load(args.pricing)
    legacy = Scenario(name="Current", base_multiplier=args.base_mult)

    print("=" * 78)
    print("REGRESSION — engine vs workbook, legacy settings (0.21, group 1, no stock)")
    print("=" * 78)
    checked = collections.Counter()
    bad = collections.defaultdict(list)
    for plan, exp in expected.items():
        if plan not in lines or plan not in counts:
            continue
        got = price_plan(
            lines[plan], skus, price_group=1, scenario=legacy, rates=exp["rates"],
            counts=counts[plan], hardware_count=hard_count.get(plan, 0.0),
            hardware_sel=exp["hard_sel"],
        )
        for field, mine in (("cabinet_cost", got.cabinet_cost),
                            ("total_labor", got.total_labor),
                            ("total_cogs", got.total_cogs),
                            ("sale_price", got.sale_price)):
            theirs = exp[field]
            if theirs is None:
                continue
            checked[field] += 1
            if abs(mine - theirs) > TOL:
                bad[field].append((plan, theirs, mine))

    total = sum(checked.values())
    wrong = sum(len(v) for v in bad.values())
    for f in checked:
        n, b = checked[f], len(bad[f])
        print(f"  {'OK  ' if b == 0 else 'DIFF'} {f:<16} {n - b}/{n}")
        for plan, theirs, mine in bad[f][:4]:
            print(f"        {plan[:34]:<36} workbook={theirs:<12.2f} engine={mine:.2f}")
    print(f"\n  {total - wrong}/{total} match")

    # ---- what a stocking program would be worth -----------------------------
    spend = collections.Counter()
    for plan, ls in lines.items():
        for ln in ls:
            info = skus.get(ln.sku)
            if info:
                spend[ln.sku] += ln.qty * (info.price(1) or 0.0)
    stocked = frozenset(s for s, _ in spend.most_common(args.top))
    program = Scenario(name=f"Stocking top {args.top}", base_multiplier=args.base_mult,
                       stocked_multiplier=args.stocked_mult, stocked_skus=stocked)

    print("\n" + "=" * 78)
    print(f"SCENARIO — stocking the top {args.top} SKUs at {args.stocked_mult} "
          f"(special order stays {args.base_mult})")
    print("=" * 78)
    tot_a = tot_b = 0.0
    rows = []
    for plan, exp in expected.items():
        if plan not in lines or plan not in counts:
            continue
        kw = dict(skus=skus, price_group=1, rates=exp["rates"], counts=counts[plan],
                  hardware_count=hard_count.get(plan, 0.0), hardware_sel=exp["hard_sel"])
        a = price_plan(lines[plan], scenario=legacy, **kw)
        b = price_plan(lines[plan], scenario=program, **kw)
        c = compare(a, b)
        tot_a += a.total_cogs
        tot_b += b.total_cogs
        rows.append((plan, c["total_cogs"]["delta"], b.stocked_lines, len(b.lines)))

    rows.sort(key=lambda x: x[1])
    print(f"  {len(rows)} plans | COGS today ${tot_a:,.0f} -> ${tot_b:,.0f} "
          f"| saving ${tot_a - tot_b:,.0f} ({(tot_a - tot_b) / tot_a * 100:.1f}%)")
    print("\n  biggest savers:")
    for plan, d, st, n in rows[:8]:
        print(f"    {plan[:34]:<36} ${-d:>8,.0f}   {st}/{n} lines stocked")


if __name__ == "__main__":
    main()
