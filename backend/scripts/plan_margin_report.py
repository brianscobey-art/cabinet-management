"""Priced plans vs what DR Horton actually pays — real margin, and what to review.

Joins three systems:
    VS combined report  PO Amount, by Job Number (the 9-digit BUID)
    CabinetTron         BUID -> job -> plan name
    Sterling            plan -> COGS and calculated sale price

Margin is measured against COGS, not `total`. The DRH cabinet PO covers cabinets
and install; countertops are a separate line, so folding tops into the cost
would understate margin on the plans that have them.

    python -m scripts.plan_margin_report --vs <xlsx> [--out-dir <dir>]
"""

import argparse
import collections
import statistics
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CARTER_GREEN = "125952"
RED = "DE2020"
SHEET_PREFIX = "DRH Cabinets Combined"


def read_vs(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    name = next((s for s in wb.sheetnames if s.startswith(SHEET_PREFIX)), None)
    if name is None:
        raise SystemExit(f"no '{SHEET_PREFIX}*' sheet in {Path(path).name}")
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if not r[ix["Job Number"]]:
            continue
        out.append({
            "buid": str(r[ix["Job Number"]]).strip(),
            "region": r[ix["Region"]],
            "project": r[ix["Project"]],
            "abbr": r[ix["Plan"]],
            "po_number": r[ix["PO Number"]],
            "po_amount": r[ix["PO Amount"]] if isinstance(r[ix["PO Amount"]], (int, float)) else None,
        })
    wb.close()
    return out


def plan_by_buid():
    from app.database import SessionLocal
    from app.models import Job, OrderingChecklist

    with SessionLocal() as db:
        pairs = (
            db.query(Job, OrderingChecklist)
            .join(OrderingChecklist, OrderingChecklist.job_id == Job.id)
            .all()
        )
        return {str(c.buid).strip(): (j.job_code, (j.plan or "").strip())
                for j, c in pairs if c.buid}


def sterling_prices():
    from app.sterling_app.compute import national_pricing_rows
    from app.sterling_app.database import SessionLocal

    with SessionLocal() as db:
        return {r["plan"]: r for r in national_pricing_rows(db, None, None)}


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor=CARTER_GREEN)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(ws, headers, widths, rows, money_cols=(), pct_cols=()):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    autosize(ws, widths)
    for c in money_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = '$#,##0.00'
    for c in pct_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = '0.0%'


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--vs", required=True)
    ap.add_argument("--out-dir", default=str(Path.home() / "Downloads"))
    ap.add_argument("--target", type=float, default=None,
                    help="override the target margin (else each plan's own)")
    args = ap.parse_args()

    vs = read_vs(args.vs)
    buids = plan_by_buid()
    priced = sterling_prices()

    # A job can carry more than one PO row: the cabinet PO plus a change order
    # or backcharge (-150, +225, -250 all appear). Sum per JOB first — treating
    # each row as its own job counted a -$100 adjustment as a whole house and
    # produced -87% margins that were an artefact, not a loss.
    by_job = collections.defaultdict(float)
    job_meta = {}
    for row in vs:
        if row["po_amount"] is None:
            continue
        by_job[row["buid"]] += float(row["po_amount"])
        job_meta.setdefault(row["buid"], row)

    per_plan = collections.defaultdict(list)
    no_job = no_plan = not_priced = 0
    no_amount = len({r["buid"] for r in vs}) - len(by_job)
    for buid, total in by_job.items():
        info = buids.get(buid)
        if info is None:
            no_job += 1
            continue
        _, plan = info
        if not plan:
            no_plan += 1
            continue
        if plan not in priced:
            not_priced += 1
            continue
        meta = dict(job_meta[buid])
        meta["po_amount"] = total
        per_plan[plan].append(meta)

    stamp = date.today().strftime("%m%d%y")
    out = Path(args.out_dir)
    summary_rows, review_rows = [], []
    total_gap = 0.0

    for plan, jobs in sorted(per_plan.items()):
        p = priced[plan]
        cogs = float(p["cogs"])
        sale = float(p["sale"])
        target = args.target if args.target is not None else float(p["margin_pct"])
        if target > 1:
            target /= 100.0
        amounts = [j["po_amount"] for j in jobs]
        avg_po = statistics.mean(amounts)
        actual = (avg_po - cogs) / avg_po if avg_po else 0.0
        gap_each = (target - actual) * avg_po
        gap_total = gap_each * len(jobs) if actual < target else 0.0
        total_gap += max(gap_total, 0.0)

        if actual < 0:
            status = "LOSS — below cost"
        elif actual < target - 0.005:
            status = "Below target"
        else:
            status = "OK"

        # Cost is stable and 24 of 32 repeat plans carry an IDENTICAL PO every
        # time, so a gap is not drift or variance — it is a fixed contract price
        # that repeats on every future house. Exposure is what that costs over
        # the 12 months of POs in this report.
        spread = max(amounts) - min(amounts)
        fixed = "Fixed" if spread < 1 else f"Varies ${spread:,.0f}"
        gap_house = avg_po - sale
        exposure = gap_house * len(jobs)

        row = [
            p["division"], plan, len(jobs), fixed,
            round(avg_po, 2), round(min(amounts), 2), round(max(amounts), 2),
            round(cogs, 2), round(sale, 2), round(gap_house, 2), round(exposure, 2),
            round(actual, 4), round(target, 4), round(actual - target, 4),
            round(gap_total, 2), status,
        ]
        summary_rows.append(row)
        if status != "OK":
            review_rows.append(row)

    summary_rows.sort(key=lambda r: r[10])      # worst 12-month exposure first
    review_rows.sort(key=lambda r: r[10])

    headers = ["Division", "Plan", "Houses (12mo)", "PO price", "Avg PO",
               "Min PO", "Max PO", "Sterling COGS", "Sterling Sale",
               "Gap / house", "12-month exposure", "Actual Margin", "Target",
               "Margin Gap", "$ Short vs target", "Status"]
    widths = [17, 28, 14, 14, 12, 12, 12, 14, 14, 13, 18, 13, 9, 12, 17, 18]
    money = (5, 6, 7, 8, 9, 10, 11, 15)
    pct = (12, 13, 14)

    wb1 = Workbook()
    write_sheet(wb1.active, headers, widths, summary_rows, money, pct)
    wb1.active.title = "Plan Margins"
    f1 = out / f"Plan Margin Summary {stamp}.xlsx"
    wb1.save(f1)

    wb2 = Workbook()
    write_sheet(wb2.active, headers, widths, review_rows, money, pct)
    wb2.active.title = "Needs Review"
    for r in range(2, wb2.active.max_row + 1):
        if str(wb2.active.cell(row=r, column=16).value).startswith("LOSS"):
            for c in range(1, len(headers) + 1):
                wb2.active.cell(row=r, column=c).font = Font(name="Calibri", size=11, color=RED, bold=True)
    f2 = out / f"Plans Needing Margin Review {stamp}.xlsx"
    wb2.save(f2)

    covered = sum(len(v) for v in per_plan.values())
    print(f"VS rows {len(vs)} -> {len(by_job)} jobs with a PO | priced & matched {covered} jobs across {len(per_plan)} plans")
    print(f"  excluded: {no_amount} no PO amount, {no_job} no CabinetTron job, "
          f"{no_plan} job has no plan, {not_priced} plan not priced in Sterling")
    print(f"\nplans below target: {len(review_rows)} of {len(summary_rows)}")
    print(f"total annualised shortfall on these POs: ${total_gap:,.0f}")
    under = sum(r[10] for r in summary_rows if r[10] < 0)
    over = sum(r[10] for r in summary_rows if r[10] > 0)
    print(f"")
    print(f"below our sale price: ${-under:,.0f} | above: ${over:,.0f} | NET ${-(under + over):,.0f}")
    print(f"")
    print(f"{'plan':<28}{'n':>4}{'PO':>14}{'gap/house':>11}{'12-month':>11}")
    for r in review_rows[:12]:
        print(f"{r[1][:27]:<28}{r[2]:>4}{str(r[3]):>14}{r[9]:>11,.0f}{r[10]:>11,.0f}")


if __name__ == "__main__":
    main()
