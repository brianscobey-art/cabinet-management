"""Import a builder pricing matrix into Sterling's contract_prices.

Built for the DRH Montgomery matrix (Pricing Matrix + Price Comparison sheets),
but reads by header name so a similarly shaped sheet for another division works
without code changes.

    Pricing Matrix     Plan Type | House Name | Floorplan | White (All-In) |
                       Color Upcharge | Dove Gray (All-In) | Charcoal (All-In)
    Price Comparison   Plan Type | Floorplan | Current Price | New Price ...

The effective date is read from the sheet ("Effective 2026-07-01") so nobody has
to remember to pass it. Prior price comes from Price Comparison — it is what
lets the reprice check say "this PO used the superseded number" rather than
just "this PO looks wrong".

Dry run by default.

    python -m scripts.import_price_matrix --file <xlsx> --division "DRH Montgomery" [--commit]
"""

import argparse
import re
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook


def _norm(v) -> str:
    return " ".join(str(v or "").split()).lower()


def _find_header(ws, wanted, limit=12):
    """Row index and {normalized header: column} for the first row carrying `wanted`."""
    for r in range(1, min(limit, ws.max_row) + 1):
        cells = {_norm(ws.cell(row=r, column=c).value): c
                 for c in range(1, ws.max_column + 1)
                 if ws.cell(row=r, column=c).value is not None}
        if all(any(w in h for h in cells) for w in wanted):
            return r, cells
    return None, {}


def _effective_date(ws) -> date | None:
    for r in range(1, 8):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, str):
                m = re.search(r"(\d{4})-(\d{2})-(\d{2})", v)
                if m:
                    return date(int(m[1]), int(m[2]), int(m[3]))
                m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", v)
                if m:
                    y = int(m[3])
                    return date(y + 2000 if y < 100 else y, int(m[1]), int(m[2]))
    return None


def read_matrix(path: str) -> tuple[list[dict], date | None]:
    wb = load_workbook(path, data_only=True)
    if "Pricing Matrix" not in wb.sheetnames:
        raise SystemExit("no 'Pricing Matrix' sheet in that workbook")
    ws = wb["Pricing Matrix"]
    eff = _effective_date(ws)
    hr, cells = _find_header(ws, ("floorplan", "all-in"))
    if hr is None:
        raise SystemExit("could not find the Pricing Matrix header row")

    def col(*names):
        for n in names:
            for h, c in cells.items():
                if n in h:
                    return c
        return None

    c_plan = col("floorplan")
    c_white = col("white")
    c_up = col("color upcharge", "upcharge")

    prior = {}
    if "Price Comparison" in wb.sheetnames:
        cw = wb["Price Comparison"]
        chr_, ccells = _find_header(cw, ("floorplan", "current price"))
        if chr_ is not None:
            p_plan = next(c for h, c in ccells.items() if "floorplan" in h)
            p_cur = next(c for h, c in ccells.items() if "current price" in h)
            for r in range(chr_ + 1, cw.max_row + 1):
                plan = cw.cell(row=r, column=p_plan).value
                cur = cw.cell(row=r, column=p_cur).value
                if plan and isinstance(cur, (int, float)):
                    prior[str(plan).strip()] = float(cur)

    out = []
    for r in range(hr + 1, ws.max_row + 1):
        plan = ws.cell(row=r, column=c_plan).value
        white = ws.cell(row=r, column=c_white).value if c_white else None
        if not plan or not isinstance(white, (int, float)):
            continue
        plan = str(plan).strip()
        up = ws.cell(row=r, column=c_up).value if c_up else None
        out.append({
            "plan": plan,
            "price": float(white),
            "color_upcharge": float(up) if isinstance(up, (int, float)) else None,
            "prior_price": prior.get(plan),
        })
    wb.close()
    return out, eff


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--division", required=True)
    ap.add_argument("--effective", help="YYYY-MM-DD; else read from the sheet")
    ap.add_argument("--commit", action="store_true")
    args = ap.parse_args()

    rows, eff = read_matrix(args.file)
    if args.effective:
        eff = date.fromisoformat(args.effective)
    if eff is None:
        raise SystemExit("no effective date found in the sheet — pass --effective")

    print(f"{len(rows)} plans, effective {eff}, division {args.division!r}")
    with_prior = sum(1 for r in rows if r["prior_price"])
    print(f"  {with_prior} carry a prior price (needed for the reprice check)")
    rises = [(r["plan"], r["prior_price"], r["price"]) for r in rows
             if r["prior_price"] and r["price"] > r["prior_price"]]
    if rises:
        avg = sum((n - o) / o for _, o, n in rises) / len(rises) * 100
        print(f"  {len(rises)} price increases, averaging {avg:.1f}%")
    for plan, o, n in rises[:5]:
        print(f"     {plan[:30]:<32} ${o:>8,.0f} -> ${n:>8,.0f}  (+{(n - o) / o * 100:.1f}%)")

    import pathlib

    from app.sterling_app.database import SessionLocal
    from app.sterling_app.models import ContractPrice

    src = pathlib.Path(args.file).name
    with SessionLocal() as db:
        added = updated = 0
        for r in rows:
            row = (db.query(ContractPrice)
                   .filter(ContractPrice.division == args.division,
                           ContractPrice.plan == r["plan"],
                           ContractPrice.effective_from == eff)
                   .first())
            if row is None:
                row = ContractPrice(division=args.division, plan=r["plan"], effective_from=eff)
                db.add(row)
                added += 1
            else:
                updated += 1
            row.price = Decimal(str(r["price"]))
            row.color_upcharge = (Decimal(str(r["color_upcharge"]))
                                  if r["color_upcharge"] is not None else None)
            row.prior_price = (Decimal(str(r["prior_price"]))
                               if r["prior_price"] is not None else None)
            row.source = src
        if args.commit:
            db.commit()
            print(f"\ncommitted: {added} added, {updated} updated")
        else:
            db.rollback()
            print(f"\ndry run — would add {added}, update {updated} (pass --commit)")


if __name__ == "__main__":
    main()
