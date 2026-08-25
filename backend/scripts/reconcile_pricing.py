"""Phase 1 of moving DRH pricing into Sterling: prove the math before building UI.

Reimplements the two workbooks' calculations in Python and diffs the result
against the workbooks' own cached values. READ-ONLY - opens copies, writes
nothing. If this does not reproduce Brian's numbers to the cent, nothing else
gets built on it.

    python -m scripts.reconcile_pricing --pricing <xlsm> --tops <xlsm>
"""

import argparse
import collections
from decimal import ROUND_HALF_UP, Decimal

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci

MONEY_TOL = 0.011  # a cent, plus float slop


def xround(x, digits=0):
    """Excel ROUND is half-away-from-zero; Python's round() is banker's rounding.

    Using the wrong one lands a penny out on every .5 and reads like a pricing
    bug rather than a rounding one.
    """
    if x is None:
        return None
    q = Decimal(1).scaleb(-digits)
    return float(Decimal(str(x)).quantize(q, rounding=ROUND_HALF_UP))


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def table_range(ws, name):
    """(left_col, header_row, right_col, last_row) for a worksheet table."""
    ref = ws.tables[name].ref
    a, b = ref.split(":")
    left = "".join(c for c in a if c.isalpha())
    top = "".join(c for c in a if c.isdigit())
    right = "".join(c for c in b if c.isalpha())
    bottom = "".join(c for c in b if c.isdigit())
    return ci(left), int(top), ci(right), int(bottom)


def headers(ws, name):
    """Normalized header -> column index. Headers carry newlines and double
    spaces ("Install  Rate", "Hard\\nSel"), so collapse whitespace and lowercase."""
    lo, hr, hi, _ = table_range(ws, name)
    out = {}
    for c in range(lo, hi + 1):
        raw = ws.cell(row=hr, column=c).value
        if raw:
            key = " ".join(str(raw).split()).lower()
            out.setdefault(key, c)          # first wins
            out.setdefault(key + "#2", None)
            if out[key + "#2"] is None and out[key] != c:
                out[key + "#2"] = c
    return {k: v for k, v in out.items() if v is not None}


class Diff:
    """Collects comparisons so the report shows what matched, not just failures."""

    def __init__(self):
        self.checked = collections.Counter()
        self.bad = collections.defaultdict(list)
        self.overrides = collections.defaultdict(list)

    def cmp_cell(self, label, key, mine, ws_f, row, col, ws_v):
        """Compare, but if the sheet cell is a typed-in constant where a formula
        belongs, count it as a manual override rather than a failure of the
        engine. Those are real business decisions, not bugs — but they are
        invisible in Excel, which is the whole reason to surface them."""
        raw = ws_f.cell(row=row, column=col).value
        theirs = num(ws_v.cell(row=row, column=col).value, None)
        if raw is not None and not (isinstance(raw, str) and str(raw).startswith("=")):
            self.checked[label] += 1
            if mine is not None and theirs is not None and abs(float(mine) - float(theirs)) > MONEY_TOL:
                self.overrides[label].append((key, theirs, mine))
            return
        self.cmp(label, key, mine, theirs)

    def cmp(self, label, key, mine, theirs, money=True):
        self.checked[label] += 1
        if mine is None or theirs is None:
            if mine != theirs:
                self.bad[label].append((key, theirs, mine))
            return
        if money and isinstance(theirs, (int, float)) and isinstance(mine, (int, float)):
            if abs(float(mine) - float(theirs)) > MONEY_TOL:
                self.bad[label].append((key, theirs, mine))
        elif str(mine).strip() != str(theirs).strip():
            self.bad[label].append((key, theirs, mine))

    def report(self, title, show=4):
        print("\n" + "=" * 78)
        print(title)
        print("=" * 78)
        total = sum(self.checked.values())
        wrong = sum(len(v) for v in self.bad.values())
        print(f"{total} values checked | {total - wrong} match | {wrong} differ\n")
        for label in self.checked:
            n = self.checked[label]
            b = len(self.bad[label])
            print(f"  {'OK  ' if b == 0 else 'DIFF'} {label:<26} {n - b}/{n}")
            for key, theirs, mine in self.bad[label][:show]:
                print(f"        {str(key)[:36]:<38} workbook={str(theirs)[:13]:<14} mine={str(mine)[:13]}")
            if b > show:
                print(f"        ... and {b - show} more")
        if any(self.overrides.values()):
            print("")
            print("MANUAL OVERRIDES (typed over the formula in the workbook):")
            for label, items in self.overrides.items():
                if items:
                    print(f"    {label}: {len(items)}")
                    for key, theirs, mine in items[:show]:
                        print(f"        {str(key)[:36]:<38} typed={str(theirs)[:12]:<13} formula would give={str(mine)[:12]}")
        return wrong


def reconcile_floorplan(wb, wbf):
    """FloorplanTable: Total, Install Group, Doors, Drawers for every SKU line."""
    d = Diff()
    inv = wb["Grand Inventory By Price Group"]

    lo, hr, _, last = table_range(inv, "Table2")
    price = {}
    for r in range(hr + 1, last + 1):
        sku = inv.cell(row=r, column=lo).value
        if sku:
            # Table2 lists each SKU once per price group (1- through 5-). The
            # workbook looks up on SKU alone, so XLOOKUP always lands on group 1
            # and the other four prices are never reachable. setdefault keeps
            # that behaviour; using [] here would silently price off group 5.
            price.setdefault(str(sku).strip(), num(inv.cell(row=r, column=lo + 2).value, None))

    lo2, hr2, _, last2 = table_range(inv, "Install_Hardware")
    install = {}
    for r in range(hr2 + 1, last2 + 1):
        sku = inv.cell(row=r, column=lo2).value
        if sku:
            install[str(sku).strip()] = (
                inv.cell(row=r, column=lo2 + 1).value,   # install group
                inv.cell(row=r, column=lo2 + 4).value,   # doors
                inv.cell(row=r, column=lo2 + 5).value,   # drawers
            )

    fs = wb["Floorplan SKU"]
    fsf = wbf["Floorplan SKU"]
    h = headers(fs, "FloorplanTable")
    _, hr3, _, last3 = table_range(fs, "FloorplanTable")
    for r in range(hr3 + 1, last3 + 1):
        sku = fs.cell(row=r, column=h["base sku"]).value
        if not sku:
            continue
        sku = str(sku).strip()
        qty = num(fs.cell(row=r, column=h["qty"]).value)
        key = f"{fs.cell(row=r, column=h['floorplan']).value} / {sku}"

        p = price.get(sku)
        if p is not None:
            d.cmp_cell("Total (price x qty)", key, p * qty, fsf, r, h["total"], fs)

        ih = install.get(sku)
        if ih:
            group, doors, drawers = ih
            d.cmp("Install Group", key, group,
                  fs.cell(row=r, column=h["install group"]).value, money=False)
            d.cmp("Doors", key, num(doors) * qty,
                  num(fs.cell(row=r, column=h["doors"]).value, None))
            d.cmp("Drawers", key, num(drawers) * qty,
                  num(fs.cell(row=r, column=h["drawers"]).value, None))
    return d


def reconcile_tops(path):
    """Every plan sheet: kitchen/vanity sq ft, sinks, cutouts, total cost."""
    d = Diff()
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        if ws["A10"].value != "Countertops":
            continue
        rate = num(ws["C5"].value)

        def block(rows):
            total = 0.0
            for r in rows:
                qty = num(ws.cell(row=r, column=2).value)
                width = num(ws.cell(row=r, column=3).value)
                depth = num(ws.cell(row=r, column=4).value)
                total += qty * xround(width * depth / 144, 2)
            return total

        # tops + backsplashes + side splash, kitchen and vanity
        k = block(range(12, 18)) + block(range(29, 34)) + block(range(45, 46))
        v = block(range(21, 26)) + block(range(37, 42)) + block(range(49, 50))
        k_sf, v_sf = xround(k, 0), xround(v, 0)
        d.cmp("Kitchen SqFt", ws.title, k_sf, num(ws["H3"].value, None))
        d.cmp("Vanity SqFt", ws.title, v_sf, num(ws["H7"].value, None))
        d.cmp("Total SqFt", ws.title, k_sf + v_sf, num(ws["C6"].value, None))

        cost = (k_sf * rate + v_sf * rate
                + num(ws["B53"].value) * num(ws["D53"].value)     # kitchen sink
                + num(ws["B54"].value) * num(ws["D54"].value)     # kitchen cutout
                + num(ws["B58"].value) * num(ws["D58"].value)     # vanity sink
                + num(ws["B59"].value) * num(ws["D59"].value))    # vanity cutout
        d.cmp("Total Cost", ws.title, cost, num(ws["C8"].value, None))
    wb.close()
    return d


def reconcile_table6(wb, wbf):
    """The pricing chain, per Builder+Plan."""
    d = Diff()
    ps = wb["Pricing Sheet"]
    MULT = num(ps["D2"].value)        # list price -> cost
    TAX = num(ps["D4"].value)
    FRT = num(ps["G2"].value)
    DELIV = num(ps["G3"].value)
    R_ASSEM = num(ps["J2"].value)
    R_INST = num(ps["J3"].value)
    R_KNOB = num(ps["J4"].value)
    R_HNDL = num(ps["J5"].value)
    R_SHOE = num(ps["J6"].value)
    KNB = num(ps["E5"].value)
    HNDL = num(ps["E6"].value)
    LAM = num(ps["Y2"].value)
    SINK_CO = num(ps["AB4"].value)
    hard_rate = {str(ps["D5"].value).strip(): KNB, str(ps["D6"].value).strip(): HNDL}

    fs = wb["Floorplan SKU"]
    list_price, hard_count, shoe = {}, {}, {}
    for r in range(1, 1010):
        a = fs.cell(row=r, column=19).value      # S: Floorplan
        if a:
            list_price.setdefault(str(a).strip(), num(fs.cell(row=r, column=20).value, None))
        b = fs.cell(row=r, column=43).value      # AQ: Floorplan
        if b:
            hard_count.setdefault(str(b).strip(), num(fs.cell(row=r, column=46).value, None))
        c = fs.cell(row=r, column=48).value      # AV: Floorplan
        if c:
            shoe.setdefault(str(c).strip(), num(fs.cell(row=r, column=49).value, None))

    h8 = headers(fs, "Table8")
    _, hr8, _, last8 = table_range(fs, "Table8")
    t8 = {}
    for r in range(hr8 + 1, last8 + 1):
        p = fs.cell(row=r, column=h8["floorplan"]).value
        if p:
            t8[str(p).strip()] = {
                k: num(fs.cell(row=r, column=h8[k]).value)
                for k in ("assembly", "install", "hardware")
            }

    psf = wbf["Pricing Sheet"]
    h6 = headers(ps, "Table6")
    _, hr6, _, last6 = table_range(ps, "Table6")

    def col(r, name):
        return ps.cell(row=r, column=h6[name]).value

    for r in range(hr6 + 1, last6 + 1):
        plan = col(r, "plan")
        if not plan:
            continue
        plan = str(plan).strip()
        if plan not in list_price or plan not in t8:
            continue
        key = plan

        lp = list_price[plan]
        cost = lp * MULT
        d.cmp("Cabinet List Price", key, lp, num(col(r, "cabinet list price"), None))
        d.cmp("Cabinet Cost", key, cost, num(col(r, "cabinet cost"), None))

        hsel = str(col(r, "hard sel") or "").strip()
        hc = hard_rate.get(hsel, 0.0) * (hard_count.get(plan) or 0.0)
        d.cmp("Hard Cost", key, hc, num(col(r, "hard cost"), None))

        tot_mat = cost + hc
        tax = tot_mat * TAX
        frt = cost * FRT
        # H "Total Materials" and K "Total  Materials" collapse to one key once
        # whitespace is normalized; #2 is the second occurrence (K).
        d.cmp("Total Materials", key, tot_mat, num(col(r, "total materials"), None))
        d.cmp("Total Materials +tax/frt", key, tot_mat + tax + frt,
              num(col(r, "total materials#2"), None))
        d.cmp("Tax", key, tax, num(col(r, "tax"), None))
        d.cmp("Freight", key, frt, num(col(r, "freight"), None))

        # Each row carries its own rates (Install is 25, 27 or 27.5; Hard is
        # 1 or 2). Using the sheet-level J2/J3/J4 defaults misprices 44 plans.
        r_assem = num(col(r, "assem rate"), R_ASSEM)
        r_inst = num(col(r, "install rate"), R_INST)
        r_hard = num(col(r, "hard rate"), R_KNOB)
        assem = t8[plan]["assembly"] * r_assem
        inst = xround(t8[plan]["install"] * r_inst, 0)
        hw = t8[plan]["hardware"] * r_hard
        labor = assem + inst + hw
        d.cmp("Assem", key, assem, num(col(r, "assem"), None))
        d.cmp("Install", key, inst, num(col(r, "install"), None))
        d.cmp("Hardware", key, hw, num(col(r, "hardware"), None))
        d.cmp("Total Labor", key, labor, num(col(r, "total labor"), None))

        # Delivery is $G$3 (0) on most rows but typed as 50 on 13 of them.
        deliv = num(col(r, "delivery"), DELIV)
        cogs = tot_mat + tax + frt + labor + deliv
        d.cmp("Total COGS", key, cogs, num(col(r, "total cogs"), None))

        margin = num(col(r, "cabinet margin"))
        if 0 <= margin < 1:
            sale = xround(cogs / (1 - margin), 0)
            d.cmp_cell("Calculated Sale Price", key, sale, psf, r,
                       h6["calculated sale price"], ps)
            # Profit is Sale - COGS, so where the sale price was typed over the
            # formula the profit follows the TYPED price, not the computed one.
            raw_sale = psf.cell(row=r, column=h6["calculated sale price"]).value
            sale_eff = sale
            if raw_sale is not None and not (isinstance(raw_sale, str) and str(raw_sale).startswith("=")):
                sale_eff = num(col(r, "calculated sale price"), sale)
            d.cmp_cell("Profit", key, sale_eff - cogs, psf, r, h6["profit"], ps)

        hcount = hard_count.get(plan)
        if hcount is not None:
            d.cmp("Hrdwre Count", key, hcount, num(col(r, "hrdwre count"), None))
            d.cmp("Knob Cost", key, (hcount * KNB) * TAX + (hcount * KNB),
                  num(col(r, "knob cost"), None))
            d.cmp("Knob Labor", key, hcount * R_KNOB, num(col(r, "knob labor"), None))
            d.cmp("Handle Cost", key, (hcount * HNDL) * TAX + (hcount * HNDL),
                  num(col(r, "handle cost"), None))
            d.cmp("Handle Install", key, hcount * R_HNDL,
                  num(col(r, "handle install"), None))

        if plan in shoe:
            d.cmp("Shoe Qty", key, shoe[plan], num(col(r, "shoe qty"), None))
            d.cmp("Total Shoe Install", key, shoe[plan] * R_SHOE,
                  num(col(r, "total shoe install"), None))

        ks = num(col(r, "k-top sqft"), None)
        if ks is not None:
            d.cmp("K-Top LAM SqFt", key, ks * LAM, num(col(r, "k-top lam sqft"), None))
        vs = num(col(r, "v-top sqft"), None)
        if vs is not None:
            d.cmp("V-Top LAM SqFt", key, vs * LAM, num(col(r, "v-top lam sqft"), None))
        kco = num(col(r, "k-top sink co"), None)
        if kco is not None:
            d.cmp("K-Sink C/O", key, kco * SINK_CO, num(col(r, "k-sink c/o"), None))
    return d


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pricing", required=True)
    ap.add_argument("--tops", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.pricing, data_only=True)
    wbf = load_workbook(args.pricing, data_only=False)   # to spot typed-over cells
    bad = reconcile_floorplan(wb, wbf).report("A. FloorplanTable - SKU lookups")
    bad += reconcile_table6(wb, wbf).report("C. Table6 - the pricing chain")
    wb.close()
    wbf.close()
    bad += reconcile_tops(args.tops).report("B. Countertops - dimensions to sq ft")

    print("\n" + "=" * 78)
    if bad == 0:
        print("VERDICT: the engine reproduces both workbooks exactly.")
    else:
        print(f"VERDICT: {bad} values differ - resolve before building on this.")


if __name__ == "__main__":
    main()
