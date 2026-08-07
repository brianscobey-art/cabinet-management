"""Sync jobs from the 3.0 Online Sales Tracker's DATA table (Command Center sheet).

Read-only against the workbook. Jobs are keyed by Job Code (column A). The tracker
is the source of truth for status (its CONST LVL column) and the install date
(Actual Install Date, falling back to Requested only when there's no actual yet):

  * existing jobs are refreshed — status + install date always re-synced, other
    fields filled only when blank so richer data isn't clobbered;
  * new active rows are created;
  * new rows that are already closed (6.0-Clsd) or void (8.0-Void) are skipped —
    they'd only ever live on the Archive tab. Existing jobs that turn closed/void
    are updated so they drop off the active views.

Re-run after each tracker update to keep CabinetTron aligned with the latest sheet.

Usage (from backend/):
    python -m scripts.import_tracker "<path to .xlsm>" [--dry-run] [--selections-only]
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal
from app.feeds import _find_job
from app.sales import resolve_salesperson
from app.models import (
    Account,
    AccountType,
    Community,
    HardwareSelection,
    Job,
    JobStatus,
    JobType,
    RoomSelection,
)

SHEET = "Command Center"
TABLE_NAME = "DATA"
BLANKS = {None, 0, "0", "", "#N/A", "N/A", "NA", "TBD", "?", "NONE", "None", "none"}

DEFAULT_SALES_CONTACT = ("Brian Scobey", "850-890-0482", "Brian.Scobey@CarterLumber.com")


def clean(value):
    if isinstance(value, str):
        value = value.strip()
    return None if value in BLANKS else value


def as_date(value) -> date | None:
    value = clean(value)
    return value.date() if isinstance(value, datetime) else None


def money_str(value) -> str | None:
    value = clean(value)
    if isinstance(value, (int, float)):
        return f"${value:,.2f}"
    return None


# The tracker's "CONST LVL" column is Brian's status ladder verbatim — the same
# strings as JobStatus values (e.g. "6.0-Clsd", "8.0-Void"). It's the authoritative
# status, so we use it directly rather than guessing from milestone dates.
CONST_LVL_COL = "CONST LVL"
_CONST_LVL = {st.value: st for st in JobStatus}
INACTIVE = (JobStatus.closed, JobStatus.void)  # closed/void → Archive only, never active views


def status_from_row(row: dict) -> JobStatus:
    """Status straight from the tracker's CONST LVL column; fall back to the
    date-derived guess only when CONST LVL is blank or unrecognized."""
    raw = clean(row.get(CONST_LVL_COL))
    if raw is not None:
        member = _CONST_LVL.get(str(raw).strip())
        if member is not None:
            return member
    return derive_status(row)


def install_from_row(row: dict) -> date | None:
    """Actual Install Date is authoritative; Requested only stands in when there's
    no actual yet (so a not-yet-installed job still lands on the calendar)."""
    return as_date(row.get("Actual Install Date")) or as_date(row.get("Requested Install Date"))


def derive_status(row: dict) -> JobStatus:
    """Fallback status from the tracker's milestone dates (used only when CONST
    LVL is blank).

    Tracker milestone columns hold SCHEDULED dates too — a punch date in
    September doesn't mean the job is done in July. Only dates that have
    already passed count as reached milestones.
    """
    today = date.today()

    def past(key: str) -> bool:
        d = as_date(row.get(key))
        return d is not None and d <= today

    if past("Full Punch Date"):
        return JobStatus.closed
    if past("Actual Install Date"):
        return JobStatus.ndqw
    if past("Cabinet Receipt Date"):
        return JobStatus.ord
    if past("Cabinet Order Date") or clean(row.get("Cabinet PO#")):
        return JobStatus.ord
    if past("Actual Measure Date") or past("Req Measure Date"):
        return JobStatus.preord
    return JobStatus.track


def load_rows(path: Path) -> list[dict]:
    # Not read_only: we need ws.tables to honor the DATA table's exact range —
    # cells below the table are dashboard UI artifacts, not jobs.
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    ref = ws.tables[TABLE_NAME].ref  # e.g. "A3:CQ502"
    m = re.match(r"[A-Z]+(\d+):([A-Z]+)(\d+)", ref)
    header_row, max_row = int(m.group(1)), int(m.group(3))

    rows = ws.iter_rows(min_row=header_row, max_row=max_row, max_col=95, values_only=True)
    headers = [
        h.replace("\n", " ").strip() if isinstance(h, str) else h for h in next(rows)
    ]
    out = []
    for values in rows:
        row = dict(zip(headers, values))
        if clean(row.get("Job Code")):
            out.append(row)
    wb.close()
    return out


def get_or_create_account(db, cache: dict, name: str, acc_type: AccountType) -> Account:
    if name in cache:
        return cache[name]
    account = db.query(Account).filter(Account.name == name).first()
    if account is None:
        account = Account(name=name, type=acc_type)
        db.add(account)
        db.flush()
    cache[name] = account
    return account


def get_or_create_community(db, cache: dict, account: Account, name: str, market: str | None) -> Community:
    key = (account.id, name)
    if key in cache:
        return cache[key]
    community = (
        db.query(Community)
        .filter(Community.account_id == account.id, Community.name == name)
        .first()
    )
    if community is None:
        community = Community(account_id=account.id, name=name, market=market)
        db.add(community)
        db.flush()
    cache[key] = community
    return community


def update_existing(db, job: Job, row: dict) -> str:
    """Refresh an existing job from the latest tracker row.

    The tracker is the source of truth for status (CONST LVL) and the install
    date (Actual Install Date); those are always synced. Other fields are only
    filled when blank so richer data (sold-job files, manual edits) isn't lost.
    """
    changed = False
    status = status_from_row(row)
    if job.status != status:
        job.status = status
        changed = True

    actual = as_date(row.get("Actual Install Date"))
    requested = as_date(row.get("Requested Install Date"))
    # Actual always wins; Requested only sets a date when the job has none yet.
    new_install = actual or (requested if job.install_date is None else None)
    if new_install and job.install_date != new_install:
        job.install_date = new_install
        if job.warranty_start_date is None:
            job.warranty_start_date = new_install
        changed = True

    measure = as_date(row.get("Actual Measure Date"))
    if measure and job.measure_date != measure:
        job.measure_date = measure
        changed = True

    plan = clean(row.get("House Plan"))
    if plan and not job.plan:
        job.plan = str(plan)[:100]
        changed = True

    if _apply_selections(db, job, row, create_room=True):
        changed = True
    return "updated" if changed else "unchanged"


def import_row(db, row: dict, caches: dict) -> str:
    job_code = str(clean(row["Job Code"])).strip()
    existing = db.query(Job).filter(Job.job_code == job_code).first()
    if existing is not None:
        return update_existing(db, existing, row)

    builder = clean(row.get("Full Builder")) or clean(row.get("Builder"))
    community_name = clean(row.get("Community"))
    city = clean(row.get("City"))
    state = clean(row.get("State"))
    market = ", ".join(str(p) for p in (city, state) if p) or None

    if builder:
        account = get_or_create_account(db, caches["accounts"], str(builder), AccountType.builder)
        community = (
            get_or_create_community(db, caches["communities"], account, str(community_name), market)
            if community_name
            else None
        )
        job_type = JobType.tract
    else:
        # Retail/local rows carry the customer name in the Community column.
        account_name = str(community_name or job_code)
        account = get_or_create_account(db, caches["accounts"], account_name, AccountType.retail)
        community = None
        job_type = JobType.custom

    lot = clean(row.get("Lot #"))

    # A feed sync may have created this job before the tracker knew its code —
    # match by community + lot, adopt the code, and refresh it from the tracker.
    if community is not None and lot is not None:
        found = _find_job(db, community, lot)
        if found is not None:
            if found.job_code is None:
                found.job_code = job_code
            update_existing(db, found, row)
            return "linked"

    # Brand-new row: don't create jobs that are already closed or void — they'd
    # never appear on an active view anyway (they live on the Archive tab).
    status = status_from_row(row)
    if status in INACTIVE:
        return "skipped_inactive"

    address = clean(row.get("Address"))
    if not address:
        address = f"{community_name or account.name}" + (f" Lot {lot}" if lot else "")
    if market:
        address = f"{address}, {market}"

    note_bits = []
    for label, key in [
        ("Plan", "House Plan"),
        ("Phase", "Phase"),
        ("Installer", "Installer"),
        ("Cabinet PO#", "Cabinet PO#"),
        ("Salesperson", "Salesperson"),
    ]:
        if clean(row.get(key)):
            note_bits.append(f"{label}: {clean(row[key])}")
    if money_str(row.get("Actual PO Amount")):
        note_bits.append(f"PO Amount: {money_str(row['Actual PO Amount'])}")
    for label, key in [
        ("Measured", "Actual Measure Date"),
        ("Cabinets received", "Cabinet Receipt Date"),
        ("Punch", "Full Punch Date"),
    ]:
        if as_date(row.get(key)):
            note_bits.append(f"{label}: {as_date(row[key]).isoformat()}")
    note_bits.append("Imported from 3.0 Online Sales Tracker")

    salesperson = clean(row.get("Salesperson"))
    super_name = clean(row.get("Super"))
    install = install_from_row(row)

    job = Job(
        job_code=job_code,
        account_id=account.id,
        community_id=community.id if community else None,
        lot_number=str(lot) if lot is not None else None,
        address=str(address)[:500],
        job_type=job_type,
        plan=str(clean(row.get("House Plan")) or "")[:100] or None,
        measure_date=as_date(row.get("Actual Measure Date")),
        status=status,
        install_date=install,
        warranty_start_date=install,
        salesperson=resolve_salesperson(account.name, str(salesperson) if salesperson else None),
        sales_contact_name=DEFAULT_SALES_CONTACT[0],
        sales_contact_phone=DEFAULT_SALES_CONTACT[1],
        sales_contact_email=DEFAULT_SALES_CONTACT[2],
        field_contact_name=str(super_name) if super_name else "TBD",
        field_contact_phone=str(clean(row.get("Super Phone")) or "") or None,
        notes=" | ".join(note_bits),
    )
    db.add(job)
    db.flush()

    _apply_selections(db, job, row, create_room=True)
    return "imported"


# The selections block uses double-spaced headers in the tracker.
SEL_BRAND = "Cabinet  Brand"
SEL_SERIES = "Cabinet  Series"
SEL_DOOR_STYLE = "Door  Style"
SEL_DOOR_HW = "Door  Hrdwe"
SEL_DRW_HW = "Drw  Hrdwe"


def _apply_selections(db, job: Job, row: dict, create_room: bool) -> bool:
    """Fill the Whole House room selection + door/drawer hardware from tracker columns.

    Only fills blanks — never clobbers values that came from richer sources
    (e.g. the sold-job-file import). Returns True if anything changed.
    """
    changed = False
    brand = clean(row.get(SEL_BRAND)) or clean(row.get("Cabinet Brand"))
    series = clean(row.get(SEL_SERIES))
    door_style = clean(row.get(SEL_DOOR_STYLE))

    room = (
        db.query(RoomSelection)
        .filter(RoomSelection.job_id == job.id, RoomSelection.room == "Whole House")
        .first()
    )
    if room is None and (brand or series or door_style) and create_room:
        room = RoomSelection(job_id=job.id, room="Whole House", notes="From tracker")
        db.add(room)
        db.flush()
        changed = True
    if room is not None:
        for attr, value in (("cabinet_brand", brand), ("series", series), ("door_style", door_style)):
            if value and not getattr(room, attr):
                setattr(room, attr, str(value))
                changed = True

    for hw_type, key in (("door", SEL_DOOR_HW), ("drawer", SEL_DRW_HW)):
        code = clean(row.get(key))
        if not code:
            continue
        exists = (
            db.query(HardwareSelection)
            .filter(HardwareSelection.job_id == job.id, HardwareSelection.hardware_type == hw_type)
            .first()
        )
        if exists is None:
            db.add(
                HardwareSelection(
                    job_id=job.id, room="Whole House", hardware_type=hw_type, item=str(code)
                )
            )
            changed = True
    return changed


def backfill_selections(db, rows: list[dict]) -> dict:
    """Apply tracker selections to already-imported jobs (matched by job code)."""
    counts = {"updated": 0, "unchanged": 0, "no_job": 0}
    for row in rows:
        job_code = str(clean(row["Job Code"])).strip()
        job = db.query(Job).filter(Job.job_code == job_code).first()
        if job is None:
            counts["no_job"] += 1
            continue
        if _apply_selections(db, job, row, create_room=True):
            counts["updated"] += 1
        else:
            counts["unchanged"] += 1
    return counts


def run_sync(db, rows: list[dict], *, on_error=None) -> dict:
    """Upsert every tracker row into the db (caller commits). Returns outcome counts."""
    counts = {"imported": 0, "updated": 0, "unchanged": 0, "linked": 0,
              "skipped_inactive": 0, "failed": 0}
    caches = {"accounts": {}, "communities": {}}
    for row in rows:
        try:
            counts[import_row(db, row, caches)] += 1
        except Exception as exc:  # noqa: BLE001 - one bad row shouldn't kill the sync
            counts["failed"] += 1
            if on_error:
                on_error(clean(row.get("Job Code")), exc)
    return counts


def sync_tracker_file(db, path: Path, *, on_error=None) -> dict:
    """Load the tracker workbook and sync every row (caller commits)."""
    rows = load_rows(path)
    return {"file": path.name, "rows": len(rows), **run_sync(db, rows, on_error=on_error)}


def main() -> None:
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry_run = "--dry-run" in flags
    if not args:
        print('Usage: python -m scripts.import_tracker "<path to .xlsm>" [--dry-run] [--selections-only]')
        sys.exit(1)

    path = Path(args[0])
    rows = load_rows(path)
    print(f"{len(rows)} rows with job codes in {path.name}")

    with SessionLocal() as db:
        if "--selections-only" in flags:
            counts = backfill_selections(db, rows)
        else:
            counts = run_sync(db, rows, on_error=lambda code, exc: print(f"! {code}: {exc}"))
        if dry_run:
            db.rollback()
            print("(dry run — rolled back)")
        else:
            db.commit()
    print(counts)


if __name__ == "__main__":
    main()
