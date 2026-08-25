"""Import the DRH Top Pricing Sheet into Sterling's PlanTops / TopPiece.

One worksheet per plan, all on the same 59-row template:

    rows 12-17  Countertops   Kitchen      (qty, width, depth)
    rows 21-25  Countertops   Vanity
    rows 29-33  Backsplashes  Kitchen      (depth column is height)
    rows 37-41  Backsplashes  Vanity
    row  45     Side Splash   Kitchen
    row  49     Side Splash   Vanity
    rows 53/54  Kitchen sink qty / cutout qty
    rows 58/59  Vanity  sink qty / cutout qty

Verified against the workbook before writing: Sterling's tops_total() reproduces
every sheet's own Total Cost. Dry run by default — pass --commit to write.

    python -m scripts.import_tops --tops <xlsm> [--division "DRH PC"] [--commit]
"""

import argparse
from decimal import Decimal

from openpyxl import load_workbook

# (first_row, last_row_inclusive, area, kind)
BLOCKS = [
    (12, 17, "Kitchen", "top"),
    (21, 25, "Vanity", "top"),
    (29, 33, "Kitchen", "backsplash"),
    (37, 41, "Vanity", "backsplash"),
    (45, 45, "Kitchen", "side_splash"),
    (49, 49, "Vanity", "side_splash"),
]


def num(cell, default=0.0):
    try:
        return float(cell.value)
    except (TypeError, ValueError):
        return default


def read_sheet(ws) -> dict | None:
    """One plan's tops, or None if the sheet is not a filled-in template."""
    if ws["A10"].value != "Countertops":
        return None

    pieces = []
    for first, last, area, kind in BLOCKS:
        for r in range(first, last + 1):
            qty = num(ws.cell(row=r, column=2))
            width = num(ws.cell(row=r, column=3))
            depth = num(ws.cell(row=r, column=4))
            if qty <= 0 or width <= 0 or depth <= 0:
                continue      # blank template row
            pieces.append({"area": area, "kind": kind, "qty": int(qty),
                           "width": width, "depth": depth})
    if not pieces:
        return None           # the untouched "Template" sheet

    k_sink, k_cut = num(ws["B53"]), num(ws["B54"])
    v_sink, v_cut = num(ws["B58"]), num(ws["B59"])
    return {
        "plan": ws.title.strip(),
        "material": str(ws["C4"].value or "").strip() or "Laminate",
        "rate_sqft": num(ws["C5"], None),
        "k_sinks": int(k_sink), "v_sinks": int(v_sink),
        # Cutout counts usually equal the sink count but not always, and the
        # unit prices are supposed to come from the builder-tier table yet were
        # typed in by hand on half the sheets. Carry all of it, assume nothing.
        "k_cutouts": int(k_cut), "v_cutouts": int(v_cut),
        "k_sink_rate": num(ws["D53"], None), "k_cutout_rate": num(ws["D54"], None),
        "v_sink_rate": num(ws["D58"], None), "v_cutout_rate": num(ws["D59"], None),
        # Sterling charges one cutout per sink. Where the sheet disagrees the
        # import cannot represent it, so it is reported rather than rounded away.
        "sink_mismatch": (k_sink != k_cut or v_sink != v_cut),
        "cutouts": (int(k_cut), int(v_cut)),
        "pieces": pieces,
        "sheet_total": num(ws["C8"], None),
        "sheet_sqft": num(ws["C6"], None),
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tops", required=True)
    ap.add_argument("--division", default="DRH PC")
    ap.add_argument("--commit", action="store_true", help="write; otherwise dry run")
    args = ap.parse_args()

    wb = load_workbook(args.tops, data_only=True)
    plans = [p for p in (read_sheet(ws) for ws in wb.worksheets) if p]
    wb.close()
    print(f"{len(plans)} plans read, {sum(len(p['pieces']) for p in plans)} pieces")

    odd = [p for p in plans if p["sink_mismatch"]]
    if odd:
        print(f"\nsink/cutout mismatch on {len(odd)} sheet(s) — Sterling ties cutouts to sinks:")
        for p in odd:
            print(f"   {p['plan']}: sinks K={p['k_sinks']} V={p['v_sinks']} "
                  f"but cutouts K={p['cutouts'][0]} V={p['cutouts'][1]}")

    from app.sterling_app.compute import tops_total
    from app.sterling_app.database import SessionLocal
    from app.sterling_app.models import PlanTops, TopPiece

    with SessionLocal() as db:
        written = skipped = 0
        for p in plans:
            row = (
                db.query(PlanTops)
                .filter(PlanTops.division == args.division, PlanTops.plan == p["plan"])
                .first()
            )
            if row is None:
                row = PlanTops(division=args.division, plan=p["plan"])
                db.add(row)
            elif row.pieces and not args.commit:
                skipped += 1
            row.material = p["material"]
            row.rate_sqft = Decimal(str(p["rate_sqft"])) if p["rate_sqft"] else None
            row.k_sinks = p["k_sinks"]
            row.v_sinks = p["v_sinks"]
            row.k_cutouts = p["k_cutouts"]
            row.v_cutouts = p["v_cutouts"]
            for attr in ("k_sink_rate", "k_cutout_rate", "v_sink_rate", "v_cutout_rate"):
                val = p[attr]
                setattr(row, attr, Decimal(str(val)) if val else None)
            row.pieces = [
                TopPiece(area=q["area"], kind=q["kind"], qty=q["qty"],
                         width=Decimal(str(q["width"])), depth=Decimal(str(q["depth"])))
                for q in p["pieces"]
            ]
            db.flush()
            written += 1

        # Verify against the workbook BEFORE committing anything.
        bad = []
        for p in plans:
            row = (
                db.query(PlanTops)
                .filter(PlanTops.division == args.division, PlanTops.plan == p["plan"])
                .first()
            )
            got = tops_total(db, row)
            if p["sheet_total"] is not None:
                if abs(float(got["total"]) - p["sheet_total"]) > 0.011:
                    bad.append((p["plan"], p["sheet_total"], float(got["total"])))

        print(f"\nverify: {len(plans) - len(bad)}/{len(plans)} plans match the sheet's own total")
        for plan, sheet, mine in bad[:8]:
            print(f"   {plan[:34]:<36} sheet={sheet:<10.2f} sterling={mine:.2f}")

        if args.commit and not bad:
            db.commit()
            print(f"\ncommitted {written} plans to Sterling ({args.division})")
        elif args.commit:
            db.rollback()
            print("\nNOT committed — verification failed, nothing written")
        else:
            db.rollback()
            print("\ndry run — nothing written (pass --commit)")


if __name__ == "__main__":
    main()
