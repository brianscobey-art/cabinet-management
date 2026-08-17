"""Backfill the Order Pack board from history.

The board only fills itself going forward, so every job filed before Order Pack
existed shows blank columns. Two sources can fix that, and both live on this PC:

  1. The 3.0 Online Sales Tracker — the `DATA` table on Command Center (BUID,
     plan abbreviation, swing, PO amount, Labor Total) and the `POTracker` table
     on PO Tracking (Carter PO number).
  2. The filed job folders under Sold Job Files\\National Accounts\\
     DR Horton - All\\{Region}\\{Community} {Sub #}\\{Job folder}. The folder
     names and the files inside carry the sub number, the Everluxe SO, the
     Carter PO, and the date it was filed.

WHAT IS DELIBERATELY NOT BACKFILLED
-----------------------------------
`so_total`. POTracker has a `Cost` column that looks like the answer and is not:
on DROP-0094 it reads $2,022.75 while that job's Carter PO 750005026 reads
$2,219.69. Backfilling from it would paint hundreds of false mismatches onto the
one signal the dollar gate exists to make trustworthy. The SO total comes off the
SO PDF or it stays blank.

`so_number` likewise does NOT come from POTracker's "Our Sales Order #" — that is
a Carter number (75000015529), not the Everluxe SO. The Everluxe SO is read off
the filed PDF's own filename instead.

`elevation` and `po_date` are in neither source; they only exist in the
Selections and PO PDFs.

`install_pay` DOES come from the tracker's Labor Total, and that is not a guess:
it is the very column the installer pay sheet generator prints from. Verified
8/17/26 against four sheets — 383 / 387 / 350 / 587, exact matches.

The server applies everything fill-if-null, so this can be re-run safely and can
never overwrite something a real stage run established.

Usage (from the repo root):
    backend\\.venv\\Scripts\\python.exe agent\\backfill_history.py --dry-run
    backend\\.venv\\Scripts\\python.exe agent\\backfill_history.py
"""

import argparse
import os
import re
import sys
from datetime import date
from pathlib import Path

_JOBCODE_RE = re.compile(r"^([A-Z0-9]{2,7}-\d{3,4})\b")
_SO_RE = re.compile(r"\b(SO\d{4,})\b", re.IGNORECASE)
_CARTER_PO_RE = re.compile(r"\b(750\d{6})\b")
_SUB_RE = re.compile(r"\s(\d{4,6})$")           # "Gator's Landing 22792"
_TRACKER_DATE_RE = re.compile(r"(\d{6})\.xlsm$", re.IGNORECASE)
_NOT_A_JOB = ("_to_delete", "archive", "forms", "templates", "old")


# ---------------------------------------------------------------------------
# Source 1: the tracker
# ---------------------------------------------------------------------------
def newest_tracker(tracker_dir: Path) -> Path | None:
    """Newest workbook by the MMDDYY in its filename, not by mtime — OneDrive
    rewrites mtimes on sync and would pick the wrong one."""
    best, best_key = None, None
    for f in tracker_dir.glob("*.xlsm"):
        m = _TRACKER_DATE_RE.search(f.name)
        if not m:
            continue
        d = m.group(1)
        key = d[4:6] + d[0:2] + d[2:4]          # YYMMDD
        if best_key is None or key > best_key:
            best, best_key = f, key
    return best


def _header_map(ws, header_row: int, max_col: int) -> dict:
    out = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            out[str(v).replace("\n", " ").strip()] = c
    return out


def read_tracker(path: Path) -> dict:
    """{job_code: {...}} from the DATA and POTracker tables."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: dict[str, dict] = {}
    try:
        ws = wb["Command Center"]                       # DATA table, headers row 3
        hdr = _header_map(ws, 3, 99)
        want = {
            "buid": "Builder Job Code",
            "plan_abbr": "House Plan Abbr",
            "swing": "Swing",
            "po_total": "Actual PO Amount",
            "install_pay": "Labor Total",
        }
        idx = {k: hdr.get(v) for k, v in want.items()}
        for row in ws.iter_rows(min_row=4, max_col=99, values_only=True):
            code = row[0]
            if not code or not isinstance(code, str) or "-" not in code:
                continue
            rec = {}
            for key, col in idx.items():
                if col:
                    rec[key] = row[col - 1]
            rows[code.strip().upper()] = rec

        ws = wb["PO Tracking"]                           # POTracker table, headers row 2
        hdr = _header_map(ws, 2, 26)
        po_col = hdr.get("Our PO #")
        for row in ws.iter_rows(min_row=3, max_col=26, values_only=True):
            code = row[0]
            if not code or not isinstance(code, str) or "-" not in code:
                continue
            if po_col and row[po_col - 1]:
                rows.setdefault(code.strip().upper(), {})["carter_po_number"] = row[po_col - 1]
    finally:
        wb.close()
    return rows


# ---------------------------------------------------------------------------
# Source 2: the filed job folders
# ---------------------------------------------------------------------------
def _mmddyy_to_date(token: str) -> date | None:
    if len(token) != 6 or not token.isdigit():
        return None
    try:
        return date(2000 + int(token[4:6]), int(token[0:2]), int(token[2:4]))
    except ValueError:
        return None


def scan_sold_tree(root: Path) -> dict:
    """{job_code: {...}} for every job filed under DR Horton - All.

    Only this tree is ever read. The old Sold Jobs\\Builders\\DR Horton tree is
    deprecated and off limits.
    """
    found: dict[str, dict] = {}
    if not root.is_dir():
        return found

    for region in sorted(p for p in root.iterdir() if p.is_dir()):
        for community in sorted(p for p in region.iterdir() if p.is_dir()):
            sub_m = _SUB_RE.search(community.name)
            sub = sub_m.group(1) if sub_m else None
            comm_name = _SUB_RE.sub("", community.name).strip()
            for folder in sorted(p for p in community.iterdir() if p.is_dir()):
                if folder.name.lower().startswith(_NOT_A_JOB):
                    continue
                m = _JOBCODE_RE.match(folder.name)
                if not m:
                    continue
                try:
                    files = sorted(f.name for f in folder.iterdir() if f.is_file())
                except OSError:
                    continue

                rec = {
                    "folder_name": folder.name,
                    "folder_files": files,
                    "sub_number": sub,
                    "community": comm_name,
                    "region": region.name,
                    "installer_pay_sheet": any("installer pay sheet" in f.lower() for f in files),
                    "current_folder": "sold",
                }
                # The filed Carter PO is named "{Job} {Abbr} {SO#} {CarterPO#} {MMDDYY}.pdf",
                # which carries the SO, the Carter PO, and the date it was filed.
                for f in files:
                    po = _CARTER_PO_RE.search(f)
                    if po:
                        rec["carter_po_number"] = po.group(1)
                        so = _SO_RE.search(f)
                        if so:
                            rec["so_number"] = so.group(1).upper()
                        stem = f.rsplit(".", 1)[0].split()
                        filed = _mmddyy_to_date(stem[-1]) if stem else None
                        if filed:
                            rec["moved_to_sold_date"] = filed.isoformat()
                        break
                if "so_number" not in rec:      # fall back to the plain SO document
                    for f in files:
                        so = _SO_RE.search(f)
                        if so and not _CARTER_PO_RE.search(f):
                            rec["so_number"] = so.group(1).upper()
                            break
                found[m.group(1).upper()] = rec
    return found


# ---------------------------------------------------------------------------
# Merge
# ---------------------------------------------------------------------------
_NUMERIC = ("po_total", "install_pay")
# Excel hands these back as ints (BUID 268440027, Carter PO 750005026) but they
# are identifiers, not quantities — the columns are text and never arithmetic.
_AS_TEXT = ("buid", "carter_po_number", "sub_number", "plan_abbr", "swing", "so_number")


def _clean(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def build_records(tracker: dict, sold: dict) -> list[dict]:
    """One record per job we know anything about. Folder facts win over the
    tracker where they overlap — a filed folder is physical evidence."""
    out = []
    for code in sorted(set(tracker) | set(sold)):
        rec = {"job_code": code}
        for source in (tracker.get(code, {}), sold.get(code, {})):
            for key, value in source.items():
                value = _clean(value)
                if value is None:
                    continue
                if key in _NUMERIC:
                    try:
                        value = float(value)
                    except (TypeError, ValueError):
                        continue
                    if value <= 0:
                        continue
                elif key in _AS_TEXT:
                    value = str(int(value)) if isinstance(value, float) and value.is_integer() \
                        else str(value)
                    value = value.strip()
                    if not value or value == "0":
                        continue
                rec[key] = value
        # A row with nothing but a job code is not worth sending.
        if len(rec) > 1:
            out.append(rec)
    return out


def summarize(records: list[dict]) -> dict:
    fields = ("buid", "plan_abbr", "swing", "po_total", "install_pay",
              "carter_po_number", "sub_number", "so_number", "folder_name",
              "moved_to_sold_date", "installer_pay_sheet")
    return {f: sum(1 for r in records if r.get(f) not in (None, "")) for f in fields}


def main() -> None:
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import orderpack_agent as agent

    parser = argparse.ArgumentParser(description="Backfill the Order Pack board from history")
    parser.add_argument("--dry-run", action="store_true",
                        help="print what would be sent and change nothing")
    parser.add_argument("--tracker-dir", default=os.environ.get("TRACKER_DIR", ""))
    parser.add_argument("--sold-dir", default=os.environ.get("SOLD_FILES_DIR", ""))
    parser.add_argument("--limit", type=int, default=0, help="only send the first N records")
    args = parser.parse_args()

    tracker_dir = Path(args.tracker_dir) if args.tracker_dir else Path(
        r"C:\Users\Brian SE6\OneDrive - carterlumber.com"
        r"\Townsend Kitchen and Bath - Master Plans & Pricing\Trackers"
        r"\3.0 Online Sales Tracker 010726 Backup")
    sold_dir = Path(args.sold_dir) if args.sold_dir else Path(
        r"C:\Users\Brian SE6\OneDrive - carterlumber.com"
        r"\Townsend Kitchen and Bath - Master Plans & Pricing\Sold Job Files"
        r"\National Accounts\DR Horton - All")

    book = newest_tracker(tracker_dir)
    if book is None:
        print(f"No tracker .xlsm found in {tracker_dir}")
        raise SystemExit(1)
    print(f"tracker : {book.name}")
    tracker = read_tracker(book)
    print(f"          {len(tracker)} job codes")

    print(f"sold    : {sold_dir}")
    sold = scan_sold_tree(sold_dir)
    print(f"          {len(sold)} filed job folders")

    records = build_records(tracker, sold)
    if args.limit:
        records = records[:args.limit]
    print(f"\n{len(records)} records to send. Fields populated:")
    for field, n in summarize(records).items():
        print(f"  {field:22} {n}")

    sample = [r for r in records if r.get("folder_name")][:3]
    if sample:
        print("\nsample:")
        for r in sample:
            bits = {k: v for k, v in r.items() if k != "folder_files"}
            print(f"  {bits}")

    if args.dry_run:
        print("\nDRY RUN - nothing was sent.")
        return

    result = agent.call("/agent/backfill/apply", {"records": records})
    print(f"\napplied: {result}")


if __name__ == "__main__":
    main()
